#!/usr/bin/env python3
"""
Script to optimize EUDAMED Excel template with new fields and enhanced validation
"""
import sys
import os

# Add lib directory to path
lib_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib')
if lib_path not in sys.path:
    sys.path.insert(0, lib_path)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

def add_field_comment(cell, title, description, example="", field_id="", mandatory=""):
    """Add a detailed comment to a field header"""
    comment_text = f"{'='*50}\n"
    comment_text += f"{title}\n"
    comment_text += f"{'='*50}\n\n"
    
    if field_id:
        comment_text += f"官方字段ID: {field_id}\n"
    if mandatory:
        comment_text += f"强制性: {mandatory}\n"
    comment_text += f"\n说明:\n{description}\n"
    if example:
        comment_text += f"\n示例: {example}\n"
    comment_text += f"\n{'='*50}"
    
    comment = Comment(comment_text, "EUDAMED Tool")
    comment.width = 300
    comment.height = 200
    cell.comment = comment

def add_dropdown_validation(ws, col_letter, start_row, enum_sheet, enum_col='B', error_msg="请从下拉列表中选择"):
    """Add dropdown validation from enum sheet"""
    # Find the last row with data in enum sheet
    wb = ws.parent
    if enum_sheet not in wb.sheetnames:
        print(f"Warning: Enum sheet '{enum_sheet}' not found")
        return
    
    enum_ws = wb[enum_sheet]
    last_row = enum_ws.max_row
    
    # Create validation formula
    formula = f"={enum_sheet}!${enum_col}$2:${enum_col}${last_row}"
    
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = error_msg
    dv.errorTitle = "无效输入"
    dv.prompt = "请从下拉列表中选择有效值"
    dv.promptTitle = "选择值"
    
    # Apply to range
    dv.add(f"{col_letter}{start_row}:{col_letter}1000")
    ws.add_data_validation(dv)

def optimize_basic_udi_sheet(wb):
    """Optimize Basic UDI-DI worksheet"""
    print("\nOptimizing Basic UDI-DI sheet...")
    ws = wb['Basic UDI-DI']
    
    # Get current column count
    current_cols = ws.max_column
    next_col = current_cols + 1
    
    # Define new fields to add
    new_fields = [
        {
            'header': 'Device Model',
            'comment': {
                'title': '设备型号 (Device Model)',
                'description': '设备的型号标识。与Device Name二选一必填，两者都可以提供。',
                'example': 'CardioMonitor Pro X200',
                'field_id': 'FLD-UDID-03',
                'mandatory': '与Device Name二选一必填'
            },
            'validation': None
        },
        {
            'header': 'Is it a Kit *',
            'comment': {
                'title': '是否为套装 (Is it a Kit)',
                'description': '指示该设备是否为套装（Kit）。套装是指包含多个独立设备或组件的组合产品。',
                'example': 'TRUE 或 FALSE',
                'field_id': 'FLD-UDID-06',
                'mandatory': '必填'
            },
            'validation': {'type': 'list', 'formula': '"TRUE,FALSE"', 'error': '请输入TRUE或FALSE（大写）'}
        },
        {
            'header': 'Authorised Representative SRN',
            'comment': {
                'title': '授权代表SRN (Authorised Representative SRN)',
                'description': '授权代表的单一注册号（SRN）。对于非欧盟制造商，此字段为条件必填。',
                'example': 'DE-AR-000012345',
                'field_id': 'FLD-UDID-08',
                'mandatory': '非欧盟制造商必填'
            },
            'validation': None
        },
        {
            'header': 'Special Device Type',
            'comment': {
                'title': '特殊设备类型 (Special Device Type)',
                'description': '如果设备属于特殊类别，请从下拉列表中选择相应的特殊设备类型。',
                'example': 'CUSTOM_MADE',
                'field_id': 'FLD-UDID-07',
                'mandatory': '可选'
            },
            'validation': {'type': 'enum', 'sheet': 'Enum_DeviceType', 'col': 'B'}
        },
        {
            'header': 'Reagent',
            'comment': {
                'title': '试剂 (Reagent)',
                'description': '指示该设备是否为试剂（适用于IVDR设备）。',
                'example': 'TRUE 或 FALSE',
                'field_id': 'FLD-UDID-21',
                'mandatory': '必填'
            },
            'validation': {'type': 'list', 'formula': '"TRUE,FALSE"', 'error': '请输入TRUE或FALSE（大写）'}
        },
        {
            'header': 'Presence of Medicinal Substance',
            'comment': {
                'title': '含药物物质 (Presence of Medicinal Substance)',
                'description': '指示设备是否含有如果单独使用可被视为药物的物质。',
                'example': 'TRUE 或 FALSE',
                'field_id': 'FLD-UDID-16',
                'mandatory': '必填'
            },
            'validation': {'type': 'list', 'formula': '"TRUE,FALSE"', 'error': '请输入TRUE或FALSE（大写）'}
        },
        {
            'header': 'Is Suture/Staple/Filling/Brace (IIb Implant)',
            'comment': {
                'title': 'IIb类植入物特殊声明',
                'description': '对于风险等级为Class IIb且为植入物的设备，需要声明是否为缝合线、钉、牙填充物或牙套。',
                'example': 'TRUE 或 FALSE',
                'field_id': 'FLD-UDID-XX',
                'mandatory': 'IIb类植入物条件必填'
            },
            'validation': {'type': 'list', 'formula': '"TRUE,FALSE"', 'error': '请输入TRUE或FALSE（大写）'}
        },
        {
            'header': 'Certificate Number',
            'comment': {
                'title': '证书编号 (Certificate Number)',
                'description': '与该设备关联的CE证书编号（如适用）。',
                'example': 'CE-12345-2024',
                'field_id': 'FLD-UDID-40',
                'mandatory': '可选'
            },
            'validation': None
        }
    ]
    
    # Add new fields
    for i, field in enumerate(new_fields):
        col_idx = next_col + i
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # Set header
        header_cell = ws.cell(1, col_idx, field['header'])
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add comment
        if field['comment']:
            add_field_comment(header_cell, **field['comment'])
        
        # Add validation
        if field['validation']:
            if field['validation']['type'] == 'list':
                dv = DataValidation(type="list", formula1=field['validation']['formula'], allow_blank=True)
                dv.error = field['validation']['error']
                dv.errorTitle = "无效输入"
                dv.add(f"{col_letter}3:{col_letter}1000")
                ws.add_data_validation(dv)
            elif field['validation']['type'] == 'enum':
                add_dropdown_validation(ws, col_letter, 3, field['validation']['sheet'], field['validation']['col'])
        
        # Set column width
        ws.column_dimensions[col_letter].width = 20
        
        print(f"  ✓ Added field: {field['header']}")
    
    # Update existing field validations
    print("\n  Updating existing field validations...")
    
    # Issuing Entity (Column B)
    add_dropdown_validation(ws, 'B', 3, 'Enum_IssuingEntity', 'B')
    
    # Risk Class (Column D)
    add_dropdown_validation(ws, 'D', 3, 'Enum_RiskClass', 'B')
    
    # Applicable Legislation (Column E)
    add_dropdown_validation(ws, 'E', 3, 'Enum_Legislation', 'B')
    
    # Device Type (Column F)
    add_dropdown_validation(ws, 'F', 3, 'Enum_DeviceType', 'B')
    
    print("  ✓ Updated dropdown validations for existing fields")

def optimize_udi_di_sheet(wb):
    """Optimize UDI-DI worksheet"""
    print("\nOptimizing UDI-DI sheet...")
    ws = wb['UDI-DI']
    
    # Get current column count
    current_cols = ws.max_column
    next_col = current_cols + 1
    
    # Define new fields
    new_fields = [
        {
            'header': 'Nomenclature Code *',
            'comment': {
                'title': '命名代码 (Nomenclature Code)',
                'description': '设备的国际命名代码，通常为GMDN代码。此字段为必填，可以提供多个代码。',
                'example': '12345',
                'field_id': 'FLD-UDID-60',
                'mandatory': '必填（可多个）'
            },
            'validation': None
        },
        {
            'header': 'Nomenclature System',
            'comment': {
                'title': '命名系统 (Nomenclature System)',
                'description': '命名代码所属的系统，如GMDN、UMDNS等。',
                'example': 'GMDN',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '与Nomenclature Code配套'
            },
            'validation': {'type': 'list', 'formula': '"GMDN,UMDNS,OTHER"', 'error': '请选择命名系统'}
        },
        {
            'header': 'Clinical Size Value',
            'comment': {
                'title': '临床尺寸值 (Clinical Size Value)',
                'description': '设备的临床尺寸数值（如长度、直径、容量等）。',
                'example': '10.5',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '可选'
            },
            'validation': None
        },
        {
            'header': 'Clinical Size Unit',
            'comment': {
                'title': '临床尺寸单位 (Clinical Size Unit)',
                'description': '临床尺寸的测量单位。',
                'example': 'millimetre (mm)',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '如提供尺寸值则必填'
            },
            'validation': {'type': 'list', 'formula': '"millimetre (mm),centimetre (cm),metre (m),gram (g),kilogram (kg),millilitre (mL),litre (L)"', 'error': '请选择单位'}
        },
        {
            'header': 'Additional Description',
            'comment': {
                'title': '附加描述 (Additional Description)',
                'description': '产品的附加描述信息，可以提供多语言版本。',
                'example': 'Advanced cardiac monitoring device',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '可选'
            },
            'validation': None
        },
        {
            'header': 'Description Language',
            'comment': {
                'title': '描述语言 (Description Language)',
                'description': '附加描述所使用的语言代码（ISO 639-1）。',
                'example': 'en',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '如提供描述则必填'
            },
            'validation': {'type': 'enum', 'sheet': 'Enum_LanguageCodes', 'col': 'A'}
        },
        {
            'header': 'Public Website',
            'comment': {
                'title': '公共网站 (Public Website)',
                'description': '产品的公开网站URL。',
                'example': 'https://www.example.com/product',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '可选'
            },
            'validation': None
        },
        {
            'header': 'Public Email',
            'comment': {
                'title': '公共邮箱 (Public Email)',
                'description': '产品相关咨询的公开邮箱地址。',
                'example': 'product@example.com',
                'field_id': 'FLD-UDID-XX',
                'mandatory': '可选'
            },
            'validation': None
        }
    ]
    
    # Add new fields
    for i, field in enumerate(new_fields):
        col_idx = next_col + i
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # Set header
        header_cell = ws.cell(1, col_idx, field['header'])
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add comment
        if field['comment']:
            add_field_comment(header_cell, **field['comment'])
        
        # Add validation
        if field['validation']:
            if field['validation']['type'] == 'list':
                dv = DataValidation(type="list", formula1=field['validation']['formula'], allow_blank=True)
                dv.error = field['validation']['error']
                dv.errorTitle = "无效输入"
                dv.add(f"{col_letter}3:{col_letter}1000")
                ws.add_data_validation(dv)
            elif field['validation']['type'] == 'enum':
                add_dropdown_validation(ws, col_letter, 3, field['validation']['sheet'], field['validation']['col'])
        
        # Set column width
        ws.column_dimensions[col_letter].width = 20
        
        print(f"  ✓ Added field: {field['header']}")
    
    # Update existing field validations
    print("\n  Updating existing field validations...")
    
    # UDI-DI Issuing Entity (Column C)
    add_dropdown_validation(ws, 'C', 3, 'Enum_IssuingEntity', 'B')
    
    # Device Status (Column D)
    add_dropdown_validation(ws, 'D', 3, 'Enum_DeviceStatus', 'B')
    
    print("  ✓ Updated dropdown validations for existing fields")

def optimize_market_info_sheet(wb):
    """Optimize Market Information worksheet"""
    print("\nOptimizing Market Information sheet...")
    ws = wb['Market Information']
    
    # Update Country Code validation (Column B)
    add_dropdown_validation(ws, 'B', 3, 'Enum_CountryCodes', 'A')
    print("  ✓ Updated Country Code dropdown validation")

def main():
    template_path = 'templates/EUDAMED_Template_v2.xlsx'
    
    print(f"Loading template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    
    # Optimize worksheets
    optimize_basic_udi_sheet(wb)
    optimize_udi_di_sheet(wb)
    optimize_market_info_sheet(wb)
    
    print("\nSaving optimized template...")
    wb.save(template_path)
    print(f"✓ Template optimization complete!")
    print(f"✓ Template saved: {template_path}")

if __name__ == '__main__':
    main()
