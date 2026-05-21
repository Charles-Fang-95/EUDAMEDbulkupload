#!/usr/bin/env python3
"""
修正Excel模板中的重复字段问题
删除Basic UDI-DI和UDI-DI工作表中的重复列
"""
import sys
sys.path.insert(0, '/home/ubuntu/EUDAMED_TOOL_v2/lib')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def fix_duplicate_fields(template_path):
    """修正Excel模板中的重复字段"""
    print(f"正在加载模板: {template_path}")
    wb = load_workbook(template_path)
    
    # 定义需要删除的重复列
    sheets_to_fix = {
        'Basic UDI-DI': {
            'duplicate_cols': list(range(33, 41)),  # 列33-40 (8个重复字段)
            'description': '删除列33-40的8个重复字段'
        },
        'UDI-DI': {
            'duplicate_cols': list(range(42, 50)),  # 列42-49 (8个重复字段)
            'description': '删除列42-49的8个重复字段'
        }
    }
    
    for sheet_name, config in sheets_to_fix.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  工作表 '{sheet_name}' 不存在，跳过")
            continue
        
        ws = wb[sheet_name]
        duplicate_cols = config['duplicate_cols']
        
        print(f"\n{'='*60}")
        print(f"修正工作表: {sheet_name}")
        print(f"{'='*60}")
        print(f"操作: {config['description']}")
        
        # 获取要删除的列名（用于确认）
        cols_to_delete = []
        for col_idx in duplicate_cols:
            cell_value = ws.cell(1, col_idx).value
            if cell_value:
                cols_to_delete.append(f"列{col_idx}: {cell_value}")
        
        print(f"\n将删除以下列:")
        for col_info in cols_to_delete:
            print(f"  - {col_info}")
        
        # 从后往前删除列（避免索引变化）
        for col_idx in reversed(duplicate_cols):
            ws.delete_cols(col_idx)
            print(f"✓ 已删除列 {col_idx}")
        
        # 验证修正后的列数
        max_col = ws.max_column
        print(f"\n修正后的总列数: {max_col}")
        
        # 显示修正后的字段列表
        print(f"\n修正后的字段列表:")
        for col_idx in range(1, min(max_col + 1, 35)):  # 显示前34列
            cell_value = ws.cell(1, col_idx).value
            if cell_value:
                print(f"  {col_idx:2d}. {cell_value}")
    
    # 保存修正后的模板
    output_path = template_path.replace('.xlsx', '_fixed.xlsx')
    print(f"\n{'='*60}")
    print(f"保存修正后的模板: {output_path}")
    print(f"{'='*60}")
    wb.save(output_path)
    print("✅ 模板修正完成！")
    
    return output_path

if __name__ == '__main__':
    template_path = 'templates/EUDAMED_Template_v2.xlsx'
    fixed_path = fix_duplicate_fields(template_path)
    print(f"\n修正后的模板路径: {fixed_path}")
