#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EUDAMED批量注册XML转换工具 v2.0 - 完整集成版
"""

import sys
import os

# Add lib directory to path
lib_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib')
if lib_path not in sys.path:
    sys.path.insert(0, lib_path)

import openpyxl
import argparse
from datetime import datetime

# Import custom modules
from validator import DataValidator
from xml_builder import XMLBuilder
from logger import ReportLogger


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, filepath):
        self.filepath = filepath
        self.workbook = None
        self.data = {}
        
    def load(self):
        """加载Excel文件"""
        try:
            print(f"正在加载Excel文件: {self.filepath}")
            self.workbook = openpyxl.load_workbook(self.filepath, data_only=True)
            print(f"✓ 成功加载工作簿，包含 {len(self.workbook.sheetnames)} 个工作表")
            return True
        except FileNotFoundError:
            print(f"✗ 错误：找不到文件 '{self.filepath}'")
            return False
        except Exception as e:
            print(f"✗ 加载文件时出错：{e}")
            return False
    
    def read_sheet(self, sheet_name, skip_example_row=True):
        """读取指定工作表的数据"""
        if sheet_name not in self.workbook.sheetnames:
            print(f"  警告：工作表 '{sheet_name}' 不存在")
            return []
        
        ws = self.workbook[sheet_name]
        
        # 读取表头（第1行）
        headers = []
        for cell in ws[1]:
            if cell.value:
                header = str(cell.value).strip().replace('*', '').strip()
                headers.append(header)
            else:
                break
        
        if not headers:
            print(f"  警告：工作表 '{sheet_name}' 没有表头")
            return []
        
        # 读取数据行
        data_rows = []
        start_row = 3 if skip_example_row else 2
        
        for row_idx in range(start_row, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                
                if value is not None:
                    if isinstance(value, str):
                        value = value.strip()
                        if value:
                            has_data = True
                    elif isinstance(value, (int, float)):
                        has_data = True
                    elif isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                        has_data = True
                
                row_data[header] = value if value is not None else ""
            
            if has_data:
                row_data['_row_number'] = row_idx
                data_rows.append(row_data)
        
        print(f"  ✓ 从 '{sheet_name}' 读取了 {len(data_rows)} 行数据")
        return data_rows
    
    def read_all_sheets(self):
        """读取所有数据工作表"""
        if not self.workbook:
            if not self.load():
                return False
        
        sheets_to_read = [
            'Basic UDI-DI',
            'UDI-DI',
            'Market Information',
            'Critical Warnings',
            'Storage Conditions',
            'CMR Substances',
            'Package Information'
        ]
        
        print("\n正在读取数据工作表...")
        for sheet_name in sheets_to_read:
            self.data[sheet_name] = self.read_sheet(sheet_name)
        
        total_rows = sum(len(rows) for rows in self.data.values())
        print(f"\n✓ 数据读取完成，共 {total_rows} 行数据")
        
        return True
    
    def get_data(self, sheet_name):
        """获取指定工作表的数据"""
        return self.data.get(sheet_name, [])


class EUDAMEDConverter:
    """EUDAMED转换器主类"""
    
    def __init__(self, input_file, output_file=None, validate_only=False):
        self.input_file = input_file
        self.output_file = output_file or self.generate_output_filename()
        self.validate_only = validate_only
        self.reader = ExcelReader(input_file)
        self.data = {}
        self.errors = []
        self.warnings = []
        self.logger = ReportLogger(input_file, self.output_file)
        
    def generate_output_filename(self):
        """生成输出文件名"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"eudamed_upload_{timestamp}.xml"
    
    def run(self):
        """执行转换流程"""
        print("="*70)
        print("EUDAMED批量注册XML转换工具 v2.0")
        print("="*70)
        
        # 步骤1：读取Excel数据
        print("\n[步骤 1/4] 读取Excel数据...")
        if not self.reader.read_all_sheets():
            print("\n✗ 数据读取失败，程序终止")
            return False
        
        self.data = self.reader.data
        
        # 步骤2：数据验证
        print("\n[步骤 2/4] 数据验证...")
        validator = DataValidator(self.data)
        self.errors, self.warnings = validator.validate_all()
        
        # 更新日志统计
        self.logger.set_stats(self.data, self.errors, self.warnings)
        
        # 如果有错误，停止处理
        if self.errors:
            print(f"\n✗ 发现 {len(self.errors)} 个错误，无法生成XML")
            print("  请修复错误后重试")
            
            # 生成错误报告
            self.generate_reports()
            return False
        
        if self.warnings:
            print(f"\n⚠ 发现 {len(self.warnings)} 个警告，但可以继续生成XML")
        
        # 如果仅验证模式，到此结束
        if self.validate_only:
            print("\n✓ 验证完成（仅验证模式）")
            self.generate_reports()
            return True
        
        # 步骤3：生成XML
        print("\n[步骤 3/4] 生成XML文件...")
        if not self.generate_xml():
            print("\n✗ XML生成失败")
            return False
        
        # 步骤4：生成报告
        print("\n[步骤 4/4] 生成处理报告...")
        self.logger.finish()
        self.generate_reports()
        
        print("\n" + "="*70)
        print("✓ 转换完成！")
        print(f"✓ XML文件: {self.output_file}")
        print("="*70)
        
        return True
    
    def generate_xml(self):
        """生成XML文件"""
        try:
            builder = XMLBuilder(self.data)
            builder.build()
            builder.save(self.output_file)
            
            print(f"  ✓ XML文件已生成: {self.output_file}")
            return True
            
        except Exception as e:
            print(f"  ✗ 生成XML时出错：{e}")
            import traceback
            traceback.print_exc()
            return False
    
    def generate_reports(self):
        """生成处理报告"""
        # 控制台报告
        self.logger.generate_console_report()
        
        # HTML报告
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        html_report = f"eudamed_report_{timestamp}.html"
        self.logger.generate_html_report(html_report)
        
        # JSON报告
        json_report = f"eudamed_report_{timestamp}.json"
        self.logger.generate_json_report(json_report)


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='EUDAMED批量注册XML转换工具 v2.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  python eudamed_converter_integrated.py --input EUDAMED_Template_v2.xlsx
  python eudamed_converter_integrated.py --input template.xlsx --output output.xml
  python eudamed_converter_integrated.py --input template.xlsx --validate-only
  
注意事项:
  1. 输入文件必须是优化后的Excel模板格式
  2. 所有必填字段必须填写完整
  3. 枚举值必须从下拉列表中选择
  4. 日期格式必须为YYYY-MM-DD
  5. 布尔值必须为TRUE或FALSE（大写）
        """
    )
    
    parser.add_argument('--input', '-i', required=True,
                       help='输入Excel文件路径')
    parser.add_argument('--output', '-o',
                       help='输出XML文件路径（可选，默认自动生成）')
    parser.add_argument('--validate-only', '-v', action='store_true',
                       help='仅执行数据验证，不生成XML')
    
    args = parser.parse_args()
    
    # 检查输入文件是否存在
    if not os.path.exists(args.input):
        print(f"错误：输入文件不存在: {args.input}")
        sys.exit(1)
    
    # 创建转换器并运行
    converter = EUDAMEDConverter(args.input, args.output, args.validate_only)
    success = converter.run()
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
