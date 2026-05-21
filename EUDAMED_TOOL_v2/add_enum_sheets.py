#!/usr/bin/env python3
"""
Script to add enumeration helper sheets to EUDAMED Excel template
"""
import sys
import os

# Add lib directory to path for bundled openpyxl
lib_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lib')
if lib_path not in sys.path:
    sys.path.insert(0, lib_path)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_enum_sheet(wb, sheet_name, data, headers=['Code', 'Display Name', 'Applicable To', 'Description']):
    """Create an enumeration helper sheet"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # Set header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Hide the sheet
    ws.sheet_state = 'hidden'
    
    return ws

def main():
    template_path = 'templates/EUDAMED_Template_v2.xlsx'
    
    print(f"Loading template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    
    # 1. Enum_IssuingEntity
    print("Creating Enum_IssuingEntity...")
    issuing_entities = [
        ('GS1', 'GS1', 'All', 'Global Standards 1'),
        ('HIBCC', 'HIBCC', 'All', 'Health Industry Business Communications Council'),
        ('ICCBBA', 'ICCBBA', 'All', 'International Council for Commonality in Blood Banking Automation')
    ]
    create_enum_sheet(wb, 'Enum_IssuingEntity', issuing_entities)
    
    # 2. Enum_RiskClass
    print("Creating Enum_RiskClass...")
    risk_classes = [
        ('CLASS_I', 'Class I', 'MDR', 'Low risk medical devices'),
        ('CLASS_IIA', 'Class IIa', 'MDR', 'Medium-low risk medical devices'),
        ('CLASS_IIB', 'Class IIb', 'MDR', 'Medium-high risk medical devices'),
        ('CLASS_III', 'Class III', 'MDR', 'High risk medical devices'),
        ('CLASS_A', 'Class A', 'IVDR', 'Low risk IVD devices'),
        ('CLASS_B', 'Class B', 'IVDR', 'Medium risk IVD devices'),
        ('CLASS_C', 'Class C', 'IVDR', 'High risk IVD devices'),
        ('CLASS_D', 'Class D', 'IVDR', 'Highest risk IVD devices')
    ]
    create_enum_sheet(wb, 'Enum_RiskClass', risk_classes)
    
    # 3. Enum_Legislation
    print("Creating Enum_Legislation...")
    legislations = [
        ('MDR', 'REGULATION (EU) 2017/745 on medical devices', 'Current', 'Medical Devices Regulation'),
        ('IVDR', 'REGULATION (EU) 2017/746 on in vitro diagnostic medical devices', 'Current', 'In Vitro Diagnostic Regulation'),
        ('MDD', 'Council Directive 93/42/EEC on Medical devices', 'Legacy', 'Medical Devices Directive (legacy)'),
        ('AIMDD', 'Council Directive 90/385/EEC - Active implantable medical devices', 'Legacy', 'Active Implantable Medical Devices Directive (legacy)'),
        ('IVDD', 'Directive 98/79/EC on in vitro diagnostic medical devices', 'Legacy', 'In Vitro Diagnostic Directive (legacy)')
    ]
    create_enum_sheet(wb, 'Enum_Legislation', legislations)
    
    # 4. Enum_DeviceType
    print("Creating Enum_DeviceType...")
    device_types = [
        ('REGULAR_DEVICE', 'Regular Device', 'All', 'Single medical device'),
        ('SYSTEM', 'System', 'All', 'System of multiple devices'),
        ('PROCEDURE_PACK', 'Procedure Pack', 'All', 'Procedure pack of devices')
    ]
    create_enum_sheet(wb, 'Enum_DeviceType', device_types)
    
    # 5. Enum_DeviceStatus
    print("Creating Enum_DeviceStatus...")
    device_status = [
        ('ON_THE_MARKET', 'On the EU market', 'All', 'Device is currently on the market'),
        ('NO_LONGER_PLACED_ON_THE_MARKET', 'No longer placed on the EU market', 'All', 'Device is no longer placed on market'),
        ('NOT_INTENDED_FOR_EU_MARKET', 'Not intended for the EU market', 'All', 'Device not intended for EU market')
    ]
    create_enum_sheet(wb, 'Enum_DeviceStatus', device_status)
    
    # 6. Enum_CountryCodes (EU/EEA countries)
    print("Creating Enum_CountryCodes...")
    countries = [
        ('AT', 'Austria', 'EU', 'Österreich'),
        ('BE', 'Belgium', 'EU', 'België/Belgique'),
        ('BG', 'Bulgaria', 'EU', 'България'),
        ('HR', 'Croatia', 'EU', 'Hrvatska'),
        ('CY', 'Cyprus', 'EU', 'Κύπρος'),
        ('CZ', 'Czech Republic', 'EU', 'Česká republika'),
        ('DK', 'Denmark', 'EU', 'Danmark'),
        ('EE', 'Estonia', 'EU', 'Eesti'),
        ('FI', 'Finland', 'EU', 'Suomi'),
        ('FR', 'France', 'EU', 'France'),
        ('DE', 'Germany', 'EU', 'Deutschland'),
        ('GR', 'Greece', 'EU', 'Ελλάδα'),
        ('HU', 'Hungary', 'EU', 'Magyarország'),
        ('IE', 'Ireland', 'EU', 'Éire'),
        ('IT', 'Italy', 'EU', 'Italia'),
        ('LV', 'Latvia', 'EU', 'Latvija'),
        ('LT', 'Lithuania', 'EU', 'Lietuva'),
        ('LU', 'Luxembourg', 'EU', 'Luxembourg'),
        ('MT', 'Malta', 'EU', 'Malta'),
        ('NL', 'Netherlands', 'EU', 'Nederland'),
        ('PL', 'Poland', 'EU', 'Polska'),
        ('PT', 'Portugal', 'EU', 'Portugal'),
        ('RO', 'Romania', 'EU', 'România'),
        ('SK', 'Slovakia', 'EU', 'Slovensko'),
        ('SI', 'Slovenia', 'EU', 'Slovenija'),
        ('ES', 'Spain', 'EU', 'España'),
        ('SE', 'Sweden', 'EU', 'Sverige'),
        ('IS', 'Iceland', 'EEA', 'Ísland'),
        ('LI', 'Liechtenstein', 'EEA', 'Liechtenstein'),
        ('NO', 'Norway', 'EEA', 'Norge')
    ]
    create_enum_sheet(wb, 'Enum_CountryCodes', countries)
    
    # 7. Enum_LanguageCodes
    print("Creating Enum_LanguageCodes...")
    languages = [
        ('bg', 'Bulgarian', 'Bulgaria', '保加利亚语'),
        ('cs', 'Czech', 'Czech Republic', '捷克语'),
        ('da', 'Danish', 'Denmark', '丹麦语'),
        ('de', 'German', 'Germany, Austria', '德语'),
        ('el', 'Greek', 'Greece, Cyprus', '希腊语'),
        ('en', 'English', 'Ireland, Malta', '英语'),
        ('es', 'Spanish', 'Spain', '西班牙语'),
        ('et', 'Estonian', 'Estonia', '爱沙尼亚语'),
        ('fi', 'Finnish', 'Finland', '芬兰语'),
        ('fr', 'French', 'France, Belgium, Luxembourg', '法语'),
        ('hr', 'Croatian', 'Croatia', '克罗地亚语'),
        ('hu', 'Hungarian', 'Hungary', '匈牙利语'),
        ('it', 'Italian', 'Italy', '意大利语'),
        ('lt', 'Lithuanian', 'Lithuania', '立陶宛语'),
        ('lv', 'Latvian', 'Latvia', '拉脱维亚语'),
        ('mt', 'Maltese', 'Malta', '马耳他语'),
        ('nl', 'Dutch', 'Netherlands, Belgium', '荷兰语'),
        ('pl', 'Polish', 'Poland', '波兰语'),
        ('pt', 'Portuguese', 'Portugal', '葡萄牙语'),
        ('ro', 'Romanian', 'Romania', '罗马尼亚语'),
        ('sk', 'Slovak', 'Slovakia', '斯洛伐克语'),
        ('sl', 'Slovenian', 'Slovenia', '斯洛文尼亚语'),
        ('sv', 'Swedish', 'Sweden', '瑞典语')
    ]
    create_enum_sheet(wb, 'Enum_LanguageCodes', languages)
    
    # 8. Enum_CMRTypes
    print("Creating Enum_CMRTypes...")
    cmr_types = [
        ('CMR_1A', '1A', 'All', 'CMR substance category 1A'),
        ('CMR_1B', '1B', 'All', 'CMR substance category 1B')
    ]
    create_enum_sheet(wb, 'Enum_CMRTypes', cmr_types)
    
    print("\nSaving updated template...")
    wb.save(template_path)
    print(f"✓ Successfully added {len([s for s in wb.sheetnames if s.startswith('Enum_')])} enumeration sheets")
    print(f"✓ Template saved: {template_path}")

if __name__ == '__main__':
    main()
