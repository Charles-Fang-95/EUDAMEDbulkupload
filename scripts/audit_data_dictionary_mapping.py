#!/usr/bin/env python3
"""Generate a template/importer/exporter mapping audit from the EUDAMED data dictionary.

The report is a review aid, not a code generator. It deliberately marks complex
or uncertain mappings as needs_design instead of inventing XML paths.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path


ROOT_DIR = Path(__file__).resolve().parent.parent
VENDOR_LIB = ROOT_DIR / "EUDAMED_TOOL_v2" / "lib"
if VENDOR_LIB.exists():
    sys.path.insert(0, str(VENDOR_LIB))
sys.path.insert(0, str(ROOT_DIR))

from openpyxl import load_workbook  # noqa: E402

from local_beta import template_schema  # noqa: E402
from local_beta.constants import BASIC_FIELDS, UDI_FIELDS  # noqa: E402


TARGET_SHEETS = [
    "DD BASIC UDI",
    "DD UDI-DI",
    "DD Legacy Devices",
    "DD BASIC UDI_SPP",
    "DD UDI-DI_SPP",
    "DD Container Pack",
    "DD AR related",
]

DATA_DICTIONARY_PATH = ROOT_DIR / "official_docs" / "UDI_Devices_data_dictionary.xlsx"
DEFAULT_OUTPUT = ROOT_DIR / "docs" / "DATA_DICTIONARY_FIELD_AUDIT.md"

MANUAL_MAPPINGS = {
    "applicable regulation": {
        "template": "Basic - Applicable Legislation*",
        "status": "implemented",
        "xml_path": "payload profile / applicableLegislation",
        "notes": "Template uses the label Applicable Legislation.",
    },
    "device name": {
        "template": "Basic - Device Name*",
        "status": "implemented",
        "xml_path": "deviceName",
        "notes": "Template field is Basic - Device Name* / internal field Device Name/Model.",
    },
    "authorised representative": {
        "template": "Basic - Authorised Representative SRN",
        "status": "implemented",
        "xml_path": "basicudi:ARActorCode",
        "notes": "Template stores SRN only.",
    },
    "companion diagnostic": {
        "template": "Basic - Companion Diagnostic (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:companionDiagnostics",
        "notes": "IVDR field.",
    },
    "near patient testing": {
        "template": "Basic - Near Patient Testing (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:nearPatientTesting",
        "notes": "IVDR field.",
    },
    "patient self testing": {
        "template": "Basic - Self-Testing (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:selfTesting",
        "notes": "IVDR field.",
    },
    "professional testing": {
        "template": "Basic - Professional Testing (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:professionalTesting",
        "notes": "IVDR field.",
    },
    "reusable surgical instruments": {
        "template": "Basic - Reusable Surgical Instrument",
        "status": "implemented",
        "xml_path": "commondi:reusable",
        "notes": "Template uses singular wording.",
    },
    "instrument": {
        "template": "Basic - Instrument (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:instrument",
        "notes": "IVDR field.",
    },
    "microbial origin": {
        "template": "Basic - Microbial Origin (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:microbialSubstances",
        "notes": "IVDR field.",
    },
    "presence of cells or substances of microbial origin": {
        "template": "Basic - Microbial Origin (IVDR)",
        "status": "implemented",
        "xml_path": "commondi:microbialSubstances",
        "notes": "IVDR field.",
    },
    "suture": {
        "template": "Basic - Is Suture/Staple/Filling/Brace (IIb Implant)",
        "status": "implemented",
        "xml_path": "basicudi:IIb_implantable_exceptions",
        "notes": "Conditional MDR/MDD field.",
    },
    "administer and/or remove medicinal product": {
        "template": "Basic - Administer Medicine",
        "status": "implemented",
        "xml_path": "commondi:administeringMedicine",
        "notes": "Template label is Administer Medicine.",
    },
    "eifu": {
        "template": "UDI - eIFU URL",
        "status": "collected_not_exported",
        "xml_path": "",
        "notes": "Template collects eIFU URL, but exporter does not output it yet. Official XML path requires design confirmation.",
    },
    "instructions for use": {
        "template": "UDI - eIFU URL",
        "status": "collected_not_exported",
        "xml_path": "",
        "notes": "Potential eIFU/IFU field. Collected in template, not exported until official mapping is reviewed.",
    },
    "public email": {
        "template": "UDI - Public Email",
        "status": "collected_not_exported",
        "xml_path": "",
        "notes": "Public Website is exported; Public Email is collected but not exported.",
    },
    "email": {
        "template": "UDI - Public Email",
        "status": "collected_not_exported",
        "xml_path": "",
        "notes": "Email-related UDI field requires mapping review before XML output.",
    },
    "certificate status": {
        "template": "",
        "status": "not_in_template",
        "xml_path": "",
        "notes": "Device certificate link output does not include certificate status.",
    },
    "decision date": {
        "template": "",
        "status": "not_in_template",
        "xml_path": "",
        "notes": "Certificate decision metadata is not part of current deviceCertificateLinks output.",
    },
    "issue date": {
        "template": "",
        "status": "not_in_template",
        "xml_path": "",
        "notes": "Issue date is not part of current deviceCertificateLinks output.",
    },
    "starting validity date": {
        "template": "",
        "status": "not_in_template",
        "xml_path": "",
        "notes": "Starting validity date is not part of current deviceCertificateLinks output.",
    },
    "notified body": {
        "template": "Device Certificates / Notified Body ID",
        "status": "implemented",
        "xml_path": "basicudi:deviceCertificateLinks/links:deviceCertificateLink/links:NBActorCode",
        "notes": "Stored as Notified Body ID / NBActorCode.",
    },
    "expiry date": {
        "template": "Device Certificates / Expiry Date",
        "status": "implemented",
        "xml_path": "basicudi:deviceCertificateLinks/links:deviceCertificateLink/links:expiryDate",
        "notes": "Optional for regulation devices; often required for legacy directive certificates.",
    },
    "revision number": {
        "template": "Device Certificates / Revision Number",
        "status": "implemented",
        "xml_path": "basicudi:deviceCertificateLinks/links:deviceCertificateLink/links:certificateRevisionNumber",
        "notes": "Optional revision number.",
    },
    "certificate": {
        "template": "Device Certificates",
        "status": "implemented",
        "xml_path": "basicudi:deviceCertificateLinks/links:deviceCertificateLink",
        "notes": "Structured certificate rows are exported as deviceCertificateLinks. PR/SPP certificate handling remains out of scope.",
    },
    "is it a kit": {
        "template": "Basic - Is it a Kit",
        "status": "implemented",
        "xml_path": "commondi:kit",
        "notes": "Unified template field. Exported for IVDR/IVDD paths where current XSD provides commondi:kit; MDR/MDD are not forced into XML without a safe schema location.",
    },
    "kit": {
        "template": "Basic - Is it a Kit",
        "status": "implemented",
        "xml_path": "commondi:kit",
        "notes": "Old Basic - Kit (IVDR) is migrated into the unified Is it a Kit field.",
    },
    "presence of medicinal substance": {
        "template": "Basic - Presence of Medicinal Substance",
        "status": "collected_not_exported",
        "xml_path": "",
        "notes": "Collected separately, but current exporter outputs Basic - Medicinal Product Device to medicinalProductCheck.",
    },
    "medicinal product": {
        "template": "Basic - Medicinal Product Device",
        "status": "implemented",
        "xml_path": "basicudi:medicinalProductCheck",
        "notes": "Current exporter maps Medicinal Product Device to medicinalProductCheck.",
    },
    "clinical size": {
        "template": "Clinical Sizes sheet",
        "status": "implemented",
        "xml_path": "udidi:clinicalSizes/commondi:clinicalSize",
        "notes": "Structured Clinical Sizes sheet is exported for MDR UDI-DI only; other profiles are warned and ignored.",
    },
    "product designer": {
        "template": "UDI - Product Designer SRN / UDI - Product Designer ID",
        "status": "explicitly_out_of_scope",
        "xml_path": "",
        "notes": "Product original manufacturer/designer update service is not implemented.",
    },
    "purpose other than medical": {
        "template": "Annex XVI Purposes sheet",
        "status": "implemented",
        "xml_path": "udidi:annexXVINonMedicalDeviceTypes/udidi:nmdType",
        "notes": "Annex XVI non-medical device types are collected as 0..n rows and exported for MDR UDI-DI only.",
    },
    "measure unit description": {
        "template": "Clinical Sizes / Measure Unit Description",
        "status": "implemented",
        "xml_path": "udidi:clinicalSizes/commondi:clinicalSize/commondi:measureUnitDescription",
        "notes": "Required when Clinical Size Measure Unit is MU999 - OTHER.",
    },
}

XML_PATH_OVERRIDES = {
    "Basic UDI-DI Code": "basicudi:basicUDIIdentifier/basicUDIIdentifier",
    "Issuing Entity": "basicudi:basicUDIIdentifier/issuingEntity",
    "Manufacturer SRN": "manufacturerActorCode",
    "Risk Class": "riskClass",
    "Applicable Legislation": "payload entity selection",
    "Device Name/Model": "deviceName",
    "EMDN Code": "nomenclatureCode",
    "UDI-DI Code": "udidi:udiIdentifier/diCode",
    "UDI-DI Issuing Entity": "udidi:udiIdentifier/issuingEntity",
    "Device Status": "deviceStatus",
    "Reference Number": "referenceNumber",
    "Trade Name": "tradeNames",
    "Public Website": "udidi:website",
    "Additional Description": "additionalDescription",
}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dictionary", default=str(DATA_DICTIONARY_PATH))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    rows = read_dictionary(Path(args.dictionary))
    template_index = build_template_index()
    storage_fields = set(BASIC_FIELDS) | set(UDI_FIELDS)
    exporter_text = (ROOT_DIR / "local_beta" / "exporter.py").read_text(encoding="utf-8", errors="ignore")

    audited = [
        audit_row(row, template_index, storage_fields, exporter_text)
        for row in rows
    ]
    write_markdown(audited, Path(args.output))
    print(f"Wrote {args.output} ({len(audited)} fields)")
    return 0


def read_dictionary(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: list[dict] = []
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [clean(cell.value) for cell in ws[3]]
        header_map = {header: idx for idx, header in enumerate(headers) if header}
        for row in ws.iter_rows(min_row=4, values_only=True):
            field_id = value_at(row, header_map, "Field ID")
            label = value_at(row, header_map, "Field Label")
            field_name = value_at(row, header_map, "Field Name")
            if not field_id and not label and not field_name:
                continue
            out.append(
                {
                    "sheet": sheet_name,
                    "field_id": field_id,
                    "label": label or field_name,
                    "field_name": field_name,
                    "occurrence": value_at(row, header_map, "Occurrence"),
                    "description": value_at(row, header_map, "Field Description / Notes"),
                }
            )
    return out


def build_template_index() -> dict:
    index = {}
    for col in template_schema.ALL_COLUMNS:
        candidates = {
            col.get("field", ""),
            col.get("header", ""),
            col.get("header", "").replace("*", ""),
        }
        for candidate in candidates:
            if candidate:
                index[normalize(candidate)] = col
    return index


def audit_row(row: dict, template_index: dict, storage_fields: set[str], exporter_text: str) -> dict:
    label_text = " ".join([row.get("label", ""), row.get("field_name", "")]).lower()
    full_text = " ".join([row.get("label", ""), row.get("field_name", ""), row.get("description", "")]).lower()
    manual = manual_match(label_text, full_text)
    if manual:
        manual = context_adjusted_manual(row, manual)
        template = manual.get("template", "")
        importer_reads = bool(template)
        storage_saves = storage_state(template, storage_fields)
        exporter_outputs = manual["status"] == "implemented"
        return {
            **row,
            "template": template,
            "importer_reads": yes_no(importer_reads),
            "storage_saves": storage_saves,
            "exporter_outputs": yes_no(exporter_outputs),
            "xml_path": manual.get("xml_path", ""),
            "status": manual["status"],
            "notes": manual.get("notes", ""),
        }

    col = find_template_column(row, template_index)
    if not col:
        return {
            **row,
            "template": "",
            "importer_reads": "No",
            "storage_saves": "No",
            "exporter_outputs": "No",
            "xml_path": "",
            "status": "not_in_template",
            "notes": "",
        }

    field = col.get("field", "")
    exported = exporter_has_field(exporter_text, field)
    explicitly_out = is_explicitly_out_of_scope(col)
    if exported:
        status = "implemented"
    elif explicitly_out:
        status = "explicitly_out_of_scope"
    else:
        status = "collected_not_exported"

    return {
        **row,
        "template": col.get("header", ""),
        "importer_reads": "Yes",
        "storage_saves": storage_state(field, storage_fields),
        "exporter_outputs": yes_no(exported),
        "xml_path": XML_PATH_OVERRIDES.get(field, ""),
        "status": status,
        "notes": col.get("description", "") if status != "implemented" else "",
    }


def manual_match(label_text: str, full_text: str) -> dict | None:
    for needle, mapping in MANUAL_MAPPINGS.items():
        # Broad terms like certificate/kit/medicinal often appear in occurrence notes for unrelated fields.
        # Match those only against the official label/name, not the longer description.
        haystack = full_text if needle in {"eifu", "instructions for use"} else label_text
        if needle in haystack:
            return mapping
    return None


def context_adjusted_manual(row: dict, manual: dict) -> dict:
    """Keep broad label-based mappings honest when the official sheet changes context."""
    adjusted = dict(manual)
    xml_path = adjusted.get("xml_path", "")
    if "clinicalSizes" in xml_path and row.get("sheet") != "DD UDI-DI":
        adjusted["status"] = "explicitly_out_of_scope"
        adjusted["xml_path"] = ""
        adjusted["notes"] = (
            "Current exporter supports structured clinicalSizes only for MDR UDI-DI; "
            "legacy / other profiles are warned and ignored."
        )
    return adjusted


def find_template_column(row: dict, template_index: dict) -> dict | None:
    for candidate in [row.get("label", ""), row.get("field_name", "")]:
        norm = normalize(candidate)
        if norm in template_index:
            return template_index[norm]
    text = normalize(" ".join([row.get("label", ""), row.get("field_name", "")]))
    for norm, col in template_index.items():
        if len(norm) > 5 and norm in text:
            return col
    return None


def exporter_has_field(exporter_text: str, field: str) -> bool:
    if not field:
        return False
    patterns = [
        f'"{field}"',
        f"'{field}'",
    ]
    return any(pattern in exporter_text for pattern in patterns)


def is_explicitly_out_of_scope(col: dict) -> bool:
    description = col.get("description", "")
    return any(token in description for token in ["当前 XML 暂不输出", "当前不输出", "后续实现", "待审计", "未实现"])


def storage_state(template: str, storage_fields: set[str]) -> str:
    if not template:
        return "No"
    fields = [part.strip() for part in re.split(r"/|,|，", template)]
    for item in fields:
        item = item.replace("UDI - ", "").replace("Basic - ", "").replace("*", "").strip()
        if item in storage_fields:
            return "Yes"
    if "sheet" in template.lower() or " / " in template:
        return "JSON payload"
    return "Payload"


def yes_no(value: bool) -> str:
    return "Yes" if value else "No"


def normalize(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def clean(value) -> str:
    return " ".join(str(value or "").split())


def value_at(row: tuple, header_map: dict, header: str) -> str:
    idx = header_map.get(header)
    if idx is None or idx >= len(row):
        return ""
    return clean(row[idx])


def write_markdown(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    counts = Counter(row["status"] for row in rows)
    lines = [
        "# Data Dictionary Field Mapping Audit",
        "",
        "This report compares the official EUDAMED UDI Devices data dictionary with the current Excel template, importer/storage, and XML exporter.",
        "",
        "Status meanings:",
        "",
        "- `implemented`: collected and currently output to XML, or mapped by known exporter logic.",
        "- `collected_not_exported`: template/importer collects the field but exporter does not currently output it.",
        "- `not_in_template`: official field is not represented in the current template.",
        "- `explicitly_out_of_scope`: template deliberately marks the field as not currently output or tied to a later service.",
        "- `needs_design`: field needs a dedicated mapping design before safe XML output.",
        "",
        "## Summary",
        "",
    ]
    for status in ["implemented", "collected_not_exported", "not_in_template", "explicitly_out_of_scope", "needs_design"]:
        lines.append(f"- `{status}`: {counts.get(status, 0)}")
    lines += [
        "",
        "## Known Priority Findings",
        "",
        "- `eIFU URL` and `Public Email` are collected or partly represented, but not safely output to XML yet.",
        "- `Device Certificates` is implemented for Basic UDI-DI `deviceCertificateLinks`; PR/SPP certificate handling remains out of scope.",
        "- `Clinical Sizes` and `Annex XVI Purposes` are implemented for MDR UDI-DI via structured detail sheets.",
        "- `Is it a Kit` is unified in the template and exported where the current XSD provides `commondi:kit` (IVDR/IVDD paths).",
        "- `Product Designer` remains out of scope until the Update product original manufacturer service is designed.",
        "- `Presence of Medicinal Substance` remains documented-not-exported because `Medicinal Product Device` already maps to `medicinalProductCheck`.",
        "",
        "## Field Audit",
        "",
        "| Source Sheet | Field ID | Field Label | Occurrence | Template Field | Importer Reads | Storage Saves | Exporter Outputs | XML Path | Status | Notes |",
        "|---|---|---|---|---|---:|---:|---:|---|---|---|",
    ]
    for row in rows:
        lines.append(
            "| {sheet} | {field_id} | {label} | {occurrence} | {template} | {importer} | {storage} | {exporter} | {xml} | `{status}` | {notes} |".format(
                sheet=md(row.get("sheet", "")),
                field_id=md(row.get("field_id", "")),
                label=md(row.get("label", "")),
                occurrence=md(row.get("occurrence", "")),
                template=md(row.get("template", "")),
                importer=md(row.get("importer_reads", "")),
                storage=md(row.get("storage_saves", "")),
                exporter=md(row.get("exporter_outputs", "")),
                xml=md(row.get("xml_path", "")),
                status=md(row.get("status", "")),
                notes=md(row.get("notes", "")),
            )
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def md(value: str) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ")


if __name__ == "__main__":
    sys.exit(main())
