import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from .constants import EXPORT_DIR, TOOL_DIR, VENDOR_LIB
from .legacy_identifiers import (
    LegacyIdentifierError,
    METHOD_EXISTING_EUDAMED_PAIR,
    legacy_fields_present,
    resolve_legacy_identifiers,
)
from .storage import Repository
from .template_schema import (
    ENTRY_SHEETS,
    IMPORT_ENTRY_SHEETS,
    ENUM_SOURCES,
    RELATED_SHEETS,
    TEMPLATE_VERSION,
    columns_for_entry_sheet,
)

if str(VENDOR_LIB) not in sys.path:
    sys.path.insert(0, str(VENDOR_LIB))
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import openpyxl  # type: ignore
from validator import DataValidator  # type: ignore


BUSINESS_SHEETS = [
    "Basic UDI-DI",
    "UDI-DI",
    "Trade Names",
    "Market Information",
    "Critical Warnings",
    "Storage Conditions",
    "CMR Substances",
    "Package Information",
    "Device Certificates",
    "Clinical Sizes",
    "Annex XVI Purposes",
]
DATA_START_ROW = 4
FORMAT_RISK_FIELDS = {
    "Legacy EUDAMED DI Input",
    "Legacy EUDAMED DI",
    "Legacy EUDAMED ID",
    "Basic UDI-DI Code",
    "UDI-DI Code",
    "UDI-DI Issuing Entity",
    "Package UDI-DI Code",
    "Contains DI Code",
    "Reference Number",
    "Manufacturer SRN",
    "Authorised Representative SRN",
    "Product Designer SRN",
    "DM DI Code",
    "Unit of Use DI Code",
    "Secondary UDI-DI Code",
    "EMDN Code",
    "Nomenclature Code",
    "CAS Code",
    "EC Code",
}
CODE_LENGTH_RISK_FIELDS = {
    "UDI-DI Code",
    "Package UDI-DI Code",
    "Contains DI Code",
    "Secondary UDI-DI Code",
    "DM DI Code",
    "Unit of Use DI Code",
}
SCIENTIFIC_NOTATION_RE = re.compile(r"^[+-]?\d+(\.\d+)?E[+-]?\d+$", re.IGNORECASE)
PRESERVE_WHITESPACE_HEADERS = {
    "legacy - eudamed di input",
    "legacy eudamed di input",
}


class WorkbookImporter:
    def __init__(self, repository: Repository):
        self.repository = repository

    def import_workbook(self, workbook_path: Path) -> dict:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        parsed, import_meta, format_warnings, migration_warnings, identifier_errors = self._parse_workbook(wb)

        summary = self._summary(parsed)
        if not any(summary.values()):
            validation = {
                "errors": identifier_errors or [
                    {
                        "sheet": "Workbook",
                        "row": "",
                        "field": "",
                        "value": workbook_path.name,
                        "error_type": "unsupported_template",
                        "message": "未识别到可导入的 EUDAMED 模板数据。请使用当前 v2.12 的 MDR、MDD_AIMDD、IVDR 或 IVDD 主表。",
                        "suggestion": f"旧模板或客户原始清单不要直接导入；请先迁移/映射到当前 {TEMPLATE_VERSION} 模板字段，确认字段含义后再导入。",
                    }
                ],
                "warnings": [],
            }
            import_id = self.repository.create_import(workbook_path.name, summary, validation)
            return {
                "import_id": import_id,
                "summary": summary,
                "validation": validation,
                "changes": [],
                "change_summary": {"created": 0, "updated": 0, "unchanged": 0},
                "legacy_results": import_meta.get("legacy_results", []),
                "normalized_filename": "",
            }

        market_map = defaultdict(list)
        warning_map = defaultdict(list)
        storage_map = defaultdict(list)
        package_map = defaultdict(list)
        cmr_map = defaultdict(list)
        cert_map = defaultdict(list)
        trade_name_map = defaultdict(list)
        clinical_size_map = defaultdict(list)
        annex_xvi_map = defaultdict(list)

        for row in parsed.get("Trade Names", []):
            trade_name_map[row.get("UDI-DI Code", "")].append(row)
        for row in parsed.get("Market Information", []):
            market_map[row.get("UDI-DI Code", "")].append(row)
        for row in parsed.get("Critical Warnings", []):
            warning_map[row.get("UDI-DI Code", "")].append(row)
        for row in parsed.get("Storage Conditions", []):
            storage_map[row.get("UDI-DI Code", "")].append(row)
        for row in parsed.get("Package Information", []):
            package_map[row.get("UDI-DI Code", "")].append(row)
        for row in parsed.get("CMR Substances", []):
            cmr_map[row.get("Basic UDI-DI Code", "")].append(row)
        for row in parsed.get("Device Certificates", []):
            cert_map[row.get("Basic UDI-DI Code", "")].append(row)
        for row in parsed.get("Clinical Sizes", []):
            clinical_size_map[row.get("UDI-DI Code", "")].append(row)
        for row in parsed.get("Annex XVI Purposes", []):
            annex_xvi_map[row.get("UDI-DI Code", "")].append(row)

        self._merge_trade_name_shortcuts(parsed, trade_name_map)
        validator = DataValidator(parsed)
        errors, warnings = validator.validate_all()
        extra_errors = self._validate_related_rules(parsed)

        validation = {
            "errors": identifier_errors + [error.to_dict() for error in errors] + extra_errors,
            "warnings": [warning.to_dict() for warning in warnings] + format_warnings + migration_warnings,
        }
        import_id = self.repository.create_import(workbook_path.name, summary, validation)
        changes = []

        for row in parsed.get("Basic UDI-DI", []):
            basic_code = row.get("Basic UDI-DI Code", "")
            change = self.repository.upsert_basic(
                import_id=import_id,
                row_number=row.get("_row_number", 0),
                payload=self._clean_row(row),
                cmr_rows=[self._clean_row(item) for item in cmr_map.get(basic_code, [])],
                cert_rows=[self._clean_row(item) for item in cert_map.get(basic_code, [])],
                version=import_meta["basic_versions"].get(basic_code, ""),
            )
            if change:
                changes.append(change)
        for row in parsed.get("UDI-DI", []):
            code = row.get("UDI-DI Code", "")
            change = self.repository.upsert_udi(
                import_id=import_id,
                row_number=row.get("_row_number", 0),
                payload=self._clean_row(row),
                market_rows=[self._clean_row(item) for item in market_map.get(code, [])],
                warning_rows=[self._clean_row(item) for item in warning_map.get(code, [])],
                storage_rows=[self._clean_row(item) for item in storage_map.get(code, [])],
                package_rows=[self._clean_row(item) for item in package_map.get(code, [])],
                trade_name_rows=self._trade_name_rows_for_udi(row, trade_name_map.get(code, [])),
                clinical_size_rows=[self._clean_row(item) for item in clinical_size_map.get(code, [])],
                annex_xvi_rows=[self._clean_row(item) for item in annex_xvi_map.get(code, [])],
                version=import_meta["udi_versions"].get(code, ""),
            )
            if change:
                changes.append(change)

        market_warnings = self._market_change_warnings(changes)
        consistency_warnings = self._consistency_warnings(changes)
        if market_warnings or consistency_warnings:
            validation["warnings"].extend(market_warnings)
            validation["warnings"].extend(consistency_warnings)
            self.repository.update_import_validation(import_id, validation)

        normalized = None
        if not import_meta.get("normalization_blocked") and import_meta.get("normalization_updates"):
            normalized = self._write_normalized_workbook(
                workbook_path,
                import_meta["normalization_updates"],
            )

        return {
            "import_id": import_id,
            "summary": summary,
            "validation": validation,
            "changes": changes,
            "change_summary": self._change_summary(changes),
            "legacy_results": import_meta.get("legacy_results", []),
            "normalized_filename": normalized.name if normalized else "",
        }

    def _parse_workbook(self, wb) -> tuple[dict, dict, list[dict], list[dict], list[dict]]:
        parsed = {sheet: [] for sheet in BUSINESS_SHEETS}
        basic_index = {}
        import_meta = {
            "basic_versions": {},
            "udi_versions": {},
            "legacy_results": [],
            "normalization_updates": {},
            "normalization_blocked": False,
        }
        format_warnings = []
        migration_warnings = self._template_version_warnings(wb)
        identifier_errors = []
        local_record_map = {}
        duplicate_local_ids = self._duplicate_local_record_ids(wb)

        for sheet_name in IMPORT_ENTRY_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            self._parse_entry_sheet(
                wb[sheet_name],
                parsed,
                basic_index,
                import_meta,
                format_warnings,
                migration_warnings,
                identifier_errors,
                local_record_map,
                duplicate_local_ids,
            )

        self._filter_duplicate_legacy_results(
            parsed,
            import_meta,
            local_record_map,
            identifier_errors,
        )

        for sheet_name, spec in RELATED_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            rows = self._parse_related_sheet(wb[sheet_name], spec["columns"], format_warnings, migration_warnings)
            self._resolve_related_local_ids(
                sheet_name,
                spec["target"],
                rows,
                local_record_map,
                identifier_errors,
                import_meta["normalization_updates"],
            )
            parsed[spec["target"]].extend(rows)

        return parsed, import_meta, format_warnings, migration_warnings, identifier_errors

    def _filter_duplicate_legacy_results(self, parsed: dict, import_meta: dict, local_record_map: dict, errors: list[dict]):
        origins = defaultdict(list)
        for row in parsed.get("UDI-DI", []):
            final_di = str(row.get("Legacy EUDAMED DI") or "").strip()
            final_id = str(row.get("Legacy EUDAMED ID") or row.get("UDI-DI Code") or "").strip()
            if final_di:
                origins[(final_di, final_id)].append(row)
        duplicate_rows = {
            (str(row.get("_sheet_name") or ""), int(row.get("_row_number") or 0))
            for rows in origins.values() if len(rows) > 1
            for row in rows
        }
        if not duplicate_rows:
            return
        import_meta["normalization_blocked"] = True
        for (final_di, final_id), rows in origins.items():
            if len(rows) < 2:
                continue
            for row in rows:
                errors.append(self._identifier_error(
                    str(row.get("_sheet_name") or "Workbook"),
                    int(row.get("_row_number") or 0),
                    "Legacy EUDAMED DI / ID",
                    f"{final_di} / {final_id}",
                    "LEGACY_IDENTIFIER_DUPLICATE",
                    "多条主表记录计算或提供了相同的最终 Legacy 标识；重复结果对应的所有行均未入库。",
                    "为每条器械使用制造商确定的唯一主体或真实 UDI-DI，并在 EUDAMED 环境确认最终唯一性。",
                ))

        parsed["UDI-DI"] = [
            row for row in parsed.get("UDI-DI", [])
            if (str(row.get("_sheet_name") or ""), int(row.get("_row_number") or 0)) not in duplicate_rows
        ]
        valid_parent_codes = {str(row.get("Parent Basic UDI-DI") or "") for row in parsed["UDI-DI"]}
        parsed["Basic UDI-DI"] = [
            row for row in parsed.get("Basic UDI-DI", [])
            if str(row.get("Basic UDI-DI Code") or "") in valid_parent_codes
        ]
        invalid_local_ids = {
            str(row.get("Local Record ID") or "")
            for rows in origins.values() if len(rows) > 1
            for row in rows
            if row.get("Local Record ID")
        }
        for local_id in invalid_local_ids:
            local_record_map.pop(local_id, None)
        import_meta["legacy_results"] = [
            item for item in import_meta.get("legacy_results", [])
            if (str(item.get("sheet") or ""), int(item.get("row") or 0)) not in duplicate_rows
        ]
        import_meta["normalization_updates"] = {
            key: value for key, value in import_meta.get("normalization_updates", {}).items()
            if (str(key[0]), int(key[1])) not in duplicate_rows
        }

    def _duplicate_local_record_ids(self, wb) -> set[str]:
        occurrences = defaultdict(int)
        for sheet_name in IMPORT_ENTRY_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = self._headers(ws)
            local_col = next(
                (index for index, header in enumerate(headers, start=1)
                 if self._canonical_header(header) == "local - record id"),
                None,
            )
            if local_col is None:
                continue
            for row_idx in range(DATA_START_ROW, self._last_data_row(ws, headers) + 1):
                value = str(ws.cell(row_idx, local_col).value or "").strip()
                if value:
                    occurrences[value] += 1
        return {value for value, count in occurrences.items() if count > 1}

    def _identifier_error(
        self,
        sheet: str,
        row: int,
        field: str,
        value,
        error_type: str,
        message: str,
        suggestion: str,
    ) -> dict:
        return {
            "sheet": sheet,
            "row": row,
            "field": field,
            "value": value,
            "error_type": error_type,
            "message": message,
            "suggestion": suggestion,
        }

    def _main_normalization_updates(
        self,
        updates: dict,
        sheet: str,
        row: int,
        resolution,
        *,
        preserve_assigned_basic: bool = False,
    ):
        values = {
            "Legacy Has Assigned UDI-DI": "TRUE" if resolution.has_assigned_udi else "FALSE",
            "Legacy EUDAMED DI Input": resolution.eudamed_di_input,
            "Legacy EUDAMED DI": resolution.eudamed_di,
            "Legacy EUDAMED ID": resolution.eudamed_id,
            "UDI-DI Code": resolution.identifier_code,
            "UDI-DI Issuing Entity": resolution.identifier_issuing_entity,
        }
        if not preserve_assigned_basic:
            values["Basic UDI-DI Code"] = resolution.eudamed_di
            values["Issuing Entity"] = "EUDAMED"
        for field, value in values.items():
            updates[(sheet, row, field)] = value

    def _resolve_related_local_ids(
        self,
        sheet_name: str,
        target: str,
        rows: list[dict],
        local_record_map: dict,
        errors: list[dict],
        normalization_updates: dict,
    ):
        basic_target = target in {"CMR Substances", "Device Certificates"}
        link_field = "Basic UDI-DI Code" if basic_target else "UDI-DI Code"
        by_formal = {}
        for local_id, mapping in local_record_map.items():
            candidates = (
                {mapping.get("basic_storage_code"), mapping.get("basic_final_code")}
                if basic_target else {mapping.get("udi_code")}
            )
            for code in candidates:
                if code:
                    by_formal[str(code)] = (local_id, mapping)

        valid_rows = []
        for row in rows:
            local_id = str(row.get("Local Record ID") or "").strip()
            formal = str(row.get(link_field) or "").strip()
            row_number = int(row.get("_row_number") or 0)
            mapping = local_record_map.get(local_id) if local_id else None
            if local_id and not mapping:
                errors.append(self._identifier_error(
                    sheet_name, row_number, "Local - Record ID", local_id,
                    "LOCAL_RECORD_ID_UNRESOLVED",
                    "关联 sheet 的 Local - Record ID 无法解析到一条有效主表记录；该关联行未入库。",
                    "确认主表键存在、唯一，且对应 Legacy 标识计算成功。",
                ))
                continue
            if not mapping and formal and formal in by_formal:
                _, mapping = by_formal[formal]
            if mapping:
                expected_formal = mapping.get("basic_final_code") if basic_target else mapping.get("udi_code")
                storage_code = mapping.get("basic_storage_code") if basic_target else mapping.get("udi_code")
                accepted = {str(code) for code in (expected_formal, storage_code) if code}
                if formal and formal not in accepted:
                    errors.append(self._identifier_error(
                        sheet_name, row_number, f"Local - Record ID / {link_field}",
                        f"{local_id} / {formal}", "LOCAL_AND_FORMAL_REFERENCE_CONFLICT",
                        "Local - Record ID 与同一行正式编码指向不同的主表记录；该关联行未入库。",
                        "保留一种关联方式，或把两者改为指向同一主表行。",
                    ))
                    continue
                row[link_field] = storage_code
                # Assigned-UDI Legacy records may intentionally use a local Basic key.
                # Keep related Basic references aligned with the stored/main-sheet key.
                normalization_updates[(sheet_name, row_number, link_field)] = storage_code
            valid_rows.append(row)
        rows[:] = valid_rows

    def _write_normalized_workbook(self, source_path: Path, updates: dict) -> Path | None:
        if not updates:
            return None
        source_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=False)
        detected = self._detect_template_version(source_wb)
        source_sheetnames = set(source_wb.sheetnames)
        source_wb.close()
        needs_migration = (
            detected != TEMPLATE_VERSION
            or bool(source_sheetnames.intersection(set(IMPORT_ENTRY_SHEETS) - set(ENTRY_SHEETS)))
        )
        row_mapping = {}
        if needs_migration:
            from .template_migrator import migrate_workbook
            with TemporaryDirectory(prefix="eudamed_normalize_") as temp_dir:
                migrated = migrate_workbook(source_path, Path(temp_dir))
                if not migrated.get("ok") or not migrated.get("output_filename"):
                    return None
                wb = openpyxl.load_workbook(Path(temp_dir) / migrated["output_filename"], data_only=False)
                row_mapping = {
                    (str(item.get("source_sheet") or ""), int(item.get("source_row") or 0)): (
                        str(item.get("target_sheet") or ""),
                        int(item.get("target_row") or 0),
                    )
                    for item in migrated.get("report", {}).get("row_mappings", [])
                }
        else:
            wb = openpyxl.load_workbook(source_path, data_only=False)
        for (sheet_name, row_number, field), value in updates.items():
            if row_mapping:
                mapped = row_mapping.get((str(sheet_name), int(row_number)))
                if not mapped:
                    continue
                sheet_name, row_number = mapped
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            columns = columns_for_entry_sheet(sheet_name) if sheet_name in ENTRY_SHEETS else RELATED_SHEETS.get(sheet_name, {}).get("columns", [])
            header_by_field = {item["field"]: item["header"] for item in columns}
            header = header_by_field.get(field)
            if not header:
                continue
            headers = self._headers(ws)
            canonical = self._canonical_header(header)
            col_idx = next(
                (index for index, candidate in enumerate(headers, start=1)
                 if self._canonical_header(candidate) == canonical),
                None,
            )
            if col_idx:
                ws.cell(int(row_number), col_idx, value)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        output = EXPORT_DIR / f"NORMALIZED_EUDAMED_Template_{TEMPLATE_VERSION}_{timestamp}.xlsx"
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return output

    def _market_change_warnings(self, changes: list[dict]) -> list[dict]:
        warnings = []
        for item in changes:
            if item.get("entity_type") != "udi" or item.get("action") != "updated":
                continue
            if "Market Info" not in item.get("changed_fields", []):
                continue
            warnings.append(
                {
                    "sheet": "Market Info",
                    "row": item.get("row_number", ""),
                    "field": "Market Information",
                    "value": item.get("code", ""),
                    "warning_type": "market_information_updated",
                    "message": "检测到该 UDI-DI 的市场信息发生变化。国家/市场信息错误应优先通过 EUDAMED update/create new version 修正，不应默认删除 UDI-DI 重建。",
                    "suggestion": "只有当 UDI-DI、器械身份或 Basic UDI-DI 关联本身错误且无法更新纠正时，才考虑 discard/逻辑删除并重新注册。",
                }
            )
        return warnings

    def _consistency_warnings(self, changes: list[dict]) -> list[dict]:
        if not changes:
            return []
        warnings = []
        only_codes = {
            "basic": [item.get("code", "") for item in changes if item.get("entity_type") == "basic"],
            "udi": [item.get("code", "") for item in changes if item.get("entity_type") == "udi"],
        }
        for finding in self.repository.consistency_findings(only_codes=only_codes):
            warnings.append(
                {
                    "sheet": "Repository",
                    "row": "",
                    "field": finding.get("type", ""),
                    "value": ", ".join(finding.get("codes") or []),
                    "warning_type": finding.get("type", "consistency"),
                    "message": finding.get("message", ""),
                    "suggestion": "这是跨记录一致性提示，不会阻止导入；请在导出前核对是否属于真实业务差异。",
                }
            )
        return warnings

    def _parse_entry_sheet(
        self,
        ws,
        parsed: dict,
        basic_index: dict,
        import_meta: dict,
        format_warnings: list[dict],
        migration_warnings: list[dict],
        identifier_errors: list[dict] | None = None,
        local_record_map: dict | None = None,
        duplicate_local_ids: set[str] | None = None,
    ):
        identifier_errors = identifier_errors if identifier_errors is not None else []
        local_record_map = local_record_map if local_record_map is not None else {}
        duplicate_local_ids = duplicate_local_ids if duplicate_local_ids is not None else set()
        import_meta.setdefault("legacy_results", [])
        import_meta.setdefault("normalization_updates", {})
        headers = self._headers(ws)
        schema_by_header = self._schema_by_header(columns_for_entry_sheet(ws.title))
        if any(header in {"UDI - eIFU URL", "eIFU URL"} for header in headers):
            migration_warnings.append(
                {
                    "sheet": ws.title,
                    "row": "",
                    "field": "UDI - eIFU URL",
                    "value": "",
                    "warning_type": "LEGACY_EIFU_URL_NOT_OUTPUT",
                    "message": "检测到旧模板字段 UDI - eIFU URL。0.9.6 模板曾明确该字段不输出到 XML；本次导入不会自动把它映射到官方 URL for additional information。",
                    "suggestion": "如该链接确实需要提交到 EUDAMED，请在当前 v2.12 模板中人工复制到 UDI - Additional Information URL / eIFU webpage 后再导入。",
                }
            )

        for row_idx in range(DATA_START_ROW, self._last_data_row(ws, headers) + 1):
            raw, has_data = self._row_values(ws, headers, row_idx)
            if not has_data:
                continue

            basic_payload = {}
            udi_payload = {}
            basic_version = ""
            udi_version = ""
            local_record_id = ""

            for header, value in raw.items():
                item = schema_by_header.get(header) or schema_by_header.get(self._canonical_header(header))
                if not item:
                    continue
                self._collect_format_risks(ws, row_idx, headers.index(header) + 1, item["field"], value, format_warnings)
                if item.get("validation") in {"special_device_mdr", "special_device_ivdr"}:
                    value = self._enum_code(value)
                if item["entity"] == "basic":
                    basic_payload[item["field"]] = value
                elif item["entity"] == "udi":
                    udi_payload[item["field"]] = value
                elif item["entity"] == "meta":
                    if item["field"] == "record_id":
                        local_record_id = str(value or "").strip()
                    elif item["field"] == "basic_version":
                        basic_version = str(value).strip()
                    elif item["field"] == "udi_version":
                        udi_version = str(value).strip()

            self._normalize_basic_enums(basic_payload, ws.title, row_idx, migration_warnings)
            if local_record_id:
                basic_payload["Local Record ID"] = local_record_id
                udi_payload["Local Record ID"] = local_record_id

            if ws.title in ENTRY_SHEETS:
                legislation = str(basic_payload.get("Applicable Legislation") or "").strip().upper()
                allowed = set(ENTRY_SHEETS[ws.title].get("legislations") or [])
                if legislation and legislation not in allowed:
                    import_meta["normalization_blocked"] = True
                    identifier_errors.append(self._identifier_error(
                        ws.title,
                        row_idx,
                        "Applicable Legislation",
                        legislation,
                        "ENTRY_SHEET_LEGISLATION_MISMATCH",
                        f"{ws.title} sheet 不接受 {legislation}；该主表行未入库。",
                        "请把整行移到与 Applicable Legislation 对应的法规主表。",
                    ))
                    continue

            if local_record_id and local_record_id in duplicate_local_ids:
                import_meta["normalization_blocked"] = True
                identifier_errors.append(self._identifier_error(
                    ws.title,
                    row_idx,
                    "Local - Record ID",
                    local_record_id,
                    "LOCAL_RECORD_ID_DUPLICATE",
                    "Local - Record ID 在主表中重复；重复键对应的所有行均未写入产品库。",
                    "为每条主表记录分配整本工作簿唯一的 Local - Record ID。",
                ))
                continue

            original_basic_code = str(basic_payload.get("Basic UDI-DI Code") or "").strip()
            try:
                resolution = resolve_legacy_identifiers(basic_payload, udi_payload)
                if (
                    resolution
                    and not resolution.has_assigned_udi
                    and resolution.method != METHOD_EXISTING_EUDAMED_PAIR
                    and not local_record_id
                ):
                    raise LegacyIdentifierError("Legacy Device 没有已分配 UDI-DI 时，Local - Record ID 条件必填。")
            except LegacyIdentifierError as exc:
                import_meta["normalization_blocked"] = True
                identifier_errors.append(self._identifier_error(
                    ws.title,
                    row_idx,
                    "Legacy identifiers",
                    local_record_id,
                    "LEGACY_IDENTIFIER_INVALID",
                    str(exc),
                    "迁移到 v2.12，确认 Legacy 路径并修正输入；工具不会猜测或静默覆盖标识。",
                ))
                continue

            if resolution:
                audit = resolution.audit_payload()
                udi_payload.update(audit)
                basic_payload["Legacy EUDAMED DI"] = resolution.eudamed_di
                if not resolution.has_assigned_udi:
                    basic_payload["Basic UDI-DI Code"] = resolution.eudamed_di
                    basic_payload["Issuing Entity"] = "EUDAMED"
                    udi_payload["UDI-DI Code"] = resolution.eudamed_id
                    udi_payload["UDI-DI Issuing Entity"] = "EUDAMED"
                elif not str(basic_payload.get("Basic UDI-DI Code") or "").strip():
                    basic_payload["Basic UDI-DI Code"] = resolution.eudamed_di
                    basic_payload["Issuing Entity"] = "EUDAMED"
                import_meta["legacy_results"].append({
                    "sheet": ws.title,
                    "row": row_idx,
                    "local_record_id": local_record_id,
                    "method": resolution.method,
                    "eudamed_di": resolution.eudamed_di,
                    "eudamed_id": resolution.eudamed_id,
                    "identifier_code": resolution.identifier_code,
                })
                self._main_normalization_updates(
                    import_meta["normalization_updates"],
                    ws.title,
                    row_idx,
                    resolution,
                    preserve_assigned_basic=bool(resolution.has_assigned_udi and original_basic_code),
                )
            elif legacy_fields_present(udi_payload):
                migration_warnings.append({
                    "sheet": ws.title,
                    "row": row_idx,
                    "field": "Legacy fields",
                    "value": local_record_id,
                    "warning_type": "LEGACY_FIELDS_NOT_APPLICABLE",
                    "message": "MDR/IVDR 或 System/Procedure Pack 不进入 Legacy 标识生成流程，已忽略 Legacy 专属输入。",
                    "suggestion": "请清空 Legacy 专属字段，避免把本地计算字段误认为 EUDAMED XML 字段。",
                })

            basic_code = str(basic_payload.get("Basic UDI-DI Code", "")).strip()
            udi_code = str(udi_payload.get("UDI-DI Code", "")).strip()
            if basic_code:
                udi_payload["Parent Basic UDI-DI"] = basic_code

            if resolution and resolution.has_assigned_udi:
                self._warn_legacy_eudamed_di_derivation(
                    basic_payload,
                    udi_payload,
                    ws.title,
                    row_idx,
                    migration_warnings,
                )

            if local_record_id:
                if local_record_id in local_record_map:
                    # Defensive guard; duplicates are normally filtered by the pre-pass.
                    identifier_errors.append(self._identifier_error(
                        ws.title, row_idx, "Local - Record ID", local_record_id,
                        "LOCAL_RECORD_ID_DUPLICATE", "Local - Record ID 重复。", "请使用唯一键。",
                    ))
                    continue
                local_record_map[local_record_id] = {
                    "sheet": ws.title,
                    "row": row_idx,
                    "basic_storage_code": basic_code,
                    "basic_final_code": resolution.eudamed_di if resolution else basic_code,
                    "udi_code": udi_code,
                }

            if self._should_create_basic(basic_payload, basic_version):
                basic_payload["_row_number"] = row_idx
                clean_basic = self._clean_row(basic_payload)
                if basic_code not in basic_index:
                    basic_index[basic_code] = clean_basic
                    parsed["Basic UDI-DI"].append(basic_payload)
                else:
                    self._merge_missing_basic_fields(basic_index[basic_code], clean_basic)
                    for idx, item in enumerate(parsed["Basic UDI-DI"]):
                        if item.get("Basic UDI-DI Code") == basic_code:
                            merged = dict(item)
                            self._merge_missing_basic_fields(merged, clean_basic)
                            merged["_row_number"] = item.get("_row_number", row_idx)
                            parsed["Basic UDI-DI"][idx] = merged
                            break
                if basic_version:
                    import_meta["basic_versions"][basic_code] = basic_version

            if udi_code:
                udi_payload["_row_number"] = row_idx
                udi_payload["_sheet_name"] = ws.title
                parsed["UDI-DI"].append(udi_payload)
                if udi_version:
                    import_meta["udi_versions"][udi_code] = udi_version

    def _parse_related_sheet(
        self,
        ws,
        columns: list[dict],
        format_warnings: list[dict],
        migration_warnings: list[dict],
    ) -> list[dict]:
        headers = self._headers(ws)
        schema_by_header = self._schema_by_header(columns)
        rows = []
        for row_idx in range(DATA_START_ROW, self._last_data_row(ws, headers) + 1):
            raw, has_data = self._row_values(ws, headers, row_idx)
            if not has_data:
                continue
            row_data = {}
            for header, value in raw.items():
                item = schema_by_header.get(header) or schema_by_header.get(self._canonical_header(header))
                if item:
                    self._collect_format_risks(ws, row_idx, headers.index(header) + 1, item["field"], value, format_warnings)
                    if item.get("validation") in {
                        "critical_warning",
                        "storage_condition",
                        "certificate_type",
                        "clinical_size_type",
                        "clinical_size_unit",
                        "annex_xvi_nmd",
                        "substance_type",
                    }:
                        value = self._enum_code(value)
                    if item.get("validation") == "substance_type":
                        normalized = self._normal_substance_type(value)
                        if normalized and str(value or "").strip() != normalized:
                            migration_warnings.append(self._normalization_warning(
                                ws.title,
                                row_idx,
                                item["field"],
                                value,
                                normalized,
                                "旧模板/旧写法中的 Substance Type 已按当前模板下拉值自动归一；请核对是否符合实际物质类型。",
                            ))
                        value = normalized or value
                    row_data[item["field"]] = value
            if any(value not in ("", None) for value in row_data.values()):
                row_data["_row_number"] = row_idx
                rows.append(row_data)
        return rows

    def _headers(self, ws) -> list[str]:
        headers = []
        for cell in ws[1]:
            if cell.value is None:
                break
            headers.append(str(cell.value).strip())
        return headers

    def _schema_by_header(self, columns: list[dict]) -> dict[str, dict]:
        mapping = {}
        for item in columns:
            header = str(item.get("header") or "").strip()
            if not header:
                continue
            mapping[header] = item
            mapping.setdefault(self._canonical_header(header), item)
        return mapping

    def _canonical_header(self, header: str) -> str:
        return str(header or "").strip().rstrip("*").strip().casefold()

    def _last_data_row(self, ws, headers: list[str]) -> int:
        if not headers:
            return DATA_START_ROW - 1
        max_col = len(headers)
        for row_idx in range(ws.max_row, DATA_START_ROW - 1, -1):
            for col_idx in range(1, max_col + 1):
                if ws.cell(row_idx, col_idx).value not in (None, ""):
                    return row_idx
        return DATA_START_ROW - 1

    def _row_values(self, ws, headers: list[str], row_idx: int) -> tuple[dict, bool]:
        row_data = {}
        has_data = False
        for col_idx, header in enumerate(headers, start=1):
            value = ws.cell(row_idx, col_idx).value
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            if isinstance(value, str) and self._canonical_header(header) not in PRESERVE_WHITESPACE_HEADERS:
                value = value.strip()
            if value not in (None, ""):
                has_data = True
            row_data[header] = "" if value is None else value
        return row_data, has_data

    def _collect_format_risks(self, ws, row_idx: int, col_idx: int, field: str, value, warnings: list[dict]):
        if field not in FORMAT_RISK_FIELDS or value in (None, ""):
            return
        cell = ws.cell(row_idx, col_idx)
        text = str(value).strip()
        if SCIENTIFIC_NOTATION_RE.match(text):
            warnings.append(self._format_warning(ws.title, row_idx, field, text, "疑似科学计数法。请核对原始标签/条码，避免 UDI/GTIN/Reference 被 Excel/WPS 自动改写。"))
            return
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            warnings.append(self._format_warning(ws.title, row_idx, field, text, "该编码单元格是数字类型，可能已被 Excel/WPS 自动转换并丢失前导 0。请按文本格式维护。"))
        if cell.number_format not in {"@", "General"}:
            warnings.append(self._format_warning(ws.title, row_idx, field, text, f"该编码单元格格式为 {cell.number_format}，建议改为文本格式。"))
        if field in CODE_LENGTH_RISK_FIELDS and text.isdigit() and len(text) < 8:
            warnings.append(self._format_warning(ws.title, row_idx, field, text, "该 DI/UDI 编码长度较短，请核对是否丢失前导 0 或被自动转成数字。"))

    def _format_warning(self, sheet: str, row: int, field: str, value, message: str) -> dict:
        return {
            "sheet": sheet,
            "row": row,
            "field": field,
            "value": value,
            "warning_type": "FORMAT_RISK",
            "message": message,
            "suggestion": "在 Excel/WPS 中把该列设置为文本格式，并按标签/证书原值重新填写。",
        }

    def _template_version_warnings(self, wb) -> list[dict]:
        warnings = []
        detected = self._detect_template_version(wb)
        legacy_layout = bool(set(wb.sheetnames).intersection(set(IMPORT_ENTRY_SHEETS) - set(ENTRY_SHEETS)))
        if detected == TEMPLATE_VERSION and not legacy_layout:
            return warnings
        label = detected or "未知版本 / unknown"
        if detected == TEMPLATE_VERSION and legacy_layout:
            warnings.append(
                {
                    "sheet": "Workbook",
                    "row": "",
                    "field": "Template Layout",
                    "value": "MDR_MDD / IVDR_IVDD",
                    "warning_type": "TEMPLATE_LAYOUT_MIGRATION",
                    "message": "检测到拆表前的 v2.12 混合主表；本次仍兼容导入，规范化副本会按法规迁移到四个主表。",
                    "suggestion": "后续请使用包含 MDR、MDD_AIMDD、IVDR、IVDD 的当前 v2.12 模板。",
                }
            )
            return warnings
        warnings.append(
            {
                "sheet": "Workbook",
                "row": "",
                "field": "Template Version",
                "value": label,
                "warning_type": "TEMPLATE_VERSION_RISK",
                "message": f"检测到该文件可能不是当前 {TEMPLATE_VERSION} 模板，系统已按当前规则重新校验。",
                "suggestion": "请重点核对 Legacy 标识路径、Local Record ID、Special Device Type、CMR Substance Type、Package Info、IVD 人源/动物源字段、Trade Names、UDI-PI 适用性和 URL for additional information 等历史版本变化字段；建议先使用迁移模板功能生成当前模板。",
            }
        )
        return warnings

    def _detect_template_version(self, wb) -> str:
        if "How to Use" in wb.sheetnames:
            value = str(wb["How to Use"].cell(1, 1).value or "")
            match = re.search(r"EUDAMED Template (v\d+\.\d+)", value)
            if match:
                return match.group(1)
        for sheet_name in ENTRY_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            headers = set(self._headers(wb[sheet_name]))
            if "Basic - Special Device Type" in headers and "Basic - Is Suture/Staple/Filling/Brace (IIb Implant)" in headers:
                return ""
        return ""

    def _normalization_warning(self, sheet: str, row: int, field: str, old_value, new_value, message: str) -> dict:
        return {
            "sheet": sheet,
            "row": row,
            "field": field,
            "value": f"{old_value} -> {new_value}",
            "warning_type": "TEMPLATE_VALUE_NORMALIZED",
            "message": message,
            "suggestion": "如果自动归一不符合实际，请在当前模板下拉中重新选择正确值后再导入/导出。",
        }

    def _summary(self, parsed: dict) -> dict:
        return {
            "basic_count": len(parsed.get("Basic UDI-DI", [])),
            "udi_count": len(parsed.get("UDI-DI", [])),
            "market_count": len(parsed.get("Market Information", [])),
            "warning_count": len(parsed.get("Critical Warnings", [])),
            "storage_count": len(parsed.get("Storage Conditions", [])),
            "trade_name_count": len(parsed.get("Trade Names", [])),
            "cmr_count": len(parsed.get("CMR Substances", [])),
            "package_count": len(parsed.get("Package Information", [])),
            "certificate_count": len(parsed.get("Device Certificates", [])),
            "clinical_size_count": len(parsed.get("Clinical Sizes", [])),
            "annex_xvi_count": len(parsed.get("Annex XVI Purposes", [])),
        }

    def _change_summary(self, changes: list[dict]) -> dict:
        summary = {"created": 0, "updated": 0, "unchanged": 0}
        for item in changes:
            action = item.get("action")
            if action in summary:
                summary[action] += 1
        return summary

    def _clean_row(self, row: dict) -> dict:
        return {key: value for key, value in row.items() if not key.startswith("_")}

    def _merge_trade_name_shortcuts(self, parsed: dict, trade_name_map: dict):
        for row in parsed.get("UDI-DI", []):
            code = row.get("UDI-DI Code", "")
            if row.get("Trade Name"):
                continue
            rows = trade_name_map.get(code, [])
            if not rows:
                continue
            row["Trade Name"] = rows[0].get("Trade Name", "")
            row["Trade Name Language"] = rows[0].get("Language", "")

    def _trade_name_rows_for_udi(self, payload: dict, rows: list[dict]) -> list[dict]:
        merged = []
        if payload.get("Trade Name"):
            merged.append(
                {
                    "UDI-DI Code": payload.get("UDI-DI Code", ""),
                    "Trade Name": payload.get("Trade Name", ""),
                    "Language": payload.get("Trade Name Language", "") or "en",
                }
            )
        for row in rows:
            merged.append(self._clean_row(row))

        seen = set()
        deduped = []
        for row in merged:
            key = (
                str(row.get("UDI-DI Code", "")).strip(),
                str(row.get("Trade Name", "")).strip(),
                str(row.get("Language", "")).strip().lower(),
            )
            if not key[1] or key in seen:
                continue
            seen.add(key)
            deduped.append(row)
        return deduped

    def _enum_code(self, value):
        if isinstance(value, str) and " - " in value:
            return value.split(" - ", 1)[0].strip()
        return value

    def _normalize_basic_enums(self, payload: dict, sheet: str, row_idx: int, migration_warnings: list[dict]):
        original = payload.get("Special Device Type")
        special = self._special_device_code(original, payload)
        if special:
            # 只在【真正发生归一】时才警告：当前模板下拉值是「代码 - 标签」格式，剥成代码属正常操作，
            # 不应报「自动归一」。用提取后的代码与归一结果比较，避免对正确的当前下拉输入误报。
            if str(self._enum_code(original) or "").strip() != special:
                migration_warnings.append(self._normalization_warning(
                    sheet,
                    row_idx,
                    "Special Device Type",
                    original,
                    special,
                    "旧模板/旧写法中的 Special Device Type 已按当前法规枚举自动归一；请核对是否属于该产品的真实特殊类型。普通器械应留空。",
                ))
            payload["Special Device Type"] = special

    def _warn_legacy_eudamed_di_derivation(
        self,
        basic_payload: dict,
        udi_payload: dict,
        sheet: str,
        row_idx: int,
        migration_warnings: list[dict],
    ):
        legislation = str(basic_payload.get("Applicable Legislation") or "").strip().upper()
        if legislation not in {"MDD", "AIMDD", "IVDD"}:
            return
        udi_code = str(udi_payload.get("UDI-DI Code") or "").strip()
        raw_udi_issuing_entity = str(udi_payload.get("UDI-DI Issuing Entity") or "").strip()
        udi_issuing_entity = str(self._enum_code(raw_udi_issuing_entity) or "").strip().upper()
        if not udi_code or not raw_udi_issuing_entity or udi_issuing_entity == "EUDAMED":
            return
        expected_code = f"B-{udi_code}"
        supplied_code = str(basic_payload.get("Basic UDI-DI Code") or "").strip()
        supplied_entity = str(self._enum_code(basic_payload.get("Issuing Entity")) or "").strip().upper()
        if supplied_code == expected_code and supplied_entity == "EUDAMED":
            return
        migration_warnings.append(
            {
                "sheet": sheet,
                "row": row_idx,
                "field": "Basic UDI-DI Code / Issuing Entity",
                "value": f"{supplied_code or '(blank)'} / {supplied_entity or '(blank)'}",
                "warning_type": "LEGACY_EUDAMED_DI_DERIVED",
                "message": (
                    f"Legacy Device with assigned UDI-DI：XML 中 EUDAMED DI 将由 UDI-DI {udi_code} "
                    f"自动派生为 {expected_code}，issuing entity 固定为 EUDAMED；模板中的 Basic 字段"
                    "仅用于本地关联，不会原样写入本次 XML。"
                ),
                "suggestion": (
                    "请核对 UDI-DI Code 及其真实 issuing entity。无需为了 XML 手工把本地关联键改成 B-<UDI-DI>。"
                ),
            }
        )

    def _normal_substance_type(self, value) -> str:
        text = str(self._enum_code(value) or "").strip()
        if not text:
            return ""
        supported = set(ENUM_SOURCES.get("substance_type", []))
        if text in supported:
            return text
        aliases = {
            "CMR_1A": "CMR 1A",
            "CMR 1A": "CMR 1A",
            "CMR_1B": "CMR 1B",
            "CMR 1B": "CMR 1B",
            "ENDOCRINE_DISRUPTING_SUBSTANCE": "Endocrine Disrupting",
            "ENDOCRINE DISRUPTING": "Endocrine Disrupting",
            "MEDICINAL_PRODUCT_SUBSTANCE": "Medicinal Product Substance",
            "MEDICINAL PRODUCT SUBSTANCE": "Medicinal Product Substance",
            "HUMAN_PRODUCT_SUBSTANCE": "Human Blood or Plasma Substance",
            "HUMAN_BLOOD_OR_PLASMA_SUBSTANCE": "Human Blood or Plasma Substance",
            "HUMAN PRODUCT SUBSTANCE": "Human Blood or Plasma Substance",
            "HUMAN BLOOD OR PLASMA SUBSTANCE": "Human Blood or Plasma Substance",
        }
        return aliases.get(text.upper().replace("-", " ").replace("/", " "), "")

    def _special_device_code(self, value, payload: dict) -> str:
        text = str(self._enum_code(value) or "").strip()
        if not text:
            return ""
        values = self._special_device_values_for_legislation(payload.get("Applicable Legislation"))
        allowed_codes = {str(self._enum_code(item) or "").strip() for item in values}
        if text in allowed_codes:
            return text
        normalized_code = text.upper().replace(" ", "_").replace("-", "_")
        if normalized_code in allowed_codes:
            return normalized_code
        legacy_map = {
            "SOFTWARE": "SOFTWARE",
            "ORTHOPEDIC": "ORTHOPEDIC",
            "ORTHOPAEDIC": "ORTHOPEDIC",
            "STANDARD SOFT CONTACT LENSES": "STANDARD_SOFT_CONTACT_LENSES",
            "RIGID GAS PERMEABLE": "RIGID_GAS_PERMEABLE",
            "MADE TO ORDER": "MADE_TO_ORDER",
            "SPECTACLES FRAMES": "SPECTACLES_FRAMES",
            "SPECTACLES LENSES": "SPECTACLES_LENSES",
            "READY MADE SPECTACLES": "READY_MADE_SPECTACLES",
        }
        suffix = legacy_map.get(text.upper().replace("_", " ").replace("-", " "))
        if suffix:
            candidate = f"{self._special_device_preferred_prefix(payload.get('Applicable Legislation'))}_{suffix}"
            if candidate in allowed_codes:
                return candidate
        label_matches = []
        for item in values:
            code, label = self._split_enum_label(item)
            if label and label.lower() == text.lower():
                label_matches.append(code)
        if len(label_matches) == 1:
            return label_matches[0]
        if label_matches:
            prefix = self._special_device_preferred_prefix(payload.get("Applicable Legislation"))
            for code in label_matches:
                if code.startswith(prefix):
                    return code
        return ""

    def _special_device_values_for_legislation(self, legislation) -> list[str]:
        value = str(legislation or "").strip().upper()
        if value in {"IVDR", "IVDD"}:
            return ENUM_SOURCES.get("special_device_ivdr", [])
        return ENUM_SOURCES.get("special_device_mdr", [])

    def _special_device_preferred_prefix(self, legislation) -> str:
        value = str(legislation or "").strip().upper()
        if value in {"MDR", "MDD", "AIMDD", "IVDR", "IVDD"}:
            return value
        return "MDR"

    def _split_enum_label(self, value: str) -> tuple[str, str]:
        text = str(value or "")
        if " - " in text:
            return tuple(text.split(" - ", 1))  # type: ignore[return-value]
        return text, ""

    def _validate_related_rules(self, parsed: dict) -> list[dict]:
        errors = []
        udi_codes = {row.get("UDI-DI Code", "") for row in parsed.get("UDI-DI", [])}
        storage_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("storage_condition", [])}
        warning_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("critical_warning", [])}
        certificate_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("certificate_type", [])}
        clinical_size_type_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("clinical_size_type", [])}
        clinical_size_unit_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("clinical_size_unit", [])}
        annex_xvi_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("annex_xvi_nmd", [])}
        substance_codes = {self._enum_code(value) for value in ENUM_SOURCES.get("substance_type", [])}
        language_any_codes = set(ENUM_SOURCES.get("language_any", []))
        basic_codes = {str(row.get("Basic UDI-DI Code", "")).strip() for row in parsed.get("Basic UDI-DI", [])}

        self._validate_main_field_rules(parsed, errors)

        for row in parsed.get("Trade Names", []):
            udi_code = str(row.get("UDI-DI Code", "")).strip()
            trade_name = str(row.get("Trade Name", "")).strip()
            language = str(row.get("Language", "")).strip()
            if not udi_code:
                errors.append(self._validation_error(row, "Trade Names", "UDI-DI Code", "", "Trade Names 行缺少 UDI-DI Code。"))
            elif udi_code not in udi_codes:
                errors.append(self._validation_error(row, "Trade Names", "UDI-DI Code", udi_code, "Trade Names 引用的 UDI-DI Code 不存在于主表。"))
            if not trade_name:
                errors.append(self._validation_error(row, "Trade Names", "Trade Name", "", "Trade Names 行缺少 Trade Name。"))
            if not language:
                errors.append(self._validation_error(row, "Trade Names", "Language", "", "Trade Names 行缺少 Language。"))
            elif not self._valid_language_token(language, language_any_codes):
                errors.append(self._validation_error(row, "Trade Names", "Language", language, "Trade Names 的 Language 必须是官方语言代码或 ANY。"))

        for row in parsed.get("Critical Warnings", []):
            code = str(row.get("Warning Type", "")).strip()
            comment = str(row.get("Comment", "")).strip()
            language = str(row.get("Language", "")).strip().upper()
            if not code:
                errors.append(self._validation_error(row, "Critical Warnings", "Warning Type", "", "Critical Warnings 行缺少 Warning Type。"))
            elif code not in warning_codes:
                errors.append(self._validation_error(row, "Critical Warnings", "Warning Type", code, "Warning Type 不在官方 CriticalWarningEnum 中。"))
            if code == "CW999" and not comment:
                errors.append(self._validation_error(row, "Critical Warnings", "Comment", "", "CW999 - OTHER 必须填写 Comment。"))
            if code == "CW999" and language in {"", "ANY"}:
                errors.append(self._validation_error(row, "Critical Warnings", "Language", language, "CW999 - OTHER 必须选择具体语言，不能为 ANY。"))
        for row in parsed.get("Storage Conditions", []):
            code = str(row.get("Storage Condition Type", "")).strip()
            description = str(row.get("Description", "")).strip()
            language = str(row.get("Language", "")).strip().upper()
            if not code:
                errors.append(self._validation_error(row, "Storage Conditions", "Storage Condition Type", "", "Storage Conditions 行缺少 Storage Condition Type。"))
            elif code not in storage_codes:
                errors.append(self._validation_error(row, "Storage Conditions", "Storage Condition Type", code, "Storage Condition Type 不在官方 StorageHandlingConditionEnum 中。"))
            if code == "SHC099" and not description:
                errors.append(self._validation_error(row, "Storage Conditions", "Description", "", "SHC099 - OTHER 必须填写 Description。"))
            if code == "SHC099" and language in {"", "ANY"}:
                errors.append(self._validation_error(row, "Storage Conditions", "Language", language, "SHC099 - OTHER 必须选择具体语言，不能为 ANY。"))
        self._validate_market_rules(parsed, errors)
        self._validate_package_rules(parsed, errors, udi_codes)
        self._validate_substance_rules(parsed, errors, basic_codes, substance_codes)
        self._validate_certificate_rules(parsed, errors, basic_codes, certificate_codes)
        self._validate_clinical_size_rules(parsed, errors, udi_codes, clinical_size_type_codes, clinical_size_unit_codes)
        self._validate_annex_xvi_rules(parsed, errors, udi_codes, annex_xvi_codes)
        return errors

    def _validate_main_field_rules(self, parsed: dict, errors: list[dict]):
        for row in parsed.get("Basic UDI-DI", []):
            basic_code = str(row.get("Basic UDI-DI Code", "")).strip()
            special_device = str(self._enum_code(row.get("Special Device Type")) or "").strip()
            if special_device:
                allowed = self._special_device_codes_for_legislation(row.get("Applicable Legislation"))
                if allowed and special_device not in allowed:
                    errors.append(self._validation_error(
                        row,
                        "Basic UDI-DI",
                        "Special Device Type",
                        special_device,
                        f"Basic UDI-DI {basic_code} 的 Special Device Type 不属于当前法规对应的官方枚举；普通器械应留空。",
                    ))

            ii_b_exception = row.get("Is Suture/Staple/Filling/Brace (IIb Implant)")
            if ii_b_exception not in ("", None) and not self._valid_bool_token(ii_b_exception):
                errors.append(self._validation_error(
                    row,
                    "Basic UDI-DI",
                    "Is Suture/Staple/Filling/Brace (IIb Implant)",
                    ii_b_exception,
                    "该字段必须使用 TRUE/FALSE；仅 Class IIb + Implantable 时适用。",
                ))

    def _validate_substance_rules(
        self,
        parsed: dict,
        errors: list[dict],
        basic_codes: set[str],
        substance_codes: set[str],
    ):
        for row in parsed.get("CMR Substances", []):
            basic_code = str(row.get("Basic UDI-DI Code", "")).strip()
            substance_type = str(self._enum_code(row.get("Substance Type")) or "").strip()
            if not basic_code:
                errors.append(self._validation_error(row, "CMR Substances", "Basic UDI-DI Code", "", "CMR Substances 行缺少 Basic UDI-DI Code。"))
            elif basic_code not in basic_codes:
                errors.append(self._validation_error(row, "CMR Substances", "Basic UDI-DI Code", basic_code, "CMR Substances 引用的 Basic UDI-DI Code 不存在于主表。"))
            if not substance_type:
                errors.append(self._validation_error(row, "CMR Substances", "Substance Type", "", "CMR Substances 行必须选择 Substance Type。"))
            elif substance_type not in substance_codes:
                errors.append(self._validation_error(row, "CMR Substances", "Substance Type", substance_type, "Substance Type 不在本工具当前支持的安全输出类型中。"))

    def _validate_clinical_size_rules(
        self,
        parsed: dict,
        errors: list[dict],
        udi_codes: set[str],
        clinical_size_type_codes: set[str],
        clinical_size_unit_codes: set[str],
    ):
        for row in parsed.get("Clinical Sizes", []):
            udi_code = str(row.get("UDI-DI Code", "")).strip()
            size_type = str(row.get("Clinical Size Type", "")).strip()
            type_description = str(row.get("Clinical Size Type Description", "")).strip()
            precision = str(row.get("Precision", "")).strip()
            unit = str(row.get("Measure Unit", "")).strip()
            unit_description = str(row.get("Measure Unit Description", "")).strip()

            if not udi_code:
                errors.append(self._validation_error(row, "Clinical Sizes", "UDI-DI Code", "", "Clinical Sizes 行缺少 UDI-DI Code。"))
            elif udi_code not in udi_codes:
                errors.append(self._validation_error(row, "Clinical Sizes", "UDI-DI Code", udi_code, "Clinical Sizes 引用的 UDI-DI Code 不存在于主表。"))
            if not size_type:
                errors.append(self._validation_error(row, "Clinical Sizes", "Clinical Size Type", "", "Clinical Sizes 行缺少 Clinical Size Type。"))
            elif size_type not in clinical_size_type_codes:
                errors.append(self._validation_error(row, "Clinical Sizes", "Clinical Size Type", size_type, "Clinical Size Type 不在官方 ClinicalSizeTypeEnum 中。"))
            if size_type == "CST999" and not type_description:
                errors.append(self._validation_error(row, "Clinical Sizes", "Clinical Size Type Description", "", "CST999 - OTHER 必须填写 Clinical Size Type Description。"))
            if precision not in {"Range", "Value", "Text"}:
                errors.append(self._validation_error(row, "Clinical Sizes", "Precision", precision, "Precision 必须为 Range、Value 或 Text。"))
            if precision == "Range":
                if not self._is_number(row.get("Minimum")):
                    errors.append(self._validation_error(row, "Clinical Sizes", "Minimum", row.get("Minimum"), "Precision=Range 时 Minimum 必须填写数字。"))
                if not self._is_number(row.get("Maximum")):
                    errors.append(self._validation_error(row, "Clinical Sizes", "Maximum", row.get("Maximum"), "Precision=Range 时 Maximum 必须填写数字。"))
                self._validate_clinical_unit(row, errors, unit, unit_description, clinical_size_unit_codes)
            elif precision == "Value":
                if not self._is_number(row.get("Value")):
                    errors.append(self._validation_error(row, "Clinical Sizes", "Value", row.get("Value"), "Precision=Value 时 Value 必须填写数字。"))
                self._validate_clinical_unit(row, errors, unit, unit_description, clinical_size_unit_codes)
            elif precision == "Text" and not str(row.get("Text Value", "")).strip():
                errors.append(self._validation_error(row, "Clinical Sizes", "Text Value", "", "Precision=Text 时 Text Value 必须填写。"))

    def _validate_clinical_unit(self, row: dict, errors: list[dict], unit: str, unit_description: str, unit_codes: set[str]):
        if not unit:
            errors.append(self._validation_error(row, "Clinical Sizes", "Measure Unit", "", "Precision=Range/Value 时必须填写 Measure Unit。"))
        elif unit not in unit_codes:
            errors.append(self._validation_error(row, "Clinical Sizes", "Measure Unit", unit, "Measure Unit 不在官方 ClinicalSizeUnitEnum 中。"))
        if unit == "MU999" and not unit_description:
            errors.append(self._validation_error(row, "Clinical Sizes", "Measure Unit Description", "", "MU999 - OTHER 必须填写 Measure Unit Description。"))

    def _validate_annex_xvi_rules(self, parsed: dict, errors: list[dict], udi_codes: set[str], annex_xvi_codes: set[str]):
        seen = set()
        for row in parsed.get("Annex XVI Purposes", []):
            udi_code = str(row.get("UDI-DI Code", "")).strip()
            nmd_type = str(row.get("Non-Medical Device Type", "")).strip()
            if not udi_code:
                errors.append(self._validation_error(row, "Annex XVI Purposes", "UDI-DI Code", "", "Annex XVI Purposes 行缺少 UDI-DI Code。"))
            elif udi_code not in udi_codes:
                errors.append(self._validation_error(row, "Annex XVI Purposes", "UDI-DI Code", udi_code, "Annex XVI Purposes 引用的 UDI-DI Code 不存在于主表。"))
            if not nmd_type:
                errors.append(self._validation_error(row, "Annex XVI Purposes", "Non-Medical Device Type", "", "Annex XVI Purposes 行缺少 Non-Medical Device Type。"))
            elif nmd_type not in annex_xvi_codes:
                errors.append(self._validation_error(row, "Annex XVI Purposes", "Non-Medical Device Type", nmd_type, "Non-Medical Device Type 不在官方 NonMedicalDeviceEnum 中。"))
            key = (udi_code, nmd_type)
            if udi_code and nmd_type and key in seen:
                errors.append(self._validation_error(row, "Annex XVI Purposes", "Non-Medical Device Type", nmd_type, "同一 UDI-DI 的 Annex XVI Purpose 不应重复。"))
            seen.add(key)

    def _validate_certificate_rules(self, parsed: dict, errors: list[dict], basic_codes: set[str], certificate_codes: set[str]):
        for row in parsed.get("Device Certificates", []):
            basic_code = str(row.get("Basic UDI-DI Code", "")).strip()
            certificate_type = str(row.get("Certificate Type", "")).strip()
            nb_actor = str(row.get("Notified Body ID", "")).strip()
            expiry = str(row.get("Expiry Date", "")).strip()
            if not basic_code:
                errors.append(self._validation_error(row, "Device Certificates", "Basic UDI-DI Code", "", "Device Certificates 行缺少 Basic UDI-DI Code。"))
            elif basic_code not in basic_codes:
                errors.append(self._validation_error(row, "Device Certificates", "Basic UDI-DI Code", basic_code, "Device Certificates 引用的 Basic UDI-DI Code 不存在于主表。"))
            if not certificate_type:
                errors.append(self._validation_error(row, "Device Certificates", "Certificate Type", "", "Device Certificates 行缺少 Certificate Type。"))
            elif certificate_type not in certificate_codes:
                errors.append(self._validation_error(row, "Device Certificates", "Certificate Type", certificate_type, "Certificate Type 不在官方 GenericCertificateTypeEnum 中。"))
            if not nb_actor:
                errors.append(self._validation_error(row, "Device Certificates", "Notified Body ID", "", "Device Certificates 行缺少 Notified Body ID。"))
            elif not self._valid_nb_actor_code(nb_actor):
                errors.append(self._validation_error(row, "Device Certificates", "Notified Body ID", nb_actor, "Notified Body ID / NANDO ID 必须为 4 位数字，例如 0483。"))
            if expiry and not re.match(r"^\d{4}-\d{2}-\d{2}$", expiry):
                errors.append(self._validation_error(row, "Device Certificates", "Expiry Date", expiry, "Expiry Date 必须使用 YYYY-MM-DD 格式。"))

    def _valid_nb_actor_code(self, value) -> bool:
        return bool(re.fullmatch(r"\d{4}", str(value or "").strip()))

    def _validate_market_rules(self, parsed: dict, errors: list[dict]):
        udi_rows = {str(row.get("UDI-DI Code", "")).strip(): row for row in parsed.get("UDI-DI", [])}
        market_rows_by_udi = defaultdict(list)
        for row in parsed.get("Market Information", []):
            udi_code = str(row.get("UDI-DI Code", "")).strip()
            market_rows_by_udi[udi_code].append(row)

        for udi_code, udi_row in udi_rows.items():
            if str(udi_row.get("Device Status", "")).strip() != "On the EU market":
                continue
            rows = market_rows_by_udi.get(udi_code, [])
            if not rows:
                errors.append(self._validation_error(udi_row, "Market Info", "UDI-DI Code", udi_code, "On the EU market 的 UDI-DI 必须填写 Market Info。"))
                continue
            first_rows = [row for row in rows if self._is_true(row.get("Originally Placed on Market"))]
            if len(first_rows) != 1:
                row = first_rows[0] if first_rows else rows[0]
                errors.append(self._validation_error(
                    row,
                    "Market Info",
                    "Originally Placed on Market",
                    len(first_rows),
                    "每个 On the EU market UDI-DI 必须且只能有一条 Originally Placed on Market = TRUE；其它 made available 国家应填写 FALSE。",
                ))

    def _is_true(self, value) -> bool:
        if isinstance(value, bool):
            return value
        return str(value).strip().upper() in {"TRUE", "YES", "Y", "1"}

    def _valid_language_token(self, value, valid_values: set[str]) -> bool:
        text = str(value or "").strip()
        if not text:
            return False
        return text.upper() in valid_values or text.lower() in valid_values

    def _valid_bool_token(self, value) -> bool:
        if isinstance(value, bool):
            return True
        return str(value or "").strip().upper() in {"TRUE", "FALSE", "YES", "NO", "Y", "N", "1", "0"}

    def _special_device_codes_for_legislation(self, legislation) -> set[str]:
        value = str(legislation or "").strip().upper()
        if value in {"IVDR", "IVDD"}:
            return {self._enum_code(item) for item in ENUM_SOURCES.get("special_device_ivdr", [])}
        if value in {"MDR", "MDD", "AIMDD"}:
            return {self._enum_code(item) for item in ENUM_SOURCES.get("special_device_mdr", [])}
        return set()

    def _validate_package_rules(self, parsed: dict, errors: list[dict], udi_codes: set[str]):
        rows_by_udi = defaultdict(list)
        for row in parsed.get("Package Information", []):
            rows_by_udi[str(row.get("UDI-DI Code", "")).strip()].append(row)

        for udi_code, rows in rows_by_udi.items():
            if not udi_code:
                for row in rows:
                    errors.append(self._validation_error(row, "Package Info", "UDI-DI Code", "", "Package Info 行缺少 UDI-DI Code。"))
                continue
            if udi_code not in udi_codes:
                for row in rows:
                    errors.append(self._validation_error(row, "Package Info", "UDI-DI Code", udi_code, "Package Info 引用的 UDI-DI Code 不存在于主表。"))
                continue

            package_by_code = {}
            for row in rows:
                package_code = str(row.get("Package UDI-DI Code", "")).strip()
                package_issuing = str(row.get("Package Issuing Entity", "")).strip()
                quantity = row.get("Quantity per Package")
                child_code = str(row.get("Contains DI Code", "")).strip() or udi_code

                if not package_code:
                    errors.append(self._validation_error(row, "Package Info", "Package UDI-DI Code", "", "Package Info 行缺少 Package UDI-DI Code。"))
                elif package_code in package_by_code:
                    errors.append(self._validation_error(row, "Package Info", "Package UDI-DI Code", package_code, "同一 UDI-DI 下 Package UDI-DI Code 重复。"))
                else:
                    package_by_code[package_code] = row

                if not package_issuing:
                    errors.append(self._validation_error(row, "Package Info", "Package Issuing Entity", "", "Package Info 行缺少 Package Issuing Entity。"))
                if not self._positive_integer(quantity):
                    errors.append(self._validation_error(row, "Package Info", "Quantity per Package", quantity, "Quantity per Package 必须为正整数。"))
                if package_code and child_code == package_code:
                    errors.append(self._validation_error(row, "Package Info", "Contains DI Code", child_code, "Package DI 不能包含自己。"))

            valid_children = {udi_code} | set(package_by_code.keys())
            for row in rows:
                child_code = str(row.get("Contains DI Code", "")).strip() or udi_code
                if child_code not in valid_children:
                    errors.append(self._validation_error(row, "Package Info", "Contains DI Code", child_code, "child DI 必须是主 UDI-DI 或同一 UDI-DI 包装结构中已定义的 Package DI。"))

            edges = {}
            row_by_package = {}
            for row in rows:
                package_code = str(row.get("Package UDI-DI Code", "")).strip()
                child_code = str(row.get("Contains DI Code", "")).strip() or udi_code
                if package_code and child_code in package_by_code and child_code != package_code:
                    edges[package_code] = child_code
                    row_by_package[package_code] = row
            reported_cycles = set()
            for package_code in list(edges):
                cycle = self._package_cycle(package_code, edges)
                if cycle:
                    cycle_key = frozenset(cycle)
                    if cycle_key in reported_cycles:
                        continue
                    reported_cycles.add(cycle_key)
                    errors.append(self._validation_error(row_by_package.get(package_code, {}), "Package Info", "Contains DI Code", " -> ".join(cycle), "Package Info 存在循环包含关系。"))

    def _positive_integer(self, value) -> bool:
        try:
            return int(str(value).strip()) > 0
        except (TypeError, ValueError):
            return False

    def _is_number(self, value) -> bool:
        text = str(value or "").strip()
        if not text:
            return False
        try:
            float(text)
            return True
        except (TypeError, ValueError):
            return False

    def _package_cycle(self, start: str, edges: dict[str, str]) -> list[str]:
        seen = []
        current = start
        while current in edges:
            if current in seen:
                return seen[seen.index(current):] + [current]
            seen.append(current)
            current = edges[current]
        return []

    def _validation_error(self, row: dict, sheet: str, field: str, value, message: str) -> dict:
        return {
            "sheet": sheet,
            "row": row.get("_row_number", "?"),
            "field": field,
            "value": value,
            "error_type": "BUSINESS_RULE_ERROR",
            "message": message,
            "suggestion": "请按模板说明和 EUDAMED 业务规则补充该字段。",
        }

    def _should_create_basic(self, payload: dict, version: str) -> bool:
        basic_code = str(payload.get("Basic UDI-DI Code", "")).strip()
        if not basic_code:
            return False
        for field, value in payload.items():
            if field == "Basic UDI-DI Code":
                continue
            if value not in ("", None):
                return True
        return bool(version)

    def _merge_missing_basic_fields(self, existing: dict, incoming: dict):
        for field, value in incoming.items():
            if field == "_row_number" or value in ("", None):
                continue
            if existing.get(field) in ("", None):
                existing[field] = value


def parse_json_array(text: str) -> list[dict]:
    text = (text or "").strip()
    if not text:
        return []
    parsed = json.loads(text)
    if not isinstance(parsed, list):
        raise ValueError("JSON内容必须是数组")
    clean = []
    for item in parsed:
        if not isinstance(item, dict):
            raise ValueError("JSON数组中的每一项都必须是对象")
        clean.append(item)
    return clean
