#!/usr/bin/env python3

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TOOL_LIB = ROOT / "EUDAMED_TOOL_v2" / "lib"
TMP_DEPS = Path("/tmp/eudamed_pydeps")
for path in (TOOL_LIB, ROOT, TMP_DEPS):
    if path.exists() and str(path) not in sys.path:
        sys.path.insert(0, str(path))

import xlrd  # type: ignore
from openpyxl import load_workbook  # type: ignore

try:
    from .build_unified_template import build_workbook
except ImportError:
    from local_beta.build_unified_template import build_workbook


APPENDIX_PATH = ROOT / "Test sample" / "5432_Appendix A_B_C (MDR_EN ISO 13485) - Details on Types of Devices, Facilities and Suppliers_Rev.10.xlsx"
UDI_LIST_PATH = ROOT / "Test sample" / "UM-QR-9.0-12-02 UDI-DI清单(1) 新 - 副本.xls"
OUTPUT_PATH = ROOT / "Test sample" / "EUDAMED_Customer_Test_Template_Unimax_v2.11.xlsx"
ROW_LIMIT = 10


def main():
    appendix = load_workbook(APPENDIX_PATH, read_only=True, data_only=True)
    appendix_info = appendix["A.0| Basic Information"]
    appendix_products = appendix["A.1| Basic Product Info"]
    udi_book = xlrd.open_workbook(UDI_LIST_PATH)
    udi_sheet = udi_book.sheet_by_name("Sheet1")

    workbook = build_workbook()
    main_sheet = workbook["MDR_MDD"]
    market_sheet = workbook["Market Info"]
    package_sheet = workbook["Package Info"]
    trade_sheet = workbook["Trade Names"]
    annex_sheet = workbook["Annex XVI Purposes"]
    headers = {main_sheet.cell(1, col).value: col for col in range(1, main_sheet.max_column + 1)}
    market_headers = {market_sheet.cell(1, col).value: col for col in range(1, market_sheet.max_column + 1)}
    package_headers = {package_sheet.cell(1, col).value: col for col in range(1, package_sheet.max_column + 1)}
    trade_headers = {trade_sheet.cell(1, col).value: col for col in range(1, trade_sheet.max_column + 1)}
    annex_headers = {annex_sheet.cell(1, col).value: col for col in range(1, annex_sheet.max_column + 1)}

    manufacturer_srn = text(appendix_info["D10"].value)
    for idx in range(ROW_LIMIT):
        source_row = 4 + idx
        udi_row = 5 + idx
        target_row = 4 + idx
        if udi_row >= udi_sheet.nrows:
            break

        trade_name = text(appendix_products.cell(source_row, 1).value)
        model = text(appendix_products.cell(source_row, 2).value)
        reference = text(appendix_products.cell(source_row, 3).value)
        basic_udi = text(appendix_products.cell(source_row, 4).value)
        emdn = text(appendix_products.cell(source_row, 6).value)
        intended_purpose = text(appendix_products.cell(source_row, 18).value)
        risk_class = normalize_risk_class(text(appendix_products.cell(source_row, 19).value))
        medicinal_product = yes_no(appendix_products.cell(source_row, 21).value)
        annex_xvi = yes_no(appendix_products.cell(source_row, 23).value)
        reprocessing = yes_no(appendix_products.cell(source_row, 26).value)
        sterile = yes_no(appendix_products.cell(source_row, 41).value)

        udi_code = text(udi_sheet.cell_value(udi_row, 3))
        package_code = text(udi_sheet.cell_value(udi_row, 4))

        values = {
            "Local - Record ID": f"UNIMAX-{idx + 1:03d}",
            "Basic - Basic UDI-DI Code*": basic_udi,
            "Basic - Issuing Entity*": "GS1",
            "Basic - Manufacturer SRN*": manufacturer_srn,
            "Basic - Risk Class*": risk_class,
            "Basic - Applicable Legislation*": "MDR",
            "Basic - Device Type*": "Regular Device",
            "Basic - Device Name*": trade_name,
            "Basic - EMDN Code*": emdn,
            "Basic - Active Device": "FALSE",
            "Basic - Measuring Function": "FALSE",
            "Basic - Administer Medicine": "FALSE",
            "Basic - Implantable": "FALSE",
            "Basic - Reusable Surgical Instrument": "FALSE",
            "Basic - Presence of Human Tissues": "FALSE",
            "Basic - Presence of Animal Tissues": "FALSE",
            "Basic - Medicinal Product Device": bool_text(medicinal_product),
            "Basic - Additional Description": intended_purpose,
            "Basic - Device Model": "",
            "Basic - Is it a Kit": "FALSE",
            "Basic - Authorised Representative SRN": "NL-AR-000000247",
            "Basic - Reagent": "FALSE",
            "Basic - Presence of Medicinal Substance": bool_text(medicinal_product),
            "UDI - UDI-DI Code*": udi_code,
            "UDI - UDI-DI Issuing Entity*": "GS1",
            "UDI - Device Status*": "On the EU market",
            "UDI - Quantity of Device": "1",
            "UDI - Single Use Device*": "TRUE",
            "UDI - Max Number of Reuses": "0",
            "UDI - Device Labelled as Sterile*": bool_text(sterile),
            "UDI - Needs Sterilisation Before Use": "FALSE",
            "UDI - Containing Latex*": "FALSE",
            "UDI - Reprocessed Single Use Device": bool_text(reprocessing),
            "UDI - New Device (IVDR)": "FALSE",
            "UDI - Direct Marking": "FALSE",
            "UDI - Trade Name Applicable*": "TRUE",
            "UDI - Trade Name": trade_name,
            "UDI - Trade Name Language": "en",
            "UDI - Reference Number*": reference,
            "UDI - PI Lot/Batch Number": "TRUE",
            "UDI - PI Expiration Date": "TRUE",
            "UDI - PI Manufacturing Date": "TRUE",
            "UDI - PI Serial Number": "FALSE",
            "UDI - PI Software Identification": "FALSE",
            "UDI - Nomenclature Code*": emdn,
            "UDI - Nomenclature System": "EMDN",
            "UDI - Additional Description": model,
            "UDI - Description Language": "en",
        }

        for header, value in values.items():
            col = headers.get(header)
            if col:
                main_sheet.cell(target_row, col, value)

        market_row = 4 + idx
        market_values = {
            "UDI-DI Code*": udi_code,
            "Country Code*": "IT",
            "Placed on Market": "TRUE",
            "Originally Placed on Market*": "TRUE",
        }
        for header, value in market_values.items():
            col = market_headers.get(header)
            if col:
                market_sheet.cell(market_row, col, value)

        trade_row = 4 + idx
        trade_values = {
            "UDI-DI Code*": udi_code,
            "Trade Name*": trade_name,
            "Language*": "en",
        }
        for header, value in trade_values.items():
            col = trade_headers.get(header)
            if col:
                trade_sheet.cell(trade_row, col, value)

        if package_code:
            package_row = 4 + idx
            package_values = {
                "UDI-DI Code*": udi_code,
                "Local - Package Level": "Middle layer",
                "Local - Package Type": "Middle package",
                "Package UDI-DI Code*": package_code,
                "Package Issuing Entity*": "GS1",
                "Contains DI Code": udi_code,
                "Contains DI Issuing Entity": "GS1",
                "Quantity per Package*": "10",
            }
            for header, value in package_values.items():
                col = package_headers.get(header)
                if col:
                    package_sheet.cell(package_row, col, value)

        if annex_xvi:
            annex_row = 4 + idx
            annex_values = {
                "UDI-DI Code*": udi_code,
                # Source data only says Annex XVI applies; the specific official enum must be selected by a user.
                "Non-Medical Device Type*": "",
            }
            for header, value in annex_values.items():
                col = annex_headers.get(header)
                if col:
                    annex_sheet.cell(annex_row, col, value)

    workbook.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def yes_no(value) -> bool:
    return text(value).lower().startswith("yes")


def bool_text(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def normalize_risk_class(value: str) -> str:
    compact = value.lower().replace(" ", "")
    mapping = {
        "i": "Class I",
        "iia": "Class IIa",
        "iib": "Class IIb",
        "iii": "Class III",
    }
    return mapping.get(compact, value)


if __name__ == "__main__":
    main()
