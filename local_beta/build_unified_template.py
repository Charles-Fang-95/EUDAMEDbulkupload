#!/usr/bin/env python3

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TOOL_LIB = ROOT / "EUDAMED_TOOL_v2" / "lib"
if str(TOOL_LIB) not in sys.path:
    sys.path.insert(0, str(TOOL_LIB))
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl  # type: ignore
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore

try:
    from .template_schema import (
        ALL_COLUMNS,
        ENTRY_SHEETS,
        ENUM_SOURCES,
        RELATED_SHEETS,
        TEMPLATE_VERSION,
        columns_for_entry_sheet,
    )
except ImportError:
    from local_beta.template_schema import (
        ALL_COLUMNS,
        ENTRY_SHEETS,
        ENUM_SOURCES,
        RELATED_SHEETS,
        TEMPLATE_VERSION,
        columns_for_entry_sheet,
    )


OUTPUTS = [
    ("zh", ROOT / f"EUDAMED_Template_{TEMPLATE_VERSION}.xlsx"),
    ("zh", ROOT / "EUDAMED_TOOL_v2" / "templates" / f"EUDAMED_Template_{TEMPLATE_VERSION}.xlsx"),
    ("en", ROOT / f"EUDAMED_Template_{TEMPLATE_VERSION}_EN.xlsx"),
    ("en", ROOT / "EUDAMED_TOOL_v2" / "templates" / f"EUDAMED_Template_{TEMPLATE_VERSION}_EN.xlsx"),
]
DEFAULT_MAX_DATA_ROWS = 3000
HIGH_VOLUME_RELATED_SHEETS = {
    "Trade Names",
    "Market Info",
    "Package Info",
    "Critical Warnings",
    "Storage Conditions",
}
HIGH_VOLUME_MAX_DATA_ROWS = 10000
DATA_START_ROW = 4
PROTECTION_PASSWORD = "eudamed"

GROUP_FILLS = {
    "Meta": "D9E8FB",
    "Basic": "DFF2E1",
    "UDI": "FFF0C7",
    "Related": "F8DEE0",
}
REQUIREMENT_FILLS = {
    "required": "F8D7DA",
    "conditional": "FFF3CD",
    "optional": "EAF4EA",
}

HEADER_FONT = Font(name="Arial", bold=True, color="1F1F1F")
REQUIRED_HEADER_FONT = Font(name="Arial", bold=True, color="8A1C1C")
CONDITIONAL_HEADER_FONT = Font(name="Arial", bold=True, color="7A4D00")
HELP_TITLE = Font(name="宋体", bold=True, size=14)
HELP_HEAD = Font(name="宋体", bold=True, color="1F1F1F")
CHINESE_FONT = Font(name="宋体")
ENGLISH_FONT = Font(name="Arial")
THIN = Side(style="thin", color="C8BEAC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
UNLOCKED = Protection(locked=False)
HIGH_RISK_TEXT_TOKENS = (
    "UDI-DI Code",
    "Basic UDI-DI Code",
    "Package UDI-DI Code",
    "DM DI Code",
    "Unit of Use DI Code",
    "Secondary UDI-DI Code",
    "Reference Number",
    "SRN",
    "EMDN Code",
    "Nomenclature Code",
    "CAS Code",
    "EC Code",
    "Current Version",
)


def build_workbook(locale: str = "zh"):
    locale = "en" if locale == "en" else "zh"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "MDR_MDD"

    _build_entry_sheet(first, "MDR_MDD", locale)
    _build_entry_sheet(wb.create_sheet("IVDR_IVDD"), "IVDR_IVDD", locale)
    for sheet_name, spec in RELATED_SHEETS.items():
        _build_related_sheet(wb.create_sheet(sheet_name), sheet_name, spec["columns"], locale)

    _build_how_to_use(wb.create_sheet("How to Use"), locale)
    _build_glossary(wb.create_sheet("Glossary"), locale)
    _build_critical_warning_glossary(wb.create_sheet("Critical Warning Glossary"), locale)
    enums = wb.create_sheet("Enumerations")
    _build_enums(enums)
    enums.sheet_state = "hidden"
    return wb


def _build_entry_sheet(ws, sheet_name: str, locale: str):
    _build_table_sheet(ws, columns_for_entry_sheet(sheet_name), max_data_rows=DEFAULT_MAX_DATA_ROWS, locale=locale)
    ws.sheet_view.showGridLines = True


def _build_related_sheet(ws, sheet_name: str, columns: list[dict], locale: str):
    max_data_rows = HIGH_VOLUME_MAX_DATA_ROWS if sheet_name in HIGH_VOLUME_RELATED_SHEETS else DEFAULT_MAX_DATA_ROWS
    _build_table_sheet(ws, columns, max_data_rows=max_data_rows, locale=locale)


def _build_table_sheet(ws, columns: list[dict], max_data_rows: int, locale: str):
    for col_idx, item in enumerate(columns, start=1):
        header = ws.cell(1, col_idx, item["header"])
        header.font = _header_font(item)
        header.fill = PatternFill("solid", fgColor=GROUP_FILLS[item["group"]])
        header.alignment = WRAP
        header.border = BOX

        description = ws.cell(2, col_idx, _description_line(item, locale))
        description.font = ENGLISH_FONT if locale == "en" else CHINESE_FONT
        description.fill = PatternFill("solid", fgColor=REQUIREMENT_FILLS[item["requirement"]])
        description.alignment = WRAP
        description.border = BOX

        example = ws.cell(3, col_idx, item["example"])
        example.font = ENGLISH_FONT
        example.fill = PatternFill("solid", fgColor="F3F8E8")
        example.alignment = WRAP
        example.border = BOX

        width_base = max(len(item["header"]), len(str(item["example"] or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(42, width_base * 0.95))

    for row_idx, height in ((1, 42), (2, 84), (3, 50)):
        ws.row_dimensions[row_idx].height = height

    last_col = get_column_letter(len(columns))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A1:{last_col}1"
    _unlock_data_area(ws, len(columns), max_data_rows)
    _format_data_area(ws, len(columns), max_data_rows)
    _add_data_validations(ws, columns, max_data_rows, locale)
    _protect_sheet(ws)


def _description_line(item: dict, locale: str = "zh") -> str:
    if locale == "en":
        required = {"required": "Required", "conditional": "Conditional", "optional": "Optional"}[item["requirement"]]
        fmt = f"; Format: {item['format_en']}" if item.get("format_en") else ""
        description = str(item.get("description_en") or "")
        if description.startswith(f"{required}:"):
            return f"{description}{fmt}"
        return f"{required}. {description}{fmt}"
    required = {"required": "必填", "conditional": "条件必填", "optional": "可选"}[item["requirement"]]
    fmt = f"；格式：{item['format']}" if item["format"] else ""
    description = str(item["description"] or "")
    if description.startswith(f"{required}：") or description.startswith(f"{required}/"):
        return f"{description}{fmt}"
    return f"{required}。{description}{fmt}"


def _header_font(item: dict):
    if item["requirement"] == "required":
        return REQUIRED_HEADER_FONT
    if item["requirement"] == "conditional":
        return CONDITIONAL_HEADER_FONT
    return HEADER_FONT


def _unlock_data_area(ws, column_count: int, max_data_rows: int):
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=max_data_rows, min_col=1, max_col=column_count):
        for cell in row:
            cell.protection = UNLOCKED


def _format_data_area(ws, column_count: int, max_data_rows: int):
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=max_data_rows, min_col=1, max_col=column_count):
        for cell in row:
            cell.font = ENGLISH_FONT
            cell.number_format = "@"


def _protect_sheet(ws):
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False


def _build_enums(ws):
    for col_idx, (name, values) in enumerate(ENUM_SOURCES.items(), start=1):
        ws.cell(1, col_idx, name)
        ws.cell(1, col_idx).font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col_idx)].width = 44 if name in {"storage_condition", "critical_warning"} else 18
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)
            ws.cell(row_idx, col_idx).font = ENGLISH_FONT


def _build_how_to_use(ws, locale: str = "zh"):
    if locale == "en":
        lines = _how_to_use_lines_en()
    else:
        lines = _how_to_use_lines_zh()
    row = 1
    for text, font in lines:
        cell = ws.cell(row, 1, text)
        cell.alignment = WRAP
        cell.font = font or (ENGLISH_FONT if locale == "en" else CHINESE_FONT)
        row += 1
    ws.column_dimensions["A"].width = 120


def _how_to_use_lines_zh():
    lines = [
        (f"EUDAMED Template {TEMPLATE_VERSION} - How to Use", HELP_TITLE),
        ("1. 选择法规 sheet", HELP_HEAD),
        ("MDR/MDD 产品填写 MDR_MDD；IVDR/IVDD 产品填写 IVDR_IVDD。正式数据从第 4 行开始。", None),
        ("2. 三行表头", HELP_HEAD),
        ("第 1 行是程序识别字段名；第 2 行是中文说明；第 3 行是示例。前三行已锁定，请不要修改。", None),
        ("3. 推荐维护方式", HELP_HEAD),
        ("Excel template 是主维护文件；本地网页端用于批量导入、校验、筛选、选择 service 和导出 XML。", None),
        ("网页端数据库是工作库/导出历史库，不建议长期绕过 Excel 直接维护唯一数据源。", None),
        ("4. 一个 UDI-DI 一行", HELP_HEAD),
        ("一个 Basic UDI-DI 下有多个 UDI-DI 时，在主表重复填写 Basic 列，并逐行填写不同 UDI-DI。", None),
        ("5. 明细 sheet", HELP_HEAD),
        ("Trade Names、Market Info、Package Info、Device Certificates、Clinical Sizes、Annex XVI Purposes、Critical Warnings、Storage Conditions、CMR Substances 使用独立列填写，不再使用 | 分隔符。", None),
        ("Package Info 不是所有产品都要填：只有存在 container package / 多层包装 DI 时才填写；没有包装层级时整张 Package Info sheet 可留空。", None),
        ("Market Info 属于 UDI-DI 层，不属于 BUDI 层；独立 Update market information service 已在网页端开放，更新既有市场信息时可在导出页选择。", None),
        ("同一 UDI-DI 可以有多个 made available 国家：请在 Market Info sheet 为同一个 UDI-DI Code 填多行，每行一个 Country Code。", None),
        ("同一 UDI-DI 的 Originally Placed on Market 必须且只能有一个 TRUE，其它国家填写 FALSE。", None),
        ("国家/市场信息填报错误时，应优先通过 EUDAMED update/create new version 纠正，不应默认删除 UDI-DI 重新注册。", None),
        ("只有当 UDI-DI、器械身份或 Basic UDI-DI 关联本身错误，且无法通过更新纠正时，才考虑 discard/逻辑删除并重新注册。", None),
        ("6. IVDR / IVDD 常见字段", HELP_HEAD),
        ("IVDR/IVDD 也需要确认 Basic - Presence of Human Tissues 和 Basic - Presence of Animal Tissues。", None),
        ("Containing Latex 与 Reprocessed Single Use Device 不适用于 IVDR/IVDD，模板不会显示，导入也不应要求填写。", None),
        ("New Device (IVDR) 仅 IVDR Regulation Device 输出；IVDD Legacy 不输出。", None),
        ("UDI-PI 类型字段（Lot/Batch、Expiration Date、Manufacturing Date、Serial Number、Software Identification）不适用于 IVDD Legacy；可留空，导出时不会写入 XML。", None),
        ("7. UDI/GTIN/Reference 文本格式", HELP_HEAD),
        ("UDI-DI、Basic UDI-DI、Package UDI-DI、Reference、SRN、EMDN/Nomenclature 等字段必须按文本维护。", None),
        ("不要使用数字格式保存这些编码，避免 Excel/WPS 自动改成科学计数法或丢失前导 0。", None),
        ("8. 多语言 Trade Name", HELP_HEAD),
        ("主表 Trade Name 是快捷列；同一 UDI-DI 有多个商品名或多语言时，请在 Trade Names sheet 中逐行填写。", None),
        ("Trade Name 的 Language 可选 ANY，表示该商品名不限定具体语言；ANY 不会自动翻译，也不代表只能填写一个商品名。", None),
        ("如果同一商品名不限定具体语言，Language 请选择 ANY，不需要为 27 种官方语言重复建立 27 行。", None),
        ("只有同一 UDI-DI 在不同语言下确实有不同商品名/译名时，才需要在 Trade Names sheet 按语言逐行填写。", None),
        ("v2.9 起 Trade Names、Market Info、Package Info、Critical Warnings、Storage Conditions 等高频一对多明细表预设到第 10000 行；主表和低频明细表预设到第 3000 行，以兼顾容量和 Excel 性能。", None),
        ("9. Reference Number", HELP_HEAD),
        ("UDI - Reference Number* 是 EUDAMED XML 必填字段。没有内部 reference/catalogue number 时，请先与客户确认可提交的 reference。", None),
        ("10. Storage / Critical Warning 枚举", HELP_HEAD),
        ("下拉值采用 CODE - English label。程序导入时会自动取 CODE 输出 XML。SHC099 / CW999 为 OTHER，必须填写说明并选择具体语言。", None),
        ("Critical Warning Glossary sheet 提供完整官方 CW 枚举，便于按说明书或标签查找对应 warning。", None),
        ("11. Package Info 多层包装", HELP_HEAD),
        ("Package Info 每一行表示“一个 Package DI 包含一个 child DI”。child 可为主 UDI-DI，也可为同组下一级 Package DI。", None),
        ("如果填写 Package Info 行，UDI-DI Code、Package UDI-DI Code、Package Issuing Entity、Quantity per Package 为条件必填。", None),
        ("Contains DI Code 留空时兼容旧填法，默认 child 为主 UDI-DI；Local - Package Level / Type 只作本地辅助说明，不输出到 XML。", None),
        ("12. Special Device Type / CMR Substance", HELP_HEAD),
        ("Basic - Special Device Type 是官方枚举下拉，不是产品名称/型号；普通器械留空。MDR_MDD 与 IVDR_IVDD 使用不同枚举范围。", None),
        ("Basic - Is Suture/Staple/Filling/Brace 仅 Class IIb + Implantable 时适用，必须使用 TRUE/FALSE。", None),
        ("CMR Substances 的 Substance Type 必须从下拉中选择本工具当前支持的 5 类，不能自由输入。", None),
        ("CMR/Endocrine 类型可输出 CAS/EC；Medicinal Product / Human Blood or Plasma 类型只输出名称和官方 type，不输出 CAS/EC。", None),
        ("13. 上传 service", HELP_HEAD),
        ("新建 Basic + 多个 UDI-DI 使用 DEVICE.POST；已有 Basic 追加 UDI-DI 使用 UDI_DI.POST。新上传时可随 UDI-DI 一起输出 container package。", None),
        ("MDR/IVDR 会按 Regulation Device XML 输出；MDD/AIMDD/IVDD 会按 Legacy Device / EUDI XML 输出。", None),
        ("MDR/IVDR Regulation Device 或 SPP 如需声明 UDI-PI 类型，请在主表 UDI-PI 字段选择 TRUE/FALSE；MDD/AIMDD/IVDD Legacy 不输出 productionIdentifier。", None),
        ("AIMDD 的 Risk Class 通常选择 AIMDD；IVDD 可选择 IVD Annex II List A/B、IVD Self Testing 或 IVD General。", None),
        ("14. 特殊下拉值", HELP_HEAD),
        ("国家代码请使用官方 XSD 代码：希腊是 EL，不是 GR。", None),
        ("Issuing Entity 的 EUDAMED 通常用于 EUDAMED DI 等 legacy 场景；普通 UDI 请按实际发码机构选择 GS1/HIBCC/ICCBBA/IFA。", None),
        ("IFA 是官方 issuing entity 选项之一；是否适用以客户实际发码机构为准。", None),
        ("15. Device Certificates", HELP_HEAD),
        ("触发 MDR Art. 29(3) / IVDR Art. 26(2) 或 legacy 指令证书场景时，请在 Device Certificates sheet 逐行填写 product certificate 信息。", None),
        ("证书信息会输出到 Basic UDI-DI 的 deviceCertificateLinks；PR/SPP 证书结构暂未实现。", None),
        ("16. Clinical Sizes / Annex XVI", HELP_HEAD),
        ("Clinical Sizes 仅适用于 MDR UDI-DI；请在 Clinical Sizes sheet 按 Type、Precision、数值/文本和单位逐行填写。IVDR/MDD/AIMDD/IVDD 填写后导出会忽略并提示。", None),
        ("Annex XVI Purposes 仅适用于 MDR Annex XVI 非医疗目的产品；一个 UDI-DI 可在 Annex XVI Purposes sheet 填多行 Non-Medical Device Type。", None),
        ("17. 当前不输出字段", HELP_HEAD),
        ("eIFU URL、Public Email、Product Designer、Basic - Presence of Medicinal Substance 等字段仍在官方字段映射审计中；当前不会写入普通 UDI-DI XML。", None),
        ("18. 更新 service", HELP_HEAD),
        ("Basic_UDI.PATCH 需要在主表 B 列 Basic - Current Version 填写 EUDAMED 网页中的当前 Basic 版本号。", None),
        ("UDI_DI.PATCH 需要在主表 UDI 区域的 UDI - Current Version 列填写 EUDAMED 网页中的当前 UDI-DI 版本号。", None),
        ("独立 Update container package service 和 Update market information service 已在网页端开放；请在导出页选择对应 service。", None),
        ("19. Basic Model", HELP_HEAD),
        ("如果 EUDAMED 中 Model 不适用于 BUDI，请只填 Basic - Device Name*，并留空 Basic - Device Model。", None),
        ("20. WPS / Excel 兼容说明", HELP_HEAD),
        ("模板为无宏 .xlsx，理论上可在 WPS 中填写；实际兼容性以后续用户测试为准。导入时系统会提示疑似自动格式化风险。", None),
        ("21. 字体和锁定", HELP_HEAD),
        ("英文/代码/枚举值使用 Arial；中文说明使用宋体。前三行已锁定，正式填写从第 4 行开始。", None),
    ]
    return lines


def _how_to_use_lines_en():
    title_font = Font(name="Arial", bold=True, size=14)
    head_font = Font(name="Arial", bold=True, color="1F1F1F")
    return [
        (f"EUDAMED Template {TEMPLATE_VERSION}_EN - How to Use", title_font),
        ("1. Choose the legislation sheet", head_font),
        ("Use MDR_MDD for MDR/MDD/AIMDD devices and IVDR_IVDD for IVDR/IVDD devices. Start entering data from row 4.", None),
        ("2. Three header rows", head_font),
        ("Row 1 is the program-recognised field name; row 2 is the user instruction; row 3 is an example. The first three rows are locked.", None),
        ("3. Recommended workflow", head_font),
        ("Use Excel as the master data file. Use the local web tool only for import, validation, filtering, service selection and XML export.", None),
        ("The web database is a working and export-history database. Do not bypass Excel and treat the web database as the sole long-term data source.", None),
        ("4. One UDI-DI per main-sheet row", head_font),
        ("If one Basic UDI-DI has multiple UDI-DIs, repeat the Basic columns and fill one different UDI-DI per row.", None),
        ("5. Related sheets", head_font),
        ("Trade Names, Market Info, Package Info, Device Certificates, Clinical Sizes, Annex XVI Purposes, Critical Warnings, Storage Conditions and CMR Substances are maintained in separate sheets. Do not use | separators.", None),
        ("Package Info is required only when a container package or multi-level packaging DI exists. Leave the whole Package Info sheet blank when there is no package level.", None),
        ("Market Info is UDI-DI-level data, not Basic UDI-DI-level data. The standalone Update market information service is available on the export page for updating existing market information.", None),
        ("For multiple made-available countries, create multiple Market Info rows for the same UDI-DI Code, one country per row.", None),
        ("For one UDI-DI, Originally Placed on Market must have exactly one TRUE; other made-available countries should be FALSE.", None),
        ("Correct country or market information primarily through an EUDAMED update or create-new-version operation; do not default to deleting and re-registering the UDI-DI.", None),
        ("Consider discard or logical deletion and re-registration only when the UDI-DI, device identity or Basic UDI-DI association itself is wrong and cannot be corrected through an update.", None),
        ("6. Common IVDR / IVDD fields", head_font),
        ("IVDR/IVDD devices must also confirm Basic - Presence of Human Tissues and Basic - Presence of Animal Tissues.", None),
        ("Containing Latex and Reprocessed Single Use Device do not apply to IVDR/IVDD; the template does not show them and the importer must not require them.", None),
        ("New Device (IVDR) is exported only for IVDR Regulation Device, not IVDD Legacy.", None),
        ("UDI-PI type fields (Lot/Batch, Expiration Date, Manufacturing Date, Serial Number and Software Identification) do not apply to IVDD Legacy. They may be blank and are not written to XML.", None),
        ("7. Text-format identifiers", head_font),
        ("Maintain UDI-DI, Basic UDI-DI, Package UDI-DI, Reference, SRN, EMDN/Nomenclature and similar codes as text.", None),
        ("Do not store these codes as numbers, because Excel or WPS may convert them to scientific notation or remove leading zeros.", None),
        ("8. Multilingual trade names", head_font),
        ("The main-sheet Trade Name column is a quick entry. Use the Trade Names sheet for multiple names or language-specific names.", None),
        ("Language = ANY means no specific language is declared. It does not translate the trade name and does not mean only one trade name is allowed.", None),
        ("If the same trade name is not language-specific, choose ANY once; do not create 27 duplicate language rows.", None),
        ("Create separate Trade Names rows by language only when the same UDI-DI genuinely has different names or translations in different languages.", None),
        ("Since v2.9, high-volume one-to-many sheets such as Trade Names, Market Info, Package Info, Critical Warnings and Storage Conditions are preformatted to row 10000. Main sheets and lower-volume sheets are preformatted to row 3000 to balance capacity and Excel performance.", None),
        ("9. Reference Number", head_font),
        ("UDI - Reference Number is mandatory in EUDAMED XML. If there is no internal catalogue/reference number, confirm what may be submitted before upload.", None),
        ("10. Storage / Critical Warning enumerations", head_font),
        ("Dropdown values are shown as CODE - English label. The importer automatically extracts CODE for XML output. SHC099 / CW999 are OTHER values and require a comment/description plus a specific language.", None),
        ("The Critical Warning Glossary sheet contains the complete official CW enumeration so that the appropriate warning can be found from the instructions for use or label.", None),
        ("11. Multi-level Package Info", head_font),
        ("Each Package Info row means that one Package DI contains one child DI. The child may be the main UDI-DI or a lower-level Package DI in the same group.", None),
        ("When a Package Info row is used, UDI-DI Code, Package UDI-DI Code, Package Issuing Entity and Quantity per Package are conditionally required.", None),
        ("If Contains DI Code is blank, the legacy-compatible default child is the main UDI-DI. Local - Package Level / Type are local helper notes only and are not exported to XML.", None),
        ("12. Special Device Type / CMR Substance", head_font),
        ("Basic - Special Device Type is an official enumeration, not a product name or model. Leave it blank for an ordinary device. MDR_MDD and IVDR_IVDD use different enumeration ranges.", None),
        ("Basic - Is Suture/Staple/Filling/Brace applies only to Class IIb implantable devices and must use TRUE or FALSE.", None),
        ("CMR Substances - Substance Type must be selected from the five types currently supported by this tool and cannot be entered as free text.", None),
        ("CMR/Endocrine types may export CAS/EC. Medicinal Product and Human Blood or Plasma types export only the substance name and official type, not CAS/EC.", None),
        ("13. Upload services", head_font),
        ("Use DEVICE.POST to create a new Basic UDI-DI with its UDI-DI(s). Use UDI_DI.POST to add new UDI-DI(s) to an existing Basic UDI-DI.", None),
        ("MDR/IVDR are exported as Regulation Device XML; MDD/AIMDD/IVDD are exported as Legacy Device / EUDI XML.", None),
        ("For MDR/IVDR Regulation Devices or SPP, select TRUE/FALSE in the main-sheet UDI-PI fields when UDI-PI types must be declared. Production identifiers are not exported for MDD/AIMDD/IVDD Legacy devices.", None),
        ("For AIMDD, the Risk Class is normally AIMDD. For IVDD, select IVD Annex II List A/B, IVD Self Testing or IVD General as applicable.", None),
        ("14. Special dropdown values", head_font),
        ("Use official XSD country codes: Greece is EL, not GR.", None),
        ("The EUDAMED issuing entity is normally used for EUDAMED DI and similar legacy scenarios. For an ordinary UDI, select the actual issuing entity: GS1, HIBCC, ICCBBA or IFA.", None),
        ("IFA is an official issuing-entity option. Confirm applicability against the customer's actual issuing entity.", None),
        ("15. Device Certificates", head_font),
        ("If the device triggers MDR Art. 29(3), IVDR Art. 26(2) or a legacy directive certificate scenario, fill Device Certificates rows for the relevant Basic UDI-DI.", None),
        ("Certificate information is exported to the Basic UDI-DI deviceCertificateLinks. PR/SPP certificate structures are not yet implemented.", None),
        ("16. Clinical Sizes / Annex XVI", head_font),
        ("Clinical Sizes apply only to MDR UDI-DIs. Fill Type, Precision, value/text and unit as separate Clinical Sizes rows. Values entered for IVDR/MDD/AIMDD/IVDD are ignored on export with a warning.", None),
        ("Annex XVI Purposes apply only to MDR Annex XVI products without an intended medical purpose. One UDI-DI may have multiple Non-Medical Device Type rows in Annex XVI Purposes.", None),
        ("17. Fields currently not exported", head_font),
        ("eIFU URL, Public Email, Product Designer and Basic - Presence of Medicinal Substance remain under official field-mapping review and are not written to ordinary UDI-DI XML in the current tool.", None),
        ("18. Update services", head_font),
        ("For Basic_UDI.PATCH, fill Basic - Current Version in main-sheet column B with the current Basic version shown in EUDAMED.", None),
        ("For UDI_DI.PATCH, fill UDI - Current Version in the main-sheet UDI section with the current UDI-DI version shown in EUDAMED.", None),
        ("The standalone Update container package service and Update market information service are available in the web tool. Select the required service on the export page.", None),
        ("19. Basic Model", head_font),
        ("If Model is not applicable to the BUDI in EUDAMED, fill only Basic - Device Name and leave Basic - Device Model blank.", None),
        ("20. WPS / Excel compatibility", head_font),
        ("The template is a macro-free .xlsx and can in principle be completed in WPS. Actual compatibility remains subject to user testing. The importer warns about suspected automatic-formatting risks.", None),
        ("21. Fonts and locked rows", head_font),
        ("English text, codes and enumeration values use Arial; Chinese guidance uses SimSun. The first three rows are locked and data entry starts from row 4.", None),
    ]


def _build_glossary(ws, locale: str = "zh"):
    headers = ["Group", "Column", "Requirement", "Target", "Applies", "Meaning", "Format", "Example"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx, header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="E8E0CF")
        cell.border = BOX
    for row_idx, item in enumerate(ALL_COLUMNS, start=2):
        values = [
            item["group"],
            item["header"],
            (
                {"required": "Required", "conditional": "Conditional", "optional": "Optional"}[item["requirement"]]
                if locale == "en"
                else {"required": "必填", "conditional": "条件必填", "optional": "可选"}[item["requirement"]]
            ),
            f"{item['entity']} -> {item['field']}",
            item["applies"],
            item["description_en"] if locale == "en" else item["description"],
            _glossary_format(item, locale),
            item["example"],
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = WRAP
            cell.border = BOX
            cell.font = ENGLISH_FONT if locale == "en" else (CHINESE_FONT if col_idx in {3, 6} else ENGLISH_FONT)
    ws.freeze_panes = "A2"
    for idx, width in enumerate([14, 36, 10, 24, 14, 70, 28, 40], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _glossary_format(item: dict, locale: str = "zh") -> str:
    fmt = item["format_en"] if locale == "en" else item["format"]
    if _is_high_risk_text_header(item["header"]):
        if locale == "en":
            return f"{fmt}; maintain as text to avoid scientific notation or lost leading zeros" if fmt else "Maintain as text"
        return f"{fmt}；必须按文本维护，避免科学计数法或前导 0 丢失" if fmt else "必须按文本维护"
    return fmt


def _is_high_risk_text_header(header: str) -> bool:
    return any(token in header for token in HIGH_RISK_TEXT_TOKENS)


def _build_critical_warning_glossary(ws, locale: str = "zh"):
    headers = ["Code", "English Label", "OTHER", "Requires Comment", "Language Rule", "Usage Note"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx, header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="F8DEE0")
        cell.border = BOX
        cell.alignment = WRAP

    for row_idx, value in enumerate(ENUM_SOURCES.get("critical_warning", []), start=2):
        code, label = _split_enum(value)
        is_other = code == "CW999"
        values = [
            code,
            label,
            "Yes" if is_other else "No",
            "Yes" if is_other else "No",
            (
                "Choose a specific language; ANY is not allowed"
                if is_other and locale == "en"
                else "选择具体语言，不能为 ANY"
                if is_other
                else "Comment is exported with ANY if provided"
                if locale == "en"
                else "如填写 Comment，导出时使用 ANY"
            ),
            (
                "Use OTHER only when no official warning fits, and explain it in Comment."
                if is_other and locale == "en"
                else "找不到合适官方 warning 时才使用 OTHER，并在 Comment 说明。"
                if is_other
                else "Prefer the official enum; do not replace it with free text."
                if locale == "en"
                else "优先使用官方枚举，不要用自由文本替代。"
            ),
        ]
        for col_idx, item in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, item)
            cell.font = ENGLISH_FONT if locale == "en" else (CHINESE_FONT if col_idx in {5, 6} else ENGLISH_FONT)
            cell.alignment = WRAP
            cell.border = BOX

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for idx, width in enumerate([14, 80, 10, 18, 32, 54], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _split_enum(value: str) -> tuple[str, str]:
    if " - " in value:
        return tuple(value.split(" - ", 1))  # type: ignore[return-value]
    return value, ""


def _add_data_validations(ws, columns: list[dict], max_data_rows: int, locale: str = "zh"):
    enum_col_map = {name: idx for idx, name in enumerate(ENUM_SOURCES.keys(), start=1)}
    for col_idx, item in enumerate(columns, start=1):
        if not item["validation"]:
            continue
        enum_name = item["validation"]
        enum_col = enum_col_map[enum_name]
        enum_last_row = len(ENUM_SOURCES[enum_name]) + 1
        enum_letter = get_column_letter(enum_col)
        formula = f"=Enumerations!${enum_letter}$2:${enum_letter}${enum_last_row}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.showInputMessage = True
        dv.showErrorMessage = True
        dv.errorStyle = "stop"
        dv.promptTitle = "Field guidance" if locale == "en" else "填写说明"
        dv.errorTitle = "Invalid value" if locale == "en" else "输入无效"
        dv.prompt = item["description_en"] if locale == "en" else item["description"]
        dv.error = "Please select a valid dropdown value." if locale == "en" else "请输入下拉列表中的有效值。"
        target = f"{get_column_letter(col_idx)}{DATA_START_ROW}:{get_column_letter(col_idx)}{max_data_rows}"
        dv.add(target)
        ws.add_data_validation(dv)


def save_outputs():
    for locale, path in OUTPUTS:
        wb = build_workbook(locale)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        print(f"saved: {path}")


if __name__ == "__main__":
    save_outputs()
