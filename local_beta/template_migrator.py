from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from .build_unified_template import DATA_START_ROW, build_workbook
from .constants import EXPORT_DIR, TOOL_DIR, VENDOR_LIB
from .template_schema import ENTRY_SHEETS, ENUM_SOURCES, RELATED_SHEETS, TEMPLATE_VERSION, columns_for_entry_sheet

if str(VENDOR_LIB) not in sys.path:
    sys.path.insert(0, str(VENDOR_LIB))
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))

import openpyxl  # type: ignore


SOURCE_ROW_NUMBER = "_source_row_number"
LEGACY_EIFU_HEADERS = {"UDI - eIFU URL", "eIFU URL"}

LEGACY_RELATED_SHEET_MAP = {
    "Market Information": "Market Info",
    "Package Information": "Package Info",
}

RELATED_SHEET_ALIASES = {
    "Market Info": ["Market Info", "Market Information"],
    "Package Info": ["Package Info", "Package Information"],
    "Critical Warnings": ["Critical Warnings"],
    "Storage Conditions": ["Storage Conditions"],
    "CMR Substances": ["CMR Substances"],
    "Trade Names": ["Trade Names"],
    "Device Certificates": ["Device Certificates"],
    "Clinical Sizes": ["Clinical Sizes"],
    "Annex XVI Purposes": ["Annex XVI Purposes"],
}

PACKAGE_HEADER_ALIASES = {
    "Package Level": "Local - Package Level",
    "Package Type": "Local - Package Type",
}

MAIN_HEADER_ALIASES = {
    "Basic - Kit (IVDR)": "Basic - Is it a Kit",
    "Kit (IVDR)": "Basic - Is it a Kit",
}


def migrate_workbook(source_path: Path, output_dir: Path = EXPORT_DIR) -> dict:
    """Copy known EUDAMED template data into the current generated template.

    The migrator only auto-copies fields it can match by current header, legacy
    header, or stable field name. Unknown customer columns are reported instead
    of guessed.
    """
    if source_path.suffix.lower() != ".xlsx":
        return {
            "ok": False,
            "errors": ["目前迁移工具只支持 .xlsx。旧 .xls 请先用 Excel/WPS 另存为 .xlsx。"],
            "warnings": [],
            "output_filename": "",
            "report": {},
        }

    output_dir.mkdir(parents=True, exist_ok=True)
    source = openpyxl.load_workbook(source_path, data_only=True)
    target = build_workbook()
    report = {
        "source_file": source_path.name,
        "source_sheets": list(source.sheetnames),
        "mode": "unknown",
        "copied_rows": defaultdict(int),
        "unmapped_headers": defaultdict(list),
        "warnings": [],
        "legacy_eifu_migrations": [],
    }
    detected_version = _detect_template_version(source)
    if detected_version != TEMPLATE_VERSION:
        report["warnings"].append(
            f"源文件模板版本为 {detected_version or '未知版本'}，当前模板为 {TEMPLATE_VERSION}；迁移后请重点核对 Special Device Type、CMR Substance Type、Is Suture/Staple/Filling/Brace、Package Info 和 IVD 人源/动物源字段。"
        )

    if any(sheet in source.sheetnames for sheet in ENTRY_SHEETS):
        report["mode"] = "current_unified_template"
        _copy_unified_sheets(source, target, report)
    elif "Basic UDI-DI" in source.sheetnames and "UDI-DI" in source.sheetnames:
        report["mode"] = "legacy_split_template"
        _copy_legacy_split_sheets(source, target, report)
    else:
        report["warnings"].append("未识别到 MDR_MDD / IVDR_IVDD 或旧版 Basic UDI-DI + UDI-DI sheet；未自动搬迁数据。")

    _copy_related_sheets(source, target, report)
    _add_report_sheet(target, report)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"MIGRATED_EUDAMED_Template_{TEMPLATE_VERSION}_{timestamp}.xlsx"
    target.save(output_path)
    return {
        "ok": True,
        "errors": [],
        "warnings": list(report["warnings"]),
        "output_filename": output_path.name,
        "report": _serializable_report(report),
    }


def _copy_unified_sheets(source, target, report: dict):
    for sheet_name in ENTRY_SHEETS:
        if sheet_name not in source.sheetnames:
            continue
        rows = _read_rows(source[sheet_name])
        target_headers = _headers(target[sheet_name])
        source_headers = _headers(source[sheet_name])
        header_map = _header_map(source_headers, target_headers, extra_aliases=MAIN_HEADER_ALIASES)
        _record_unmapped(report, sheet_name, source_headers, header_map)
        target_row = DATA_START_ROW
        for row in rows:
            if not _has_data(row):
                continue
            _write_row(target[sheet_name], target_headers, target_row, row, header_map)
            _migrate_legacy_eifu_url(target[sheet_name], target_headers, target_row, row, report, sheet_name)
            _normalize_main_row(target[sheet_name], target_headers, target_row, report)
            _append_old_udi_detail_rows(target, row, report)
            target_row += 1
            report["copied_rows"][sheet_name] += 1


def _copy_legacy_split_sheets(source, target, report: dict):
    basic_rows = _read_rows(source["Basic UDI-DI"])
    udi_rows = _read_rows(source["UDI-DI"])
    basics = {}
    for row in basic_rows:
        code = str(_get_by_alias(row, "Basic UDI-DI Code") or "").strip()
        if code:
            basics[code] = row

    target_next_rows = {sheet: DATA_START_ROW for sheet in ENTRY_SHEETS}
    for udi in udi_rows:
        udi_code = str(_get_by_alias(udi, "UDI-DI Code") or "").strip()
        if not udi_code:
            continue
        parent = str(_get_by_alias(udi, "Parent Basic UDI-DI") or "").strip()
        basic = basics.get(parent, {})
        legislation = str(_get_by_alias(basic, "Applicable Legislation") or "").upper()
        target_sheet = "IVDR_IVDD" if legislation in {"IVDR", "IVDD"} else "MDR_MDD"
        target_headers = _headers(target[target_sheet])
        combined = _legacy_combined_row(basic, udi)
        _migrate_legacy_split_eifu_url(combined, udi, report, target_sheet)
        _write_by_field(target[target_sheet], target_headers, target_next_rows[target_sheet], combined)
        _normalize_main_row(target[target_sheet], target_headers, target_next_rows[target_sheet], report)
        _append_old_udi_detail_rows(target, udi, report)
        target_next_rows[target_sheet] += 1
        report["copied_rows"][target_sheet] += 1

    source_headers = _headers(source["Basic UDI-DI"]) + _headers(source["UDI-DI"])
    matched = set(_legacy_field_map().keys()) | {"Parent Basic UDI-DI"}
    report["unmapped_headers"]["Legacy Basic/UDI"] = [
        header
        for header in source_headers
        if header
        and not _is_legacy_eifu_header(header)
        and _canonical(header) not in {_canonical(item) for item in matched}
    ]


def _copy_related_sheets(source, target, report: dict):
    for target_sheet, aliases in RELATED_SHEET_ALIASES.items():
        source_sheet = next((name for name in aliases if name in source.sheetnames), "")
        if not source_sheet:
            continue
        rows = _read_rows(source[source_sheet])
        target_headers = _headers(target[target_sheet])
        source_headers = _headers(source[source_sheet])
        header_map = _header_map(source_headers, target_headers, extra_aliases=PACKAGE_HEADER_ALIASES)
        _record_unmapped(report, source_sheet, source_headers, header_map)
        target_row = DATA_START_ROW
        for row in rows:
            if not _has_data(row):
                continue
            _write_row(target[target_sheet], target_headers, target_row, row, header_map)
            _normalize_related_row(target[target_sheet], target_headers, target_row, report)
            target_row += 1
            report["copied_rows"][target_sheet] += 1


def _read_rows(ws) -> list[dict]:
    headers = _headers(ws)
    rows = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row = {}
        for col_idx, header in enumerate(headers, start=1):
            if header:
                row[header] = ws.cell(row_idx, col_idx).value
        if _has_data(row):
            row[SOURCE_ROW_NUMBER] = row_idx
            rows.append(row)
    return rows


def _headers(ws) -> list[str]:
    return [str(cell.value or "").strip() for cell in ws[1]]


def _header_map(source_headers: list[str], target_headers: list[str], extra_aliases: dict | None = None) -> dict:
    extra_aliases = extra_aliases or {}
    target_by_key = {_canonical(header): header for header in target_headers if header}
    mapping = {}
    for source_header in source_headers:
        if not source_header:
            continue
        target_header = extra_aliases.get(source_header, source_header)
        key = _canonical(target_header)
        if key in target_by_key:
            mapping[source_header] = target_by_key[key]
    return mapping


def _write_row(ws, target_headers: list[str], target_row: int, row: dict, header_map: dict):
    target_index = {header: idx for idx, header in enumerate(target_headers, start=1)}
    for source_header, target_header in header_map.items():
        if target_header in target_index:
            ws.cell(target_row, target_index[target_header]).value = row.get(source_header)


def _write_by_field(ws, target_headers: list[str], target_row: int, field_values: dict):
    target_index = {header: idx for idx, header in enumerate(target_headers, start=1)}
    field_to_header = {item["field"]: item["header"] for item in columns_for_entry_sheet(ws.title)}
    for field, value in field_values.items():
        header = field_to_header.get(field)
        if header in target_index:
            ws.cell(target_row, target_index[header]).value = value


def _legacy_combined_row(basic: dict, udi: dict) -> dict:
    mapped = {}
    field_map = _legacy_field_map()
    for header, value in basic.items():
        field = field_map.get(_canonical(header))
        if field:
            mapped[field] = value
    for header, value in udi.items():
        field = field_map.get(_canonical(header))
        if field:
            mapped[field] = value
    parent = _get_by_alias(udi, "Parent Basic UDI-DI")
    if parent:
        mapped["Basic UDI-DI Code"] = parent
    return mapped


def _migrate_legacy_eifu_url(ws, target_headers: list[str], target_row: int, source_row: dict, report: dict, sheet_name: str):
    old_url = _legacy_eifu_value(source_row)
    if not old_url:
        return
    field_to_header = {item["field"]: item["header"] for item in columns_for_entry_sheet(ws.title)}
    target_header = field_to_header.get("Additional Information URL")
    if not target_header or target_header not in target_headers:
        _record_legacy_eifu_migration(report, sheet_name, source_row, old_url, "skipped_no_target", "")
        return
    cell = ws.cell(target_row, target_headers.index(target_header) + 1)
    current_value = str(cell.value or "").strip()
    if not current_value:
        cell.value = old_url
        _record_legacy_eifu_migration(report, sheet_name, source_row, old_url, "copied_to_additional_information_url", "")
        return
    if current_value == old_url:
        _record_legacy_eifu_migration(report, sheet_name, source_row, old_url, "already_same_as_current_field", current_value)
        return
    _record_legacy_eifu_migration(report, sheet_name, source_row, old_url, "kept_existing_additional_information_url", current_value)


def _migrate_legacy_split_eifu_url(mapped: dict, source_row: dict, report: dict, target_sheet: str):
    old_url = _legacy_eifu_value(source_row)
    if not old_url:
        return
    current_value = str(mapped.get("Additional Information URL") or "").strip()
    if not current_value:
        mapped["Additional Information URL"] = old_url
        _record_legacy_eifu_migration(report, target_sheet, source_row, old_url, "copied_to_additional_information_url", "")
        return
    if current_value == old_url:
        _record_legacy_eifu_migration(report, target_sheet, source_row, old_url, "already_same_as_current_field", current_value)
        return
    _record_legacy_eifu_migration(report, target_sheet, source_row, old_url, "kept_existing_additional_information_url", current_value)


def _legacy_eifu_value(row: dict) -> str:
    for header in LEGACY_EIFU_HEADERS:
        value = _get_by_alias(row, header)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _record_legacy_eifu_migration(report: dict, sheet_name: str, source_row: dict, old_url: str, result: str, current_url: str):
    udi_code = _udi_code_value(source_row)
    source_row_number = source_row.get(SOURCE_ROW_NUMBER, "")
    entry = {
        "sheet": sheet_name,
        "source_row": source_row_number,
        "udi_code": udi_code,
        "old_url": old_url,
        "current_url": current_url,
        "result": result,
    }
    report["legacy_eifu_migrations"].append(entry)
    if result == "copied_to_additional_information_url":
        report["warnings"].append(
            f"{sheet_name} 源第 {source_row_number} 行 UDI-DI {udi_code}：旧 eIFU URL 已复制到 UDI - Additional Information URL / eIFU webpage。"
            "请确认该链接确实可作为官方 URL for additional information 提交。"
        )
    elif result == "kept_existing_additional_information_url":
        report["warnings"].append(
            f"{sheet_name} 源第 {source_row_number} 行 UDI-DI {udi_code}：旧 eIFU URL 与现有 Additional Information URL 不同，"
            "迁移工具保留现有新字段，未覆盖；请人工核对两个 URL。"
        )
    elif result == "skipped_no_target":
        report["warnings"].append(
            f"{sheet_name} 源第 {source_row_number} 行 UDI-DI {udi_code}：检测到旧 eIFU URL，但当前模板缺少目标字段，未迁移。"
        )


def _legacy_field_map() -> dict:
    fields = {}
    for item in columns_for_entry_sheet("MDR_MDD") + columns_for_entry_sheet("IVDR_IVDD"):
        fields[_canonical(item["field"])] = item["field"]
        fields[_canonical(item["header"])] = item["field"]
    aliases = {
        "Device Name/Model": "Device Name/Model",
        "Device Name/Model*": "Device Name/Model",
        "Parent Basic UDI-DI": "Basic UDI-DI Code",
    }
    for source, field in aliases.items():
        fields[_canonical(source)] = field
    return fields


def _udi_code_value(row: dict) -> str:
    for header in ("UDI - UDI-DI Code", "UDI - UDI-DI Code*", "UDI-DI Code", "UDI-DI Code*"):
        value = _get_by_alias(row, header)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _append_old_udi_detail_rows(target, source_row: dict, report: dict):
    udi_code = _udi_code_value(source_row)
    if not udi_code:
        return

    old_size_value = _get_by_alias(source_row, "Clinical Size Value")
    old_size_unit = _get_by_alias(source_row, "Clinical Size Unit")
    if old_size_value not in (None, "") or old_size_unit not in (None, ""):
        _append_related_row(
            target["Clinical Sizes"],
            {
                "UDI-DI Code*": udi_code,
                "Precision*": "Value",
                "Value": old_size_value,
                "Measure Unit": old_size_unit,
            },
        )
        report["copied_rows"]["Clinical Sizes"] += 1
        report["warnings"].append(
            f"UDI-DI {udi_code} 的旧 Clinical Size Value/Unit 已迁移到 Clinical Sizes sheet；请补充 Clinical Size Type，并确认 Measure Unit 使用 {TEMPLATE_VERSION} 下拉值。"
        )

    old_annex = _get_by_alias(source_row, "Purpose Other Than Medical")
    if _truthy(old_annex):
        _append_related_row(
            target["Annex XVI Purposes"],
            {
                "UDI-DI Code*": udi_code,
            },
        )
        report["copied_rows"]["Annex XVI Purposes"] += 1
        report["warnings"].append(
            f"UDI-DI {udi_code} 的旧 Purpose Other Than Medical=TRUE 无法自动判断 Annex XVI 具体类型；请在 Annex XVI Purposes sheet 选择 Non-Medical Device Type。"
        )


def _append_related_row(ws, values_by_header: dict):
    headers = _headers(ws)
    target_index = {header: idx for idx, header in enumerate(headers, start=1)}
    row_idx = _next_data_row(ws)
    for header, value in values_by_header.items():
        if header in target_index:
            ws.cell(row_idx, target_index[header]).value = value


def _next_data_row(ws) -> int:
    for row_idx in range(DATA_START_ROW, ws.max_row + 2):
        values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            return row_idx
    return ws.max_row + 1


def _detect_template_version(wb) -> str:
    if "How to Use" in wb.sheetnames:
        value = str(wb["How to Use"].cell(1, 1).value or "")
        match = re.search(r"EUDAMED Template (v\d+\.\d+)", value)
        if match:
            return match.group(1)
    return ""


def _normalize_main_row(ws, target_headers: list[str], row_idx: int, report: dict):
    field_to_header = {item["field"]: item["header"] for item in columns_for_entry_sheet(ws.title)}
    header = field_to_header.get("Special Device Type")
    if not header or header not in target_headers:
        return
    cell = ws.cell(row_idx, target_headers.index(header) + 1)
    value = cell.value
    if value in (None, ""):
        return
    legislation_header = field_to_header.get("Applicable Legislation")
    legislation = ""
    if legislation_header in target_headers:
        legislation = str(ws.cell(row_idx, target_headers.index(legislation_header) + 1).value or "").strip()
    display = _special_device_display(value, ws.title, legislation)
    if display:
        if str(value or "").strip() != display:
            report["warnings"].append(
                f"{ws.title} 第 {row_idx} 行 Special Device Type 已自动归一：{value} -> {display}。请核对是否属于该产品的真实特殊类型；普通器械应留空。"
            )
        cell.value = display
        return
    report["warnings"].append(
        f"{ws.title} 第 {row_idx} 行 Special Device Type={value} 无法匹配官方枚举；已保留原值，请人工核对。"
    )


def _normalize_related_row(ws, target_headers: list[str], row_idx: int, report: dict):
    if ws.title != "CMR Substances" or "Substance Type" not in target_headers:
        return
    cell = ws.cell(row_idx, target_headers.index("Substance Type") + 1)
    value = cell.value
    if value in (None, ""):
        return
    normalized = _normal_substance_type(value)
    if normalized:
        if str(value or "").strip() != normalized:
            report["warnings"].append(
                f"CMR Substances 第 {row_idx} 行 Substance Type 已自动归一：{value} -> {normalized}。请核对是否符合实际物质类型。"
            )
        cell.value = normalized
        return
    report["warnings"].append(
        f"CMR Substances 第 {row_idx} 行 Substance Type={value} 无法匹配当前支持类型；已保留原值，请人工核对。"
    )


def _special_device_display(value, sheet_name: str, legislation: str = "") -> str:
    text = str(_enum_code(value) or "").strip()
    if not text:
        return ""
    values = _special_device_values(sheet_name, legislation)
    code_to_display = {_enum_code(item): item for item in values}
    if text in code_to_display:
        return code_to_display[text]
    normalized_code = text.upper().replace(" ", "_").replace("-", "_")
    if normalized_code in code_to_display:
        return code_to_display[normalized_code]
    aliases = {
        "SOFTWARE": "SOFTWARE",
        "ORTHOPEDIC": "ORTHOPEDIC",
        "ORTHOPAEDIC": "ORTHOPEDIC",
        "STANDARD SOFT CONTACT LENSES": "STANDARD_SOFT_CONTACT_LENSES",
        "RIGID GAS PERMEABLE": "RIGID_GAS_PERMEABLE",
        "MADE TO ORDER": "MADE_TO_ORDER",
        "SPECTACLES FRAMES": "SPECTACLES_FRAMES",
        "SPECTACLES LENSES": "SPECTACLES_LENSES",
        "READY MADE SPECTACLES": "READY_MADE_SPECTACLES",
    }
    suffix = aliases.get(text.upper().replace("_", " ").replace("-", " "))
    if suffix:
        candidate = f"{_special_device_prefix(sheet_name, legislation)}_{suffix}"
        return code_to_display.get(candidate, "")
    label_matches = []
    for item in values:
        code, label = _split_enum_label(item)
        if label and label.lower() == text.lower():
            label_matches.append((code, item))
    if len(label_matches) == 1:
        return label_matches[0][1]
    if label_matches:
        prefix = _special_device_prefix(sheet_name, legislation)
        for code, item in label_matches:
            if code.startswith(prefix):
                return item
    return ""


def _special_device_values(sheet_name: str, legislation: str = "") -> list[str]:
    value = str(legislation or "").strip().upper()
    if sheet_name == "IVDR_IVDD" or value in {"IVDR", "IVDD"}:
        return ENUM_SOURCES.get("special_device_ivdr", [])
    return ENUM_SOURCES.get("special_device_mdr", [])


def _special_device_prefix(sheet_name: str, legislation: str = "") -> str:
    value = str(legislation or "").strip().upper()
    if value in {"MDR", "MDD", "AIMDD", "IVDR", "IVDD"}:
        return value
    return "IVDR" if sheet_name == "IVDR_IVDD" else "MDR"


def _normal_substance_type(value) -> str:
    text = str(_enum_code(value) or "").strip()
    if not text:
        return ""
    if text in ENUM_SOURCES.get("substance_type", []):
        return text
    aliases = {
        "CMR_1A": "CMR 1A",
        "CMR 1A": "CMR 1A",
        "CMR_1B": "CMR 1B",
        "CMR 1B": "CMR 1B",
        "ENDOCRINE_DISRUPTING_SUBSTANCE": "Endocrine Disrupting",
        "ENDOCRINE DISRUPTING": "Endocrine Disrupting",
        "MEDICINAL_PRODUCT_SUBSTANCE": "Medicinal Product Substance",
        "MEDICINAL PRODUCT SUBSTANCE": "Medicinal Product Substance",
        "HUMAN_PRODUCT_SUBSTANCE": "Human Blood or Plasma Substance",
        "HUMAN_BLOOD_OR_PLASMA_SUBSTANCE": "Human Blood or Plasma Substance",
        "HUMAN PRODUCT SUBSTANCE": "Human Blood or Plasma Substance",
        "HUMAN BLOOD OR PLASMA SUBSTANCE": "Human Blood or Plasma Substance",
    }
    return aliases.get(text.upper().replace("-", " ").replace("/", " "), "")


def _enum_code(value):
    if isinstance(value, str) and " - " in value:
        return value.split(" - ", 1)[0].strip()
    return value


def _split_enum_label(value: str) -> tuple[str, str]:
    text = str(value or "")
    if " - " in text:
        return tuple(text.split(" - ", 1))  # type: ignore[return-value]
    return text, ""


def _truthy(value) -> bool:
    return str(value or "").strip().upper() in {"TRUE", "YES", "1", "Y"}


def _get_by_alias(row: dict, field: str):
    key = _canonical(field)
    for header, value in row.items():
        if _canonical(header) == key:
            return value
    return None


def _record_unmapped(report: dict, sheet_name: str, source_headers: list[str], header_map: dict):
    mapped = set(header_map)
    unmapped = [header for header in source_headers if header and header not in mapped and not _is_legacy_eifu_header(header)]
    if unmapped:
        report["unmapped_headers"][sheet_name].extend(unmapped)


def _is_legacy_eifu_header(header: str) -> bool:
    return str(header or "").strip() in LEGACY_EIFU_HEADERS


def _has_data(row: dict) -> bool:
    return any(key != SOURCE_ROW_NUMBER and value not in (None, "") for key, value in row.items())


def _canonical(value: str) -> str:
    text = str(value or "").replace("*", "")
    text = re.sub(r"^(basic|udi|local)\s*-\s*", "", text, flags=re.IGNORECASE)
    text = text.replace("UDI DI", "UDI-DI")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _add_report_sheet(wb, report: dict):
    ws = wb.create_sheet("Migration Report", 0)
    ws.append(["Item", "Value"])
    ws.append(["Source file", report["source_file"]])
    ws.append(["Detected mode", report["mode"]])
    ws.append(["Source sheets", ", ".join(report["source_sheets"])])
    ws.append([])
    ws.append(["Copied sheet", "Rows"])
    for sheet, count in sorted(report["copied_rows"].items()):
        ws.append([sheet, count])
    ws.append([])
    ws.append(["Unmapped source sheet", "Headers"])
    for sheet, headers in sorted(report["unmapped_headers"].items()):
        ws.append([sheet, ", ".join(headers)])
    ws.append([])
    ws.append(["Legacy eIFU URL migrations"])
    ws.append(["Sheet", "Source Row", "UDI-DI Code", "Old eIFU URL", "Current Additional Information URL", "Result"])
    for item in report["legacy_eifu_migrations"]:
        ws.append([
            item.get("sheet", ""),
            item.get("source_row", ""),
            item.get("udi_code", ""),
            item.get("old_url", ""),
            item.get("current_url", ""),
            item.get("result", ""),
        ])
    ws.append([])
    ws.append(["Warnings"])
    for warning in report["warnings"]:
        ws.append([warning])
    for col in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 48


def _serializable_report(report: dict) -> dict:
    return {
        "source_file": report["source_file"],
        "source_sheets": report["source_sheets"],
        "mode": report["mode"],
        "copied_rows": dict(report["copied_rows"]),
        "unmapped_headers": {key: list(value) for key, value in report["unmapped_headers"].items()},
        "warnings": list(report["warnings"]),
        "legacy_eifu_migrations": list(report["legacy_eifu_migrations"]),
    }
