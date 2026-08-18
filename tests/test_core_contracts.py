import unittest
import importlib.util
import re
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from local_beta import constants, template_schema
from local_beta import build_unified_template, template_migrator, views
from local_beta.exporter import BULK_UPLOAD_ENTITY_LIMIT, BetaXMLExporter
from local_beta.importer import WorkbookImporter


def _record(record_id, basic_code="BUDI-1", udi_code=None):
    return {
        "id": record_id,
        "basic_code": basic_code,
        "udi_code": udi_code or f"UDI-{record_id}",
    }


class VersionContractTests(unittest.TestCase):
    def test_template_version_constants_stay_aligned(self):
        self.assertEqual(constants.TEMPLATE_VERSION, template_schema.TEMPLATE_VERSION)
        self.assertEqual(constants.TEMPLATE_FILENAME, f"EUDAMED_Template_{constants.TEMPLATE_VERSION}.xlsx")
        self.assertEqual(constants.TEMPLATE_EN_FILENAME, f"EUDAMED_Template_{constants.TEMPLATE_VERSION}_EN.xlsx")
        self.assertEqual(constants.SCHEMA_VERSION, "3.0.30")

    def test_release_template_assets_exist(self):
        root = Path(__file__).resolve().parents[1]
        expected = [
            root / constants.TEMPLATE_FILENAME,
            root / constants.TEMPLATE_EN_FILENAME,
            root / "EUDAMED_TOOL_v2" / "templates" / constants.TEMPLATE_FILENAME,
            root / "EUDAMED_TOOL_v2" / "templates" / constants.TEMPLATE_EN_FILENAME,
        ]
        self.assertEqual([str(path) for path in expected if not path.is_file()], [])

    def test_v212_legacy_columns_and_related_local_keys(self):
        by_field = {item["field"]: item for item in template_schema.MAIN_COLUMNS}
        for field in (
            "Legacy Has Assigned UDI-DI",
            "Legacy EUDAMED DI Input",
            "Legacy EUDAMED DI",
            "Legacy EUDAMED ID",
        ):
            self.assertIn(field, by_field)
        self.assertFalse(by_field["Legacy EUDAMED DI"]["editable"])
        self.assertFalse(by_field["Legacy EUDAMED ID"]["editable"])
        for sheet_name, spec in template_schema.RELATED_SHEETS.items():
            with self.subTest(sheet=sheet_name):
                self.assertEqual(spec["columns"][0]["field"], "Local Record ID")

    def test_v212_uses_four_regulation_specific_main_sheets(self):
        self.assertEqual(list(template_schema.ENTRY_SHEETS), ["MDR", "MDD_AIMDD", "IVDR", "IVDD"])
        self.assertNotIn("Legacy Has Assigned UDI-DI", {
            item["field"] for item in template_schema.columns_for_entry_sheet("MDR")
        })
        self.assertNotIn("Legacy Has Assigned UDI-DI", {
            item["field"] for item in template_schema.columns_for_entry_sheet("IVDR")
        })
        self.assertIn("Legacy Has Assigned UDI-DI", {
            item["field"] for item in template_schema.columns_for_entry_sheet("MDD_AIMDD")
        })
        self.assertIn("Legacy Has Assigned UDI-DI", {
            item["field"] for item in template_schema.columns_for_entry_sheet("IVDD")
        })
        for sheet_name, expected in {
            "MDR": ["MDR"],
            "MDD_AIMDD": ["MDD", "AIMDD"],
            "IVDR": ["IVDR"],
            "IVDD": ["IVDD"],
        }.items():
            column = next(
                item for item in template_schema.columns_for_entry_sheet(sheet_name)
                if item["field"] == "Applicable Legislation"
            )
            self.assertEqual(template_schema.ENUM_SOURCES[column["validation"]], expected)
        for sheet_name in ("MDD_AIMDD", "IVDD"):
            by_field = {item["field"]: item for item in template_schema.columns_for_entry_sheet(sheet_name)}
            self.assertEqual(by_field["Legacy Has Assigned UDI-DI"]["example"], "TRUE")
            self.assertEqual(by_field["Legacy EUDAMED DI Input"]["example"], "")

    def test_v212_generated_identifier_columns_are_locked(self):
        workbook = Workbook()
        worksheet = workbook.active
        columns = [
            next(item for item in template_schema.MAIN_COLUMNS if item["field"] == field)
            for field in (
                "Legacy EUDAMED DI Input",
                "Legacy EUDAMED DI",
                "Legacy EUDAMED ID",
            )
        ]
        build_unified_template._build_table_sheet(worksheet, columns, max_data_rows=5, locale="en")
        self.assertFalse(worksheet.cell(4, 1).protection.locked)
        self.assertTrue(worksheet.cell(4, 2).protection.locked)
        self.assertTrue(worksheet.cell(4, 3).protection.locked)

    def test_v212_release_workbooks_are_synchronized_and_structurally_valid(self):
        root = Path(__file__).resolve().parents[1]
        zh_path = root / constants.TEMPLATE_FILENAME
        en_path = root / constants.TEMPLATE_EN_FILENAME
        self.assertEqual(zh_path.read_bytes(), (root / "EUDAMED_TOOL_v2" / "templates" / constants.TEMPLATE_FILENAME).read_bytes())
        self.assertEqual(en_path.read_bytes(), (root / "EUDAMED_TOOL_v2" / "templates" / constants.TEMPLATE_EN_FILENAME).read_bytes())
        zh = load_workbook(zh_path, read_only=False, data_only=False)
        en = load_workbook(en_path, read_only=False, data_only=False)
        self.assertEqual(zh.sheetnames, en.sheetnames)
        for sheet_name in list(template_schema.ENTRY_SHEETS) + list(template_schema.RELATED_SHEETS):
            with self.subTest(sheet=sheet_name):
                zh_headers = [cell.value for cell in zh[sheet_name][1] if cell.value]
                en_headers = [cell.value for cell in en[sheet_name][1] if cell.value]
                self.assertEqual(zh_headers, en_headers)
        main = zh["MDD_AIMDD"]
        headers = {cell.value: cell.column for cell in main[1]}
        self.assertTrue(main.cell(4, headers["Legacy - EUDAMED DI"]).protection.locked)
        self.assertTrue(main.cell(4, headers["Legacy - EUDAMED ID"]).protection.locked)
        self.assertFalse(main.cell(4, headers["Legacy - EUDAMED DI Input"]).protection.locked)
        self.assertTrue(any("C4:C3000" in str(validation.sqref) for validation in main.data_validations.dataValidation))
        for sheet_name in template_schema.RELATED_SHEETS:
            self.assertEqual(zh[sheet_name].cell(1, 1).value, "Local - Record ID")


class EnglishTemplateCopyTests(unittest.TestCase):
    def test_schema_fields_have_english_copy_without_cjk(self):
        cjk = re.compile(r"[\u3400-\u9fff]")
        missing = []
        leaked = []
        for item in template_schema.ALL_COLUMNS:
            for key in ("description_en", "format_en"):
                value = str(item.get(key) or "")
                if not value:
                    missing.append(f"{item['header']}:{key}")
                if cjk.search(value):
                    leaked.append(f"{item['header']}:{key}={value}")

        self.assertEqual(missing, [])
        self.assertEqual(leaked, [])

    def test_context_sensitive_english_copy_matches_field_semantics(self):
        by_header = {item["header"]: item for item in template_schema.MAIN_COLUMNS}
        medicinal = by_header["Basic - Presence of Medicinal Substance"]
        basic_description = by_header["Basic - Additional Description"]

        self.assertIn("not exported separately", medicinal["description_en"])
        self.assertIn("Basic UDI-DI-level", basic_description["description_en"])

        trade_name_columns = template_schema.RELATED_SHEETS["Trade Names"]["columns"]
        trade_name_copy = {item["field"]: item["description_en"] for item in trade_name_columns}
        self.assertIn("Trade Names row", trade_name_copy["Trade Name"])
        self.assertIn("Choose ANY", trade_name_copy["Language"])

    def test_english_how_to_use_matches_chinese_scope(self):
        zh = build_unified_template._how_to_use_lines_zh()
        en = build_unified_template._how_to_use_lines_en()

        self.assertEqual(len(en), len(zh))
        english_text = "\n".join(text for text, _ in en)
        for phrase in (
            "Update container package service",
            "Clinical Sizes",
            "Greece is EL, not GR",
            "IFA",
            "WPS / Excel compatibility",
            "Presence of Medicinal Substance",
        ):
            with self.subTest(phrase=phrase):
                self.assertIn(phrase, english_text)

    def test_dropdown_validation_displays_guidance_and_blocks_invalid_values(self):
        workbook = Workbook()
        worksheet = workbook.active
        boolean_column = next(item for item in template_schema.MAIN_COLUMNS if item["validation"] == "boolean")

        build_unified_template._add_data_validations(worksheet, [boolean_column], 20, "en")

        validation = list(worksheet.data_validations.dataValidation)[0]
        self.assertTrue(validation.showInputMessage)
        self.assertTrue(validation.showErrorMessage)
        self.assertEqual(validation.errorStyle, "stop")
        self.assertEqual(validation.promptTitle, "Field guidance")

    def test_english_web_fields_show_english_guidance_and_version_label(self):
        views.set_lang("en")
        try:
            field_html = views.field_input("Presence of Medicinal Substance", "")
            self.assertIn("field-hint", field_html)
            self.assertIn("not exported separately", field_html)
            self.assertEqual(views.version_label(), f"{constants.TOOL_VERSION} · Public Beta")
        finally:
            views.set_lang("zh")

    def test_migration_page_shows_legacy_eifu_migration_details(self):
        result = {
            "output_filename": "migrated.xlsx",
            "report": {
                "mode": "current_unified_template",
                "copied_rows": {"MDR_MDD": 1},
                "unmapped_headers": {},
                "legacy_eifu_migrations": [
                    {
                        "sheet": "MDR_MDD",
                        "source_row": 4,
                        "udi_code": "UDI-EIFU",
                        "old_url": "https://example.com/old-eifu",
                        "current_url": "",
                        "result": "copied_to_additional_information_url",
                    },
                    {
                        "sheet": "MDR_MDD",
                        "source_row": 5,
                        "udi_code": "UDI-CONFLICT",
                        "old_url": "https://example.com/old",
                        "current_url": "https://example.com/current",
                        "result": "kept_existing_additional_information_url",
                    },
                ],
            },
        }

        html = views.migrate_template_page("迁移完成", result, "warning")

        self.assertIn("旧 eIFU URL 迁移明细", html)
        self.assertIn("复制 1", html)
        self.assertIn("冲突未覆盖 1", html)
        self.assertIn("UDI-EIFU", html)
        self.assertIn("https://example.com/old-eifu", html)
        self.assertIn("冲突：保留现有新字段，未覆盖", html)


class ImportHeaderCompatibilityTests(unittest.TestCase):
    def test_old_starred_pi_headers_map_to_current_fields(self):
        importer = WorkbookImporter(repository=None)
        mapping = importer._schema_by_header(template_schema.columns_for_entry_sheet("IVDR_IVDD"))

        old_lot = "UDI - PI Lot/Batch Number*"
        old_expiration = "UDI - PI Expiration Date*"
        lot_item = mapping.get(old_lot) or mapping.get(importer._canonical_header(old_lot))
        expiration_item = mapping.get(old_expiration) or mapping.get(importer._canonical_header(old_expiration))

        self.assertEqual(lot_item["field"], "PI Lot/Batch Number")
        self.assertEqual(expiration_item["field"], "PI Expiration Date")

    def test_old_eifu_header_does_not_map_to_output_field(self):
        importer = WorkbookImporter(repository=None)
        mapping = importer._schema_by_header(template_schema.columns_for_entry_sheet("MDR_MDD"))

        old_item = mapping.get("UDI - eIFU URL") or mapping.get(importer._canonical_header("UDI - eIFU URL"))

        self.assertIsNone(old_item)


class TemplateMigrationTests(unittest.TestCase):
    def _source_workbook(self, headers: list[str], values: list):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "MDR_MDD"
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(1, col_idx).value = header
            worksheet.cell(4, col_idx).value = values[col_idx - 1]
        return workbook

    def test_migration_copies_legacy_eifu_url_to_additional_information_url(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            source_path = tmp / "old_template.xlsx"
            workbook = self._source_workbook(
                ["Basic - Basic UDI-DI Code*", "UDI - UDI-DI Code*", "UDI - eIFU URL"],
                ["BUDI-EIFU", "UDI-EIFU", "https://example.com/old-eifu"],
            )
            workbook.save(source_path)

            result = template_migrator.migrate_workbook(source_path, output_dir=tmp)
            migrated = load_workbook(tmp / result["output_filename"], data_only=True)
            worksheet = migrated["MDR"]
            headers = [cell.value for cell in worksheet[1]]
            target_col = headers.index("UDI - Additional Information URL / eIFU webpage") + 1

            self.assertEqual(worksheet.cell(4, target_col).value, "https://example.com/old-eifu")
            self.assertEqual(result["report"]["legacy_eifu_migrations"][0]["source_row"], 4)
            self.assertEqual(result["report"]["legacy_eifu_migrations"][0]["udi_code"], "UDI-EIFU")
            self.assertEqual(result["report"]["legacy_eifu_migrations"][0]["result"], "copied_to_additional_information_url")
            report_values = [cell.value for row in migrated["Migration Report"].iter_rows() for cell in row]
            self.assertIn("https://example.com/old-eifu", report_values)

    def test_migration_does_not_overwrite_current_additional_information_url(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            source_path = tmp / "mixed_template.xlsx"
            workbook = self._source_workbook(
                [
                    "Basic - Basic UDI-DI Code*",
                    "UDI - UDI-DI Code*",
                    "UDI - Additional Information URL / eIFU webpage",
                    "UDI - eIFU URL",
                ],
                ["BUDI-EIFU", "UDI-EIFU", "https://example.com/current", "https://example.com/old-eifu"],
            )
            workbook.save(source_path)

            result = template_migrator.migrate_workbook(source_path, output_dir=tmp)
            migrated = load_workbook(tmp / result["output_filename"], data_only=True)
            worksheet = migrated["MDR"]
            headers = [cell.value for cell in worksheet[1]]
            target_col = headers.index("UDI - Additional Information URL / eIFU webpage") + 1

            self.assertEqual(worksheet.cell(4, target_col).value, "https://example.com/current")
            self.assertEqual(result["report"]["legacy_eifu_migrations"][0]["result"], "kept_existing_additional_information_url")

    def test_mixed_main_sheets_are_split_by_legislation(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            source_path = tmp / "mixed-main-sheets.xlsx"
            workbook = Workbook()
            mdr_mdd = workbook.active
            mdr_mdd.title = "MDR_MDD"
            ivdr_ivdd = workbook.create_sheet("IVDR_IVDD")
            headers = ["Basic - Applicable Legislation*", "Basic - Device Name*"]
            for worksheet in (mdr_mdd, ivdr_ivdd):
                for column, header in enumerate(headers, start=1):
                    worksheet.cell(1, column, header)
            for row, values in enumerate((("MDR", "MDR row"), ("MDD", "MDD row"), ("AIMDD", "AIMDD row")), start=4):
                for column, value in enumerate(values, start=1):
                    mdr_mdd.cell(row, column, value)
            for row, values in enumerate((("IVDR", "IVDR row"), ("IVDD", "IVDD row")), start=4):
                for column, value in enumerate(values, start=1):
                    ivdr_ivdd.cell(row, column, value)
            workbook.save(source_path)

            result = template_migrator.migrate_workbook(source_path, output_dir=tmp)
            migrated = load_workbook(tmp / result["output_filename"], data_only=True)

            expected = {
                "MDR": ["MDR row"],
                "MDD_AIMDD": ["MDD row", "AIMDD row"],
                "IVDR": ["IVDR row"],
                "IVDD": ["IVDD row"],
            }
            for sheet_name, names in expected.items():
                with self.subTest(sheet=sheet_name):
                    sheet = migrated[sheet_name]
                    header_map = {cell.value: cell.column for cell in sheet[1]}
                    actual = [
                        sheet.cell(row, header_map["Basic - Device Name*"]).value
                        for row in range(4, 4 + len(names))
                    ]
                    self.assertEqual(actual, names)


class ExportBatchPlanningTests(unittest.TestCase):
    def setUp(self):
        self.exporter = BetaXMLExporter(repository=None)

    def test_simple_services_split_at_official_bulk_limit(self):
        records = [_record(i) for i in range(1, BULK_UPLOAD_ENTITY_LIMIT + 2)]

        batches = self.exporter.plan_export_batches("UDI_DI.POST", records)

        self.assertEqual([batch["record_count"] for batch in batches], [300, 1])
        self.assertEqual({batch["service_type"] for batch in batches}, {"UDI_DI.POST"})

    def test_device_post_packs_multiple_basics_into_one_file_when_possible(self):
        records = [_record(i, basic_code=f"BUDI-{i}") for i in range(1, 51)]

        batches = self.exporter.plan_export_batches("DEVICE.POST", records)

        self.assertEqual(len(batches), 1)
        self.assertEqual(batches[0]["service_type"], "DEVICE.POST")
        self.assertEqual(batches[0]["record_count"], 50)
        self.assertEqual(len(batches[0]["basic_codes"]), 50)

    def test_device_post_creates_basic_once_and_posts_remaining_udis_later(self):
        records = [_record(i, basic_code="BUDI-SAME") for i in range(1, 4)]

        batches = self.exporter.plan_export_batches("DEVICE.POST", records)

        self.assertEqual([batch["service_type"] for batch in batches], ["DEVICE.POST", "UDI_DI.POST"])
        self.assertEqual([batch["record_count"] for batch in batches], [1, 2])
        self.assertEqual(batches[1]["depends_on"], batches[0]["sequence"])

    def test_legacy_assigned_udis_do_not_share_template_basic_group(self):
        records = []
        for record_id, udi_code in enumerate(["06947145553906", "05012345678903"], start=1):
            records.append(
                {
                    "id": record_id,
                    "basic_code": "WRONG-SHARED-BASIC",
                    "udi_code": udi_code,
                    "basic_payload": {
                        "Applicable Legislation": "MDD",
                        "Basic UDI-DI Code": "WRONG-SHARED-BASIC",
                        "Issuing Entity": "GS1",
                    },
                    "payload": {
                        "UDI-DI Code": udi_code,
                        "UDI-DI Issuing Entity": "GS1",
                    },
                }
            )

        batches = self.exporter.plan_export_batches("DEVICE.POST", records)

        self.assertEqual([batch["service_type"] for batch in batches], ["DEVICE.POST"])
        self.assertEqual(batches[0]["record_count"], 2)
        self.assertEqual(
            set(batches[0]["basic_codes"]),
            {"B-06947145553906", "B-05012345678903"},
        )


class LegacyEudamedDIContractTests(unittest.TestCase):
    def setUp(self):
        self.exporter = BetaXMLExporter(repository=None)
        self.basic_payload = {
            "Applicable Legislation": "MDD",
            "Basic UDI-DI Code": "WRONG-BASIC",
            "Issuing Entity": "GS1",
        }

    def test_assigned_legacy_udi_derives_eudamed_di(self):
        code, issuing_entity = self.exporter._legacy_basic_identifier(
            self.basic_payload,
            "MDEU",
            item={
                "payload": {
                    "UDI-DI Code": "06947145553906",
                    "UDI-DI Issuing Entity": "GS1",
                }
            },
        )

        self.assertEqual(code, "B-06947145553906")
        self.assertEqual(issuing_entity, "EUDAMED")

    def test_no_assigned_udi_keeps_user_eudamed_identifier(self):
        legacy_eudamed_payload = {
            "Applicable Legislation": "IVDD",
            "Basic UDI-DI Code": "B-USER-PROVIDED",
            "Issuing Entity": "EUDAMED",
        }
        for udi_code, issuing_entity in [("", "GS1"), ("D-USER-PROVIDED", "EUDAMED")]:
            with self.subTest(udi_code=udi_code, issuing_entity=issuing_entity):
                code, entity = self.exporter._legacy_basic_identifier(
                    legacy_eudamed_payload,
                    "IVDEU",
                    item={
                        "payload": {
                            "UDI-DI Code": udi_code,
                            "UDI-DI Issuing Entity": issuing_entity,
                        }
                    },
                )
                self.assertEqual(code, "B-USER-PROVIDED")
                self.assertEqual(entity, "EUDAMED")

    def test_regulation_device_keeps_user_basic_identifier(self):
        code, issuing_entity = self.exporter._legacy_basic_identifier(
            self.basic_payload,
            "MDR",
            item={
                "payload": {
                    "UDI-DI Code": "06947145553906",
                    "UDI-DI Issuing Entity": "GS1",
                }
            },
        )

        self.assertEqual(code, "WRONG-BASIC")
        self.assertEqual(issuing_entity, "GS1")

    def test_export_precheck_warns_when_template_basic_is_ignored(self):
        warnings = []

        self.exporter._warn_legacy_eudamed_di_derivation(
            warnings,
            {
                "basic_payload": self.basic_payload,
                "payload": {
                    "UDI-DI Code": "06947145553906",
                    "UDI-DI Issuing Entity": "GS1",
                },
            },
        )

        self.assertEqual(len(warnings), 1)
        self.assertIn("B-06947145553906", warnings[0])
        self.assertIn("EUDAMED", warnings[0])

        warnings = []
        self.exporter._warn_legacy_eudamed_di_derivation(
            warnings,
            {
                "basic_payload": dict(
                    self.basic_payload,
                    **{
                        "Basic UDI-DI Code": "B-06947145553906",
                        "Issuing Entity": "EUDAMED",
                    },
                ),
                "payload": {
                    "UDI-DI Code": "06947145553906",
                    "UDI-DI Issuing Entity": "GS1",
                },
            },
        )
        self.assertEqual(warnings, [])

    def test_import_warning_is_structured_and_regulation_is_unchanged(self):
        importer = WorkbookImporter(repository=None)
        warnings = []
        importer._warn_legacy_eudamed_di_derivation(
            self.basic_payload,
            {
                "UDI-DI Code": "06947145553906",
                "UDI-DI Issuing Entity": "GS1",
            },
            "MDR_MDD",
            4,
            warnings,
        )

        self.assertEqual(len(warnings), 1)
        self.assertEqual(warnings[0]["warning_type"], "LEGACY_EUDAMED_DI_DERIVED")
        self.assertEqual(warnings[0]["row"], 4)
        self.assertIn("B-06947145553906", warnings[0]["message"])

        warnings = []
        importer._warn_legacy_eudamed_di_derivation(
            dict(self.basic_payload, **{"Applicable Legislation": "MDR"}),
            {
                "UDI-DI Code": "06947145553906",
                "UDI-DI Issuing Entity": "GS1",
            },
            "MDR_MDD",
            5,
            warnings,
        )
        self.assertEqual(warnings, [])

    def test_import_entry_sheet_surfaces_legacy_derivation_warning(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "MDR_MDD"
        headers = [
            "Basic - Basic UDI-DI Code*",
            "Basic - Issuing Entity*",
            "Basic - Applicable Legislation*",
            "UDI - UDI-DI Code*",
            "UDI - UDI-DI Issuing Entity*",
        ]
        values = ["WRONG-BASIC", "GS1", "MDD", "06947145553906", "GS1"]
        for column, (header, value) in enumerate(zip(headers, values), start=1):
            worksheet.cell(1, column, header)
            worksheet.cell(4, column, value)

        parsed = {"Basic UDI-DI": [], "UDI-DI": []}
        migration_warnings = []
        WorkbookImporter(repository=None)._parse_entry_sheet(
            worksheet,
            parsed,
            {},
            {"basic_versions": {}, "udi_versions": {}},
            [],
            migration_warnings,
        )

        warning = next(
            item
            for item in migration_warnings
            if item.get("warning_type") == "LEGACY_EUDAMED_DI_DERIVED"
        )
        self.assertEqual(warning["row"], 4)
        self.assertIn("B-06947145553906", warning["message"])


class NumberOfReusesTests(unittest.TestCase):
    def setUp(self):
        self.exporter = BetaXMLExporter(repository=None)

    def test_single_use_always_outputs_zero(self):
        value = self.exporter._number_of_reuses({
            "Single Use Device": "TRUE",
            "Max Number of Reuses": "12",
        })

        self.assertEqual(value, "0")

    def test_reusable_without_declared_limit_outputs_minus_one(self):
        value = self.exporter._number_of_reuses({
            "Single Use Device": "FALSE",
            "Max Number of Reuses": "",
        })

        self.assertEqual(value, "-1")

    def test_reusable_not_applicable_aliases_normalize_to_minus_one(self):
        for raw in ["-", "N/A", "不适用", "not defined"]:
            with self.subTest(raw=raw):
                self.assertEqual(self.exporter._normalised_number_of_reuses(raw), "-1")


class WebsiteUrlMappingTests(unittest.TestCase):
    def setUp(self):
        self.exporter = BetaXMLExporter(repository=None)

    def test_additional_information_url_exports_when_public_website_blank(self):
        value = self.exporter._website_url({
            "Additional Information URL": "https://example.com/eifu",
            "Public Website": "",
        })

        self.assertEqual(value, "https://example.com/eifu")

    def test_public_website_wins_when_both_urls_are_filled(self):
        value = self.exporter._website_url({
            "Additional Information URL": "https://example.com/eifu",
            "Public Website": "https://example.com/product",
        })

        self.assertEqual(value, "https://example.com/product")

    def test_legacy_eifu_url_does_not_export_silently(self):
        value = self.exporter._website_url({
            "eIFU URL": "https://example.com/legacy-eifu",
            "Public Website": "",
            "Additional Information URL": "",
        })

        self.assertEqual(value, "")

    def test_legacy_eifu_url_warns_before_export(self):
        warnings = []
        self.exporter._warn_website_url_selection(warnings, {
            "udi_code": "UDI-LEGACY-EIFU",
            "payload": {"eIFU URL": "https://example.com/legacy-eifu"},
        })

        self.assertEqual(len(warnings), 1)
        self.assertIn("不会自动把它写入 EUDAMED", warnings[0])


class LegacyValidatorApplicabilityTests(unittest.TestCase):
    @staticmethod
    def _validator_module():
        validator_path = Path(__file__).resolve().parents[1] / "EUDAMED_TOOL_v2" / "validator.py"
        spec = importlib.util.spec_from_file_location("legacy_validator", validator_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module

    def test_ivdd_does_not_require_pi_lot_or_expiration(self):
        module = self._validator_module()

        data = {
            "Basic UDI-DI": [{
                "Basic UDI-DI Code": "BIVDD0001",
                "Issuing Entity": "EUDAMED",
                "Manufacturer SRN": "DE-MF-000000001",
                "Risk Class": "IVD General",
                "Applicable Legislation": "IVDD",
                "Device Type": "Regular Device",
                "Device Name/Model": "Legacy IVDD Device",
                "EMDN Code": "W0101",
            }],
            "UDI-DI": [{
                "Parent Basic UDI-DI": "BIVDD0001",
                "UDI-DI Code": "UIVDD0001",
                "UDI-DI Issuing Entity": "EUDAMED",
                "Device Status": "On the EU market",
                "Single Use Device": "FALSE",
                "Device Labelled as Sterile": "FALSE",
                "Trade Name Applicable": "FALSE",
                "Nomenclature Code": "W0101",
            }],
        }

        errors, _ = module.DataValidator(data).validate_all()
        pi_errors = [
            error for error in errors
            if error.field in {"PI Lot/Batch Number", "PI Expiration Date"}
        ]
        self.assertEqual(pi_errors, [])

    def test_eudamed_b_and_d_identifiers_are_not_rejected_as_format_errors(self):
        module = self._validator_module()
        data = {
            "Basic UDI-DI": [{
                "Basic UDI-DI Code": "B-AF-MF-000000245AT",
                "Issuing Entity": "EUDAMED",
                "Manufacturer SRN": "DE-MF-000000001",
                "Risk Class": "IVD General",
                "Applicable Legislation": "IVDD",
                "Device Type": "Regular Device",
                "Device Name/Model": "Legacy IVDD Device",
                "EMDN Code": "W0101",
            }],
            "UDI-DI": [{
                "Parent Basic UDI-DI": "B-AF-MF-000000245AT",
                "UDI-DI Code": "D-AF-MF-000000245AT",
                "UDI-DI Issuing Entity": "EUDAMED",
                "Device Status": "Not intended for the EU market",
                "Single Use Device": "FALSE",
                "Device Labelled as Sterile": "FALSE",
                "Trade Name Applicable": "FALSE",
                "Nomenclature Code": "W0101",
            }],
        }

        errors, _ = module.DataValidator(data).validate_all()
        identifier_format_errors = [
            error for error in errors
            if error.error_type == "FORMAT_ERROR"
            and error.field in {"Basic UDI-DI Code", "UDI-DI Code"}
        ]

        self.assertEqual(identifier_format_errors, [])

    def test_di_code_format_uses_official_xsd_length_without_alphanumeric_regex(self):
        validator = self._validator_module().DataValidator({})

        self.assertTrue(validator._is_valid_udi_code("B-1234_1237_1"))
        self.assertTrue(validator._is_valid_udi_code("X" * 120))
        self.assertFalse(validator._is_valid_udi_code(""))
        self.assertFalse(validator._is_valid_udi_code("X" * 121))


class DataDictionaryAuditTests(unittest.TestCase):
    def test_generator_preserves_current_field_mappings_and_notes(self):
        root = Path(__file__).resolve().parents[1]
        audit_path = root / "scripts" / "audit_data_dictionary_mapping.py"
        spec = importlib.util.spec_from_file_location("data_dictionary_audit", audit_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)

        rows = module.read_dictionary(module.DATA_DICTIONARY_PATH)
        template_index = module.build_template_index()
        storage_fields = set(constants.BASIC_FIELDS) | set(constants.UDI_FIELDS)
        exporter_text = (root / "local_beta" / "exporter.py").read_text(encoding="utf-8")
        audited = [
            module.audit_row(row, template_index, storage_fields, exporter_text)
            for row in rows
        ]
        by_key = {(row["sheet"], row["field_id"]): row for row in audited}

        for sheet in ("DD UDI-DI", "DD Legacy Devices", "DD UDI-DI_SPP"):
            with self.subTest(sheet=sheet):
                mapping = by_key[(sheet, "FLD-UDID-174")]
                self.assertEqual(mapping["status"], "implemented")
                self.assertEqual(mapping["xml_path"], "udidi:website")
                self.assertIn("Additional Information URL", mapping["template"])

        self.assertIn("baseQuantity", by_key[("DD UDI-DI", "FLD-UDID-151")]["notes"])
        self.assertIn("MDR/MDD/AIMDD", by_key[("DD UDI-DI", "FLD-UDID-156")]["notes"])
        self.assertIn("MDD/AIMDD legacy", by_key[("DD Legacy Devices", "FLD-UDID-156")]["notes"])
        self.assertFalse(any("v2.7 template" in row.get("notes", "") for row in audited))

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "audit.md"
            module.write_markdown(audited, output)
            report = output.read_text(encoding="utf-8")
        self.assertIn("`FLD-UDID-174`", report)
        self.assertNotIn("eIFU URL` and `Public Email` are collected", report)


if __name__ == "__main__":
    unittest.main()
