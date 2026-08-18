import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import local_beta.importer as importer_module
from local_beta import template_migrator
from local_beta.importer import WorkbookImporter
from local_beta.legacy_identifiers import (
    LegacyIdentifierError,
    calculate_check_characters,
    generate_eudamed_pair,
    resolve_legacy_identifiers,
    validate_eudamed_pair,
)
from local_beta.storage import Repository
from local_beta.template_schema import RELATED_SHEETS, TEMPLATE_VERSION, columns_for_entry_sheet


class LegacyIdentifierAlgorithmTests(unittest.TestCase):
    def test_official_examples(self):
        self.assertEqual(generate_eudamed_pair("CR0233"), ("B-CR023368", "D-CR023368"))
        self.assertEqual(
            generate_eudamed_pair("BEMF000000106CR023335"),
            ("B-BEMF000000106CR023335WE", "D-BEMF000000106CR023335WE"),
        )

    def test_case_is_preserved_and_changes_checksum(self):
        upper = generate_eudamed_pair("ABC")
        lower = generate_eudamed_pair("abc")
        self.assertTrue(upper[0].startswith("B-ABC"))
        self.assertTrue(lower[0].startswith("B-abc"))
        self.assertNotEqual(upper, lower)

    def test_invalid_length_character_checksum_and_pair(self):
        for body in ("", "A" * 22):
            with self.subTest(body=body):
                with self.assertRaises(LegacyIdentifierError):
                    calculate_check_characters(body)
        with self.assertRaises(LegacyIdentifierError):
            generate_eudamed_pair("ABC@")
        with self.assertRaises(LegacyIdentifierError):
            generate_eudamed_pair(" CR0233 ")
        with self.assertRaises(LegacyIdentifierError):
            validate_eudamed_pair("B-CR023369", "D-CR023369")
        with self.assertRaises(LegacyIdentifierError):
            validate_eudamed_pair("B-CR023368", "D-OTHER68")

    def test_all_legacy_legislations_support_three_resolution_methods(self):
        for legislation in ("MDD", "AIMDD", "IVDD"):
            with self.subTest(legislation=legislation, path="assigned"):
                assigned = resolve_legacy_identifiers(
                    {"Applicable Legislation": legislation, "Basic UDI-DI Code": "LOCAL", "Issuing Entity": "GS1"},
                    {"Legacy Has Assigned UDI-DI": "TRUE", "UDI-DI Code": "06947145553906", "UDI-DI Issuing Entity": "GS1"},
                )
                self.assertEqual((assigned.eudamed_di, assigned.eudamed_id), ("B-06947145553906", ""))
                self.assertEqual(assigned.method, "derived_from_udi")
            with self.subTest(legislation=legislation, path="generated"):
                generated = resolve_legacy_identifiers(
                    {"Applicable Legislation": legislation},
                    {"Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"},
                )
                self.assertEqual((generated.eudamed_di, generated.eudamed_id), ("B-CR023368", "D-CR023368"))
                self.assertEqual(generated.method, "generated_from_input")
            with self.subTest(legislation=legislation, path="existing"):
                existing = resolve_legacy_identifiers(
                    {"Applicable Legislation": legislation, "Basic UDI-DI Code": "B-CR023368", "Issuing Entity": "EUDAMED"},
                    {"Legacy Has Assigned UDI-DI": "FALSE", "UDI-DI Code": "D-CR023368", "UDI-DI Issuing Entity": "EUDAMED"},
                )
                self.assertEqual(existing.method, "existing_eudamed_pair")

    def test_regulation_devices_do_not_enter_legacy_resolver(self):
        for legislation in ("MDR", "IVDR"):
            self.assertIsNone(resolve_legacy_identifiers(
                {"Applicable Legislation": legislation},
                {"Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"},
            ))

    def test_system_and_procedure_pack_do_not_enter_legacy_resolver(self):
        for device_type in ("System", "Procedure Pack"):
            with self.subTest(device_type=device_type):
                self.assertIsNone(resolve_legacy_identifiers(
                    {"Applicable Legislation": "MDD", "Device Type": device_type},
                    {
                        "Legacy Has Assigned UDI-DI": "TRUE",
                        "UDI-DI Code": "06947145553906",
                        "UDI-DI Issuing Entity": "GS1",
                    },
                ))

    def test_no_udi_path_rejects_real_issuer_data_and_partial_existing_pair(self):
        with self.assertRaises(LegacyIdentifierError):
            resolve_legacy_identifiers(
                {"Applicable Legislation": "MDD"},
                {
                    "Legacy Has Assigned UDI-DI": "FALSE",
                    "Legacy EUDAMED DI Input": "CR0233",
                    "UDI-DI Code": "06947145553906",
                    "UDI-DI Issuing Entity": "GS1",
                },
            )
        with self.assertRaises(LegacyIdentifierError):
            resolve_legacy_identifiers(
                {"Applicable Legislation": "MDD"},
                {
                    "Legacy Has Assigned UDI-DI": "FALSE",
                    "Legacy EUDAMED DI Input": "CR0233",
                    "Legacy EUDAMED DI": "B-CR023368",
                },
            )

    def test_v211_compatibility_infers_assigned_or_existing_pair_and_refuses_guessing(self):
        assigned = resolve_legacy_identifiers(
            {"Applicable Legislation": "MDD", "Basic UDI-DI Code": "LOCAL", "Issuing Entity": "GS1"},
            {"UDI-DI Code": "06947145553906", "UDI-DI Issuing Entity": "GS1"},
        )
        self.assertEqual(assigned.method, "derived_from_udi")
        existing = resolve_legacy_identifiers(
            {"Applicable Legislation": "IVDD", "Basic UDI-DI Code": "B-CR023368", "Issuing Entity": "EUDAMED"},
            {"UDI-DI Code": "D-CR023368", "UDI-DI Issuing Entity": "EUDAMED"},
        )
        self.assertEqual(existing.method, "existing_eudamed_pair")
        with self.assertRaises(LegacyIdentifierError):
            resolve_legacy_identifiers(
                {"Applicable Legislation": "MDD", "Basic UDI-DI Code": "LOCAL", "Issuing Entity": "GS1"},
                {"UDI-DI Code": "", "UDI-DI Issuing Entity": ""},
            )


class LegacyIdentifierImportTests(unittest.TestCase):
    def setUp(self):
        self.tmp_context = tempfile.TemporaryDirectory()
        self.tmp = Path(self.tmp_context.name)
        self.old_export_dir = importer_module.EXPORT_DIR
        importer_module.EXPORT_DIR = self.tmp / "exports"
        self.repo = Repository(db_path=self.tmp / "test.db")
        self.importer = WorkbookImporter(self.repo)

    def tearDown(self):
        importer_module.EXPORT_DIR = self.old_export_dir
        self.tmp_context.cleanup()

    def _workbook(self, rows, related_rows=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "MDD_AIMDD"
        columns = columns_for_entry_sheet("MDD_AIMDD")
        headers = [item["header"] for item in columns]
        field_to_col = {item["field"]: index for index, item in enumerate(columns, start=1)}
        for index, header in enumerate(headers, start=1):
            ws.cell(1, index, header)
        for row_number, values in enumerate(rows, start=4):
            defaults = {
                "Manufacturer SRN": "DE-MF-000000001",
                "Risk Class": "Class IIa",
                "Applicable Legislation": "MDD",
                "Device Type": "Regular Device",
                "Device Name/Model": "Legacy test",
                "EMDN Code": "W0101",
                "Device Status": "No longer placed on the EU market",
                "Single Use Device": "FALSE",
                "Device Labelled as Sterile": "FALSE",
                "Trade Name Applicable": "FALSE",
                "Reference Number": f"REF-{row_number}",
                "Nomenclature Code": "W0101",
            }
            defaults.update(values)
            for field, value in defaults.items():
                ws.cell(row_number, field_to_col[field], value)

        for sheet_name, related in (related_rows or {}).items():
            related_ws = wb.create_sheet(sheet_name)
            columns = RELATED_SHEETS[sheet_name]["columns"]
            related_field_to_col = {item["field"]: index for index, item in enumerate(columns, start=1)}
            for index, item in enumerate(columns, start=1):
                related_ws.cell(1, index, item["header"])
            for row_number, values in enumerate(related, start=4):
                for field, value in values.items():
                    related_ws.cell(row_number, related_field_to_col[field], value)

        how_to = wb.create_sheet("How to Use")
        how_to.cell(1, 1, f"EUDAMED Template {TEMPLATE_VERSION} - How to Use")
        return wb

    def _save(self, wb, name="input.xlsx"):
        path = self.tmp / name
        wb.save(path)
        return path

    def test_no_udi_generation_related_resolution_and_normalized_copy(self):
        path = self._save(self._workbook(
            [{
                "record_id": "ROW-001",
                "Legacy Has Assigned UDI-DI": "FALSE",
                "Legacy EUDAMED DI Input": "CR0233",
            }],
            {"Trade Names": [{"Local Record ID": "ROW-001", "Trade Name": "Legacy trade", "Language": "en"}]},
        ))

        result = self.importer.import_workbook(path)

        self.assertEqual(result["summary"]["udi_count"], 1)
        self.assertTrue(result["normalized_filename"])
        udi = self.repo.list_udis(limit=10)[0]
        self.assertEqual(udi["udi_code"], "D-CR023368")
        self.assertEqual(udi["basic_code"], "B-CR023368")
        self.assertEqual(udi["payload"]["Legacy Identifier Method"], "generated_from_input")
        self.assertEqual(udi["trade_name_rows"][0]["UDI-DI Code"], "D-CR023368")

        normalized = load_workbook(self.tmp / "exports" / result["normalized_filename"], data_only=True)
        main_headers = {cell.value: cell.column for cell in normalized["MDD_AIMDD"][1]}
        self.assertEqual(normalized["MDD_AIMDD"].cell(4, main_headers["Legacy - EUDAMED DI"]).value, "B-CR023368")
        self.assertEqual(normalized["MDD_AIMDD"].cell(4, main_headers["Legacy - EUDAMED ID"]).value, "D-CR023368")
        related_headers = {cell.value: cell.column for cell in normalized["Trade Names"][1]}
        self.assertEqual(normalized["Trade Names"].cell(4, related_headers["UDI-DI Code*"]).value, "D-CR023368")

        second_repo = Repository(db_path=self.tmp / "second.db")
        second_result = WorkbookImporter(second_repo).import_workbook(self.tmp / "exports" / result["normalized_filename"])
        self.assertFalse(any(error.get("error_type") == "LEGACY_IDENTIFIER_INVALID" for error in second_result["validation"]["errors"]))
        self.assertEqual(second_repo.list_udis(limit=10)[0]["payload"]["Legacy Identifier Method"], "generated_from_input")

    def test_invalid_identifier_row_is_not_written_but_valid_row_continues(self):
        path = self._save(self._workbook([
            {"record_id": "BAD", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "A@"},
            {"record_id": "GOOD", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"},
        ]))

        result = self.importer.import_workbook(path)

        self.assertEqual(result["summary"]["udi_count"], 1)
        self.assertEqual([item["udi_code"] for item in self.repo.list_udis(limit=10)], ["D-CR023368"])
        self.assertTrue(any(error.get("error_type") == "LEGACY_IDENTIFIER_INVALID" for error in result["validation"]["errors"]))
        self.assertEqual(result["normalized_filename"], "")

    def test_excel_import_preserves_legacy_body_whitespace_for_validation(self):
        path = self._save(self._workbook([{
            "record_id": "SPACED",
            "Legacy Has Assigned UDI-DI": "FALSE",
            "Legacy EUDAMED DI Input": " CR0233 ",
        }]))

        result = self.importer.import_workbook(path)

        self.assertEqual(result["summary"]["basic_count"], 0)
        self.assertEqual(result["summary"]["udi_count"], 0)
        errors = [
            item for item in result["validation"]["errors"]
            if item.get("error_type") == "LEGACY_IDENTIFIER_INVALID"
        ]
        self.assertEqual(len(errors), 1)
        self.assertIn("官方字符表之外", errors[0]["message"])

    def test_first_generation_without_local_record_id_remains_blocked(self):
        path = self._save(self._workbook([{
            "Legacy Has Assigned UDI-DI": "FALSE",
            "Legacy EUDAMED DI Input": "CR0233",
        }]))

        result = self.importer.import_workbook(path)

        self.assertEqual(result["summary"]["basic_count"], 0)
        self.assertEqual(result["summary"]["udi_count"], 0)
        errors = [
            item for item in result["validation"]["errors"]
            if item.get("error_type") == "LEGACY_IDENTIFIER_INVALID"
        ]
        self.assertEqual(len(errors), 1)
        self.assertIn("Local - Record ID 条件必填", errors[0]["message"])

    def test_duplicate_local_ids_exclude_all_duplicate_rows(self):
        path = self._save(self._workbook([
            {"record_id": "DUP", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"},
            {"record_id": "DUP", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "ABC"},
        ]))

        result = self.importer.import_workbook(path)

        self.assertEqual(result["summary"]["udi_count"], 0)
        self.assertEqual(self.repo.list_udis(limit=10), [])
        self.assertEqual(sum(error.get("error_type") == "LOCAL_RECORD_ID_DUPLICATE" for error in result["validation"]["errors"]), 2)

    def test_duplicate_generated_identifiers_exclude_all_duplicate_rows(self):
        path = self._save(self._workbook([
            {"record_id": "ONE", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"},
            {"record_id": "TWO", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"},
        ]))

        result = self.importer.import_workbook(path)

        self.assertEqual(result["summary"]["udi_count"], 0)
        self.assertEqual(self.repo.list_udis(limit=10), [])
        self.assertEqual(sum(error.get("error_type") == "LEGACY_IDENTIFIER_DUPLICATE" for error in result["validation"]["errors"]), 2)

    def test_conflicting_local_and_formal_related_reference_is_rejected(self):
        path = self._save(self._workbook(
            [{"record_id": "ROW-001", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"}],
            {"Trade Names": [{
                "Local Record ID": "ROW-001",
                "UDI-DI Code": "DIFFERENT",
                "Trade Name": "Conflict",
                "Language": "en",
            }]},
        ))

        result = self.importer.import_workbook(path)

        self.assertTrue(any(error.get("error_type") == "LOCAL_AND_FORMAL_REFERENCE_CONFLICT" for error in result["validation"]["errors"]))
        self.assertEqual(self.repo.list_udis(limit=10)[0]["trade_name_rows"], [])

    def test_every_related_sheet_resolves_local_record_id(self):
        related = {
            "Trade Names": [{"Local Record ID": "ROW-ALL", "Trade Name": "Name", "Language": "en"}],
            "Market Info": [{"Local Record ID": "ROW-ALL", "Country Code": "IT", "Originally Placed on Market": "TRUE"}],
            "Package Info": [{"Local Record ID": "ROW-ALL", "Package UDI-DI Code": "PKG-1", "Package Issuing Entity": "EUDAMED", "Quantity per Package": "1"}],
            "Clinical Sizes": [{"Local Record ID": "ROW-ALL", "Clinical Size Type": "CST1", "Precision": "Text", "Text Value": "small"}],
            "Annex XVI Purposes": [{"Local Record ID": "ROW-ALL", "Non-Medical Device Type": "CONTACT_LENSES"}],
            "Critical Warnings": [{"Local Record ID": "ROW-ALL", "Warning Type": "CW001"}],
            "Storage Conditions": [{"Local Record ID": "ROW-ALL", "Storage Condition Type": "SHC001"}],
            "CMR Substances": [{"Local Record ID": "ROW-ALL", "Substance Type": "CMR 1A", "Substance Name": "X"}],
            "Device Certificates": [{"Local Record ID": "ROW-ALL", "Certificate Type": "MDD_III", "Notified Body ID": "1234"}],
        }
        path = self._save(self._workbook(
            [{"record_id": "ROW-ALL", "Legacy Has Assigned UDI-DI": "FALSE", "Legacy EUDAMED DI Input": "CR0233"}],
            related,
        ))

        self.importer.import_workbook(path)

        udi = self.repo.list_udis(limit=10)[0]
        basic = self.repo.get_basic_by_code("B-CR023368")
        for key in (
            "trade_name_rows", "market_rows", "package_rows", "clinical_size_rows",
            "annex_xvi_rows", "warning_rows", "storage_rows",
        ):
            with self.subTest(key=key):
                self.assertEqual(len(udi[key]), 1)
                self.assertEqual(udi[key][0]["UDI-DI Code"], "D-CR023368")
        self.assertEqual(basic["cmr_rows"][0]["Basic UDI-DI Code"], "B-CR023368")
        self.assertEqual(basic["cert_rows"][0]["Basic UDI-DI Code"], "B-CR023368")

    def test_assigned_udi_normalized_copy_preserves_local_basic_and_certificate_link(self):
        path = self._save(self._workbook(
            [{
                "record_id": "ASSIGNED-1",
                "Legacy Has Assigned UDI-DI": "TRUE",
                "Basic UDI-DI Code": "LOCAL-BASIC",
                "Issuing Entity": "GS1",
                "UDI-DI Code": "06947145553906",
                "UDI-DI Issuing Entity": "GS1",
            }],
            {"Device Certificates": [{
                "Basic UDI-DI Code": "LOCAL-BASIC",
                "Certificate Type": "MDD_III",
                "Notified Body ID": "1234",
            }]},
        ))

        result = self.importer.import_workbook(path)

        self.assertTrue(result["normalized_filename"])
        normalized_path = self.tmp / "exports" / result["normalized_filename"]
        normalized = load_workbook(normalized_path, data_only=True)
        main_headers = {cell.value: cell.column for cell in normalized["MDD_AIMDD"][1]}
        self.assertEqual(normalized["MDD_AIMDD"].cell(4, main_headers["Basic - Basic UDI-DI Code"]).value, "LOCAL-BASIC")
        self.assertEqual(normalized["MDD_AIMDD"].cell(4, main_headers["Legacy - EUDAMED DI"]).value, "B-06947145553906")
        cert_headers = {cell.value: cell.column for cell in normalized["Device Certificates"][1]}
        self.assertEqual(normalized["Device Certificates"].cell(4, cert_headers["Basic UDI-DI Code*"]).value, "LOCAL-BASIC")

        second_repo = Repository(db_path=self.tmp / "assigned-second.db")
        second_result = WorkbookImporter(second_repo).import_workbook(normalized_path)
        self.assertEqual(second_result["summary"]["basic_count"], 1)
        self.assertEqual(second_result["summary"]["udi_count"], 1)
        self.assertFalse(any("不存在于主表" in str(error.get("message")) for error in second_result["validation"]["errors"]))

    def test_sparse_v211_row_normalization_uses_migrated_target_row(self):
        wb = self._workbook([{
            "record_id": "SPARSE-1",
            "Legacy Has Assigned UDI-DI": "TRUE",
            "Basic UDI-DI Code": "LOCAL-SPARSE",
            "Issuing Entity": "GS1",
            "UDI-DI Code": "06947145553906",
            "UDI-DI Issuing Entity": "GS1",
        }])
        ws = wb["MDD_AIMDD"]
        ws.title = "MDR_MDD"
        for cell in ws[4]:
            ws.cell(7, cell.column).value = cell.value
            cell.value = None
        how_to = wb["How to Use"]
        how_to.cell(1, 1, "EUDAMED Template v2.11 - How to Use")
        source = self._save(wb, "sparse-v211.xlsx")

        result = self.importer.import_workbook(source)

        self.assertTrue(result["normalized_filename"])
        normalized_path = self.tmp / "exports" / result["normalized_filename"]
        normalized = load_workbook(normalized_path, data_only=True)
        main = normalized["MDD_AIMDD"]
        headers = {cell.value: cell.column for cell in main[1]}
        self.assertEqual(main.cell(4, headers["Basic - Applicable Legislation*"]).value, "MDD")
        self.assertEqual(main.cell(4, headers["Legacy - EUDAMED DI"]).value, "B-06947145553906")
        self.assertTrue(all(main.cell(7, column).value in (None, "") for column in range(1, main.max_column + 1)))

        second_repo = Repository(db_path=self.tmp / "sparse-second.db")
        second_result = WorkbookImporter(second_repo).import_workbook(normalized_path)
        self.assertEqual(second_result["summary"]["basic_count"], 1)
        self.assertEqual(second_result["summary"]["udi_count"], 1)

    def test_v211_migrator_sets_explicit_legacy_paths_without_guessing_body(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "MDR_MDD"
        headers = [
            "Basic - Basic UDI-DI Code*", "Basic - Issuing Entity*",
            "Basic - Applicable Legislation*", "Basic - Device Type*",
            "UDI - UDI-DI Code*", "UDI - UDI-DI Issuing Entity*",
        ]
        for column, header in enumerate(headers, start=1):
            ws.cell(1, column, header)
        rows = [
            ["LOCAL-BASIC", "GS1", "MDD", "Regular Device", "06947145553906", "GS1"],
            ["B-CR023368", "EUDAMED", "MDD", "Regular Device", "D-CR023368", "EUDAMED"],
            ["UNKNOWN", "GS1", "MDD", "Regular Device", "", ""],
        ]
        for row_number, values in enumerate(rows, start=4):
            for column, value in enumerate(values, start=1):
                ws.cell(row_number, column, value)
        how_to = wb.create_sheet("How to Use")
        how_to.cell(1, 1, "EUDAMED Template v2.11 - How to Use")
        source = self._save(wb, "v211.xlsx")

        result = template_migrator.migrate_workbook(source, self.tmp)

        self.assertTrue(result["ok"])
        migrated = load_workbook(self.tmp / result["output_filename"], data_only=True)
        headers = {cell.value: cell.column for cell in migrated["MDD_AIMDD"][1]}
        self.assertEqual(migrated["MDD_AIMDD"].cell(4, headers["Legacy - Has Assigned UDI-DI?*"]).value, "TRUE")
        self.assertEqual(migrated["MDD_AIMDD"].cell(4, headers["Legacy - EUDAMED DI"]).value, "B-06947145553906")
        self.assertEqual(migrated["MDD_AIMDD"].cell(5, headers["Legacy - Has Assigned UDI-DI?*"]).value, "FALSE")
        self.assertEqual(migrated["MDD_AIMDD"].cell(5, headers["Legacy - EUDAMED ID"]).value, "D-CR023368")
        self.assertIsNone(migrated["MDD_AIMDD"].cell(6, headers["Legacy - Has Assigned UDI-DI?*"]).value)
        self.assertTrue(any("不会猜测主体" in warning for warning in result["warnings"]))

    def test_v211_existing_eudamed_pair_migrates_and_imports_without_local_record_id(self):
        wb = self._workbook(
            [{
                "Basic UDI-DI Code": "B-CR023368",
                "Issuing Entity": "EUDAMED",
                "UDI-DI Code": "D-CR023368",
                "UDI-DI Issuing Entity": "EUDAMED",
                "Trade Name Applicable": "TRUE",
                "Containing Latex": "FALSE",
            }],
            {"Trade Names": [{
                "UDI-DI Code": "D-CR023368",
                "Trade Name": "Existing Legacy pair",
                "Language": "en",
            }]},
        )
        main = wb["MDD_AIMDD"]
        main.title = "MDR_MDD"
        removed_headers = {
            "Local - Record ID",
            "Legacy - Has Assigned UDI-DI?*",
            "Legacy - EUDAMED DI Input",
            "Legacy - EUDAMED DI",
            "Legacy - EUDAMED ID",
        }
        for column in range(main.max_column, 0, -1):
            if main.cell(1, column).value in removed_headers:
                main.delete_cols(column)
        trade_names = wb["Trade Names"]
        for column in range(trade_names.max_column, 0, -1):
            if trade_names.cell(1, column).value == "Local - Record ID":
                trade_names.delete_cols(column)
                break
        wb["How to Use"].cell(1, 1, "EUDAMED Template v2.11 - How to Use")
        source = self._save(wb, "v211-existing-pair.xlsx")

        migration = template_migrator.migrate_workbook(source, self.tmp)

        self.assertTrue(migration["ok"])
        migrated_path = self.tmp / migration["output_filename"]
        migrated = load_workbook(migrated_path, data_only=True)
        headers = {cell.value: cell.column for cell in migrated["MDD_AIMDD"][1]}
        self.assertEqual(
            migrated["MDD_AIMDD"].cell(4, headers["Legacy - Has Assigned UDI-DI?*"]).value,
            "FALSE",
        )
        self.assertEqual(migrated["MDD_AIMDD"].cell(4, headers["Legacy - EUDAMED DI"]).value, "B-CR023368")
        self.assertEqual(migrated["MDD_AIMDD"].cell(4, headers["Legacy - EUDAMED ID"]).value, "D-CR023368")
        self.assertIsNone(migrated["MDD_AIMDD"].cell(4, headers["Local - Record ID"]).value)

        import_result = self.importer.import_workbook(migrated_path)

        self.assertFalse(any(
            error.get("error_type") == "LEGACY_IDENTIFIER_INVALID"
            for error in import_result["validation"]["errors"]
        ))
        self.assertEqual(import_result["validation"]["errors"], [])
        self.assertEqual(import_result["summary"]["basic_count"], 1)
        self.assertEqual(import_result["summary"]["udi_count"], 1)
        stored = self.repo.list_udis(limit=10)[0]
        self.assertEqual(stored["basic_code"], "B-CR023368")
        self.assertEqual(stored["udi_code"], "D-CR023368")
        self.assertEqual(stored["payload"]["Legacy Identifier Method"], "existing_eudamed_pair")
        self.assertEqual(stored["trade_name_rows"][0]["Trade Name"], "Existing Legacy pair")

    def test_migrator_appends_existing_related_rows_after_inline_details(self):
        wb = Workbook()
        main = wb.active
        main.title = "MDR_MDD"
        main_headers = [
            "Basic - Applicable Legislation*",
            "Basic - Device Type*",
            "UDI - UDI-DI Code*",
            "Clinical Size Value",
            "Clinical Size Unit",
            "Purpose Other Than Medical",
        ]
        for column, header in enumerate(main_headers, start=1):
            main.cell(1, column, header)
        for column, value in enumerate(
            ["MDR", "Regular Device", "UDI-INLINE", 5, "mm", "TRUE"],
            start=1,
        ):
            main.cell(4, column, value)

        clinical = wb.create_sheet("Clinical Sizes")
        clinical_headers = [
            "UDI-DI Code*", "Clinical Size Type*", "Precision*", "Value", "Measure Unit",
        ]
        for column, header in enumerate(clinical_headers, start=1):
            clinical.cell(1, column, header)
        for column, value in enumerate(["UDI-EXISTING", "CST1", "Value", 9, "MU01"], start=1):
            clinical.cell(4, column, value)

        annex = wb.create_sheet("Annex XVI Purposes")
        annex_headers = ["UDI-DI Code*", "Non-Medical Device Type*"]
        for column, header in enumerate(annex_headers, start=1):
            annex.cell(1, column, header)
        for column, value in enumerate(["UDI-ANNEX-EXISTING", "CONTACT_LENSES"], start=1):
            annex.cell(4, column, value)

        how_to = wb.create_sheet("How to Use")
        how_to.cell(1, 1, "EUDAMED Template v2.11 - How to Use")
        source = self._save(wb, "inline-and-related-v211.xlsx")

        result = template_migrator.migrate_workbook(source, self.tmp)

        self.assertTrue(result["ok"])
        migrated = load_workbook(self.tmp / result["output_filename"], data_only=True)
        clinical_code_column = next(
            cell.column for cell in migrated["Clinical Sizes"][1] if cell.value == "UDI-DI Code*"
        )
        annex_code_column = next(
            cell.column for cell in migrated["Annex XVI Purposes"][1] if cell.value == "UDI-DI Code*"
        )
        clinical_codes = {
            migrated["Clinical Sizes"].cell(row, clinical_code_column).value
            for row in range(4, 6)
        }
        annex_codes = {
            migrated["Annex XVI Purposes"].cell(row, annex_code_column).value
            for row in range(4, 6)
        }
        self.assertEqual(clinical_codes, {"UDI-INLINE", "UDI-EXISTING"})
        self.assertEqual(annex_codes, {"UDI-INLINE", "UDI-ANNEX-EXISTING"})
        self.assertEqual(result["report"]["copied_rows"]["Clinical Sizes"], 2)
        self.assertEqual(result["report"]["copied_rows"]["Annex XVI Purposes"], 2)


if __name__ == "__main__":
    unittest.main()
